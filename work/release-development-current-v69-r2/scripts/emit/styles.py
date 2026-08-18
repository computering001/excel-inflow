"""
Plan style records -> openpyxl style objects.

THE TRAP THIS MODULE EXISTS TO CLOSE.

openpyxl has no partial-style merge. `cell.font = Font(bold=True)` does not add
bold; it replaces the font, and every other channel on it — name, size, and
above all COLOUR — reverts to default. Font colour is the entire provenance
layer of this model now that input fills were removed (blue = typed hardcode,
black = same-sheet derivation, green = cross-sheet link, white = chrome), so a
partial font assignment anywhere in a renderer silently destroys provenance, and
does it without raising anything.

The legacy workbook library writer merged partial font assignments, which is why the Node
emitter can write `{ bold: true }` on a subtotal row and keep the blue beneath
it. That behaviour is a SEMANTIC, not an API call, and it survives no port.

The answer here is not to reimplement the merge. It is that the plan carries
COMPLETE style records — every channel resolved by the emitter — so this module
performs exactly one whole-style assignment per cell and never reads a cell's
prior state. There is no partial assignment to get wrong.

The border trap is the mirror image. Partial border assignment in legacy workbook library
silently CLEARED the perimeter edges it omitted, which is why the Node emitter
has a draw-order discipline: passes are ordered so that the last writer of a
range states every edge it wants. That workaround is not translated. The plan
states the intended edge set outright — an edge present means that edge, an edge
absent means NO edge — and openpyxl's Border does precisely that, so intent and
mechanism agree without any ordering rule at all.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle

# openpyxl's PatternFill compares child colours against Color() to decide
# whether to serialise them, so leaving a colour unset is meaningful and must
# not be filled in with a plausible default.
_NO_SIDE = Side()


def _color(value):
    """
    Plan colours are ARGB or RGB hex strings, upper case.

    Theme and indexed colours are refused by the schema and refused again here:
    they resolve against a theme part, so the rendered colour would depend on
    something outside the plan. Provenance may not depend on a lookup table the
    plan does not carry.
    """
    if value is None:
        return None
    if not isinstance(value, str):
        raise ValueError(
            f"Non-literal colour {value!r}. Theme and indexed colours are not "
            "renderable from a plan; resolve them to RGB in the emitter."
        )
    return value if len(value) == 8 else f"FF{value}"


def build_font(record):
    if not record:
        return Font()
    kwargs = {}
    if record.get("name") is not None:
        kwargs["name"] = record["name"]
    if record.get("size") is not None:
        kwargs["size"] = record["size"]
    color = _color(record.get("color"))
    if color is not None:
        kwargs["color"] = color
    # Explicit False matters: an xf pointing at a font with <b val="0"/> is a
    # different font from one with no <b/> at all, and both must round-trip.
    if "bold" in record:
        kwargs["bold"] = bool(record["bold"])
    if "italic" in record:
        kwargs["italic"] = bool(record["italic"])
    if "strike" in record:
        kwargs["strike"] = bool(record["strike"])
    underline = record.get("underline")
    if underline and underline != "none":
        kwargs["underline"] = underline
    return Font(**kwargs)


def build_fill(record):
    if not record:
        return PatternFill()
    pattern = record.get("pattern", "none")
    if pattern == "none":
        return PatternFill()
    kwargs = {"patternType": pattern}
    fg = _color(record.get("fg_color"))
    bg = _color(record.get("bg_color"))
    if fg is not None:
        kwargs["fgColor"] = fg
    if bg is not None:
        kwargs["bgColor"] = bg
    return PatternFill(**kwargs)


def build_side(record):
    if not record:
        return _NO_SIDE
    color = _color(record.get("color"))
    if color is None:
        return Side(style=record["style"])
    return Side(style=record["style"], color=color)


def build_border(record):
    record = record or {}
    return Border(
        left=build_side(record.get("left")),
        right=build_side(record.get("right")),
        top=build_side(record.get("top")),
        bottom=build_side(record.get("bottom")),
        diagonal=build_side(record.get("diagonal")),
    )


def build_alignment(record):
    if not record:
        return Alignment()
    kwargs = {}
    if record.get("horizontal"):
        kwargs["horizontal"] = record["horizontal"]
    if record.get("vertical"):
        kwargs["vertical"] = record["vertical"]
    if record.get("indent"):
        kwargs["indent"] = record["indent"]
    if record.get("wrap_text"):
        kwargs["wrap_text"] = True
    if record.get("shrink_to_fit"):
        kwargs["shrink_to_fit"] = True
    return Alignment(**kwargs)


class StyleTable:
    """
    The plan's style table, resolved once into openpyxl objects.

    Resolved once and shared: openpyxl deduplicates by identity/equality into
    the workbook's own indexed lists, so reusing one Font object across four
    thousand cells produces one `<font>` entry, not four thousand.
    """

    __slots__ = ("fonts", "fills", "borders", "alignments", "number_formats")

    def __init__(self, records):
        self.fonts = []
        self.fills = []
        self.borders = []
        self.alignments = []
        self.number_formats = []
        for record in records:
            record = record or {}
            self.fonts.append(build_font(record.get("font")))
            self.fills.append(build_fill(record.get("fill")))
            self.borders.append(build_border(record.get("border")))
            self.alignments.append(build_alignment(record.get("alignment")))
            self.number_formats.append(record.get("number_format") or "General")

    def __len__(self):
        return len(self.fonts)

    def apply(self, cell, index):
        """
        Assign the WHOLE style. Five assignments, no reads, no merge.

        Order is irrelevant here by construction — each setter writes one slot
        of the cell's style array — and that is the point: the draw-order
        discipline the Node emitter needs exists only because its writer's
        partial assignments interacted. Nothing interacts here.
        """
        cell.font = self.fonts[index]
        cell.fill = self.fills[index]
        cell.border = self.borders[index]
        cell.alignment = self.alignments[index]
        cell.number_format = self.number_formats[index]


def build_differential(record):
    """
    A dxf IS partial by contract — it states only the channels a conditional
    rule overrides — so unlike a cell style it is built as a sparse record.

    A dxf may set fill, border, bold/italic and number format. It may never set
    font colour on a body cell; that is asserted upstream by the style-token
    validator, not silently enforced here, because a renderer that quietly drops
    a channel is worse than one that renders what the plan said.
    """
    record = record or {}
    kwargs = {}
    if "font" in record:
        kwargs["font"] = build_font(record["font"])
    if "fill" in record:
        fill = record["fill"]
        if fill:
            # A dxf fill states its colour in bgColor, not fgColor. That is the
            # OOXML convention for differential formatting and is not the same
            # spelling a cell fill uses.
            kwargs["fill"] = PatternFill(
                patternType=fill.get("pattern", "solid"),
                bgColor=_color(fill.get("bg_color") or fill.get("fg_color")),
            )
    if "border" in record:
        kwargs["border"] = build_border(record["border"])
    if "number_format" in record:
        from openpyxl.styles.differential import DifferentialStyle as _DS  # noqa: F401
        from openpyxl.styles.numbers import NumberFormat

        kwargs["numFmt"] = NumberFormat(numFmtId=0, formatCode=record["number_format"])
    return DifferentialStyle(**kwargs)
