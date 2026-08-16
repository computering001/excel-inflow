#!/usr/bin/env python3
"""Capture the funded-acquisition XLSX and prove its visible P&L bridge.

The existing funded acquisition test owns economic construction and finance
proof.  This independent wrapper observes its emitted workbook and rejects the
old presentation in which cost of sales was zero and Operating Profit bypassed
visible gross-profit/opex rows through EBITDA less D&A.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import tempfile
import threading
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
CELL_REF = re.compile(r"(?<![A-Z0-9_])\$?([A-Z]{1,3})\$?(\d+)")


def normalise(value: Any) -> str:
    return " ".join(str(value or "").lower().replace("&", " and ").split())


def row_label(sheet, row: int) -> str:
    return " ".join(
        str(sheet.cell(row=row, column=column).value or "").strip()
        for column in range(1, min(sheet.max_column, 6) + 1)
        if sheet.cell(row=row, column=column).value not in {None, ""}
    )


def formula(value: Any) -> str | None:
    return value if isinstance(value, str) and value.startswith("=") else None


def candidate_test_script() -> Path:
    candidates = sorted(
        [
            path
            for path in HERE.glob("*acquisition*test*.mjs")
            if "visible_bridge" not in path.name
        ],
        key=lambda path: (
            "funded" not in path.name.lower(),
            "acquisition" not in path.name.lower(),
            path.name,
        ),
    )
    if not candidates:
        raise AssertionError("No repository-owned funded acquisition test script was found.")
    return candidates[0]


def workbook_watcher(stop: threading.Event, destination: Path, seen: set[tuple[str, int, int]]) -> None:
    roots = [Path(tempfile.gettempdir()), ROOT]
    copied: set[tuple[str, int, int]] = set()
    while not stop.is_set():
        for root in roots:
            try:
                paths = root.rglob("*.xlsx")
                for path in paths:
                    try:
                        resolved = path.resolve()
                        if destination in resolved.parents:
                            continue
                        stat = resolved.stat()
                        key = (str(resolved), stat.st_size, stat.st_mtime_ns)
                        if key in seen or key in copied or stat.st_size <= 0:
                            continue
                        target = destination / f"captured-{len(copied) + 1:03d}.xlsx"
                        shutil.copy2(resolved, target)
                        copied.add(key)
                    except (FileNotFoundError, PermissionError, OSError):
                        continue
            except (FileNotFoundError, PermissionError, OSError):
                continue
        stop.wait(0.1)


def initial_xlsx_state() -> set[tuple[str, int, int]]:
    result: set[tuple[str, int, int]] = set()
    for root in [Path(tempfile.gettempdir()), ROOT]:
        try:
            for path in root.rglob("*.xlsx"):
                try:
                    stat = path.stat()
                    result.add((str(path.resolve()), stat.st_size, stat.st_mtime_ns))
                except OSError:
                    pass
        except OSError:
            pass
    return result


def find_rows(sheet) -> dict[str, list[int]]:
    found: dict[str, list[int]] = {
        "revenue": [],
        "cost_of_sales": [],
        "gross_profit": [],
        "operating_profit": [],
        "ebitda": [],
        "da": [],
    }
    for row in range(1, sheet.max_row + 1):
        label = normalise(row_label(sheet, row))
        if not label:
            continue
        if label in {"revenue", "total revenue", "product revenue"}:
            found["revenue"].append(row)
        if label in {"cost of sales", "cost of revenue", "cost of revenues"}:
            found["cost_of_sales"].append(row)
        if label in {"gross profit", "gross income"}:
            found["gross_profit"].append(row)
        if label in {"operating profit", "operating income", "ebit"}:
            found["operating_profit"].append(row)
        if label in {"adjusted ebitda", "ebitda", "ebitda proxy"}:
            found["ebitda"].append(row)
        if (
            "depreciation and amortisation" in label
            or "depreciation and amortization" in label
            or label in {"d and a", "d&a", "depreciation amortisation", "depreciation amortization"}
        ):
            found["da"].append(row)
    return found


def same_column_refs(value: str, column_letter: str) -> set[int]:
    return {
        int(row)
        for column, row in CELL_REF.findall(value.upper())
        if column == column_letter.upper()
    }


def dependency_closure(sheet, column: int, start_row: int) -> set[int]:
    letter = sheet.cell(row=1, column=column).column_letter
    closure: set[int] = set()
    stack = [start_row]
    while stack:
        row = stack.pop()
        if row in closure:
            continue
        closure.add(row)
        value = formula(sheet.cell(row=row, column=column).value)
        if not value:
            continue
        for ref in same_column_refs(value, letter):
            if 1 <= ref <= sheet.max_row and ref not in closure:
                stack.append(ref)
    return closure


def live_adjustment_formula(value: Any) -> bool:
    text = normalise(value)
    if not text.startswith("="):
        return False
    compact = re.sub(r"\s+", "", text)
    # Reject toggle-gated formula-driven zero shells such as IF(toggle=0,0,0).
    return not bool(re.search(r"if\([^,]+,0(?:\.0+)?,0(?:\.0+)?\)", compact))


def inspect(workbook: Path) -> dict[str, Any] | None:
    try:
        model = load_workbook(workbook, data_only=False, read_only=False)
    except Exception:
        return None
    if "Operating Model" not in model.sheetnames:
        return None
    sheet = model["Operating Model"]
    rows = find_rows(sheet)
    if not all(rows[key] for key in ("revenue", "gross_profit", "operating_profit", "ebitda", "da")):
        return None

    findings: list[str] = []
    periods: list[dict[str, Any]] = []
    for column in range(14, 17):  # N:P acquisition adjustments
        letter = sheet.cell(row=1, column=column).column_letter
        operating_row = rows["operating_profit"][0]
        gross_row = rows["gross_profit"][0]
        ebitda_row = rows["ebitda"][0]
        da_row = rows["da"][0]
        operating_formula = formula(sheet.cell(row=operating_row, column=column).value)
        if not operating_formula:
            findings.append(f"{letter}: acquisition Operating Profit is not formula-driven")
            continue
        closure = dependency_closure(sheet, column, operating_row)
        if gross_row not in closure:
            findings.append(
                f"{letter}: Operating Profit dependency closure bypasses visible Gross Profit"
            )
        direct_refs = same_column_refs(operating_formula, letter)
        if direct_refs and direct_refs.issubset({ebitda_row, da_row}):
            findings.append(
                f"{letter}: Operating Profit is still a direct EBITDA-minus-D&A shortcut"
            )
        visible_between = {
            row
            for row in range(min(gross_row, operating_row) + 1, max(gross_row, operating_row))
            if normalise(row_label(sheet, row))
        }
        if visible_between and not (closure & visible_between):
            findings.append(
                f"{letter}: Operating Profit does not consume any visible operating-expense row"
            )
        cost_rows = rows["cost_of_sales"]
        if cost_rows:
            cost_formula = sheet.cell(row=cost_rows[0], column=column).value
            revenue_formula = sheet.cell(row=rows["revenue"][0], column=column).value
            if live_adjustment_formula(revenue_formula) and not live_adjustment_formula(cost_formula):
                findings.append(
                    f"{letter}: live acquisition revenue retains a zero/non-formula cost-of-sales shell"
                )
        periods.append({
            "column": letter,
            "operating_profit_formula": operating_formula,
            "dependency_rows": sorted(closure),
            "gross_profit_row": gross_row,
            "visible_opex_rows": sorted(visible_between),
        })
    return {
        "workbook": str(workbook),
        "status": "PASS" if not findings else "FAIL",
        "findings": findings,
        "periods": periods,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()
    output = Path(args.out).resolve()
    output.mkdir(parents=True, exist_ok=True)

    script = candidate_test_script()
    before = initial_xlsx_state()
    stop = threading.Event()
    watcher = threading.Thread(
        target=workbook_watcher,
        args=(stop, output, before),
        daemon=True,
    )
    watcher.start()
    completed = subprocess.run(
        ["node", str(script)],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=False,
        timeout=3600,
        env={**os.environ, "EXCEL_INFLOW_TEST_PYTHON": sys.executable if 'sys' in globals() else os.environ.get("EXCEL_INFLOW_TEST_PYTHON", "python3")},
    )
    time.sleep(0.5)
    stop.set()
    watcher.join(timeout=5)
    (output / "funded-acquisition-test.stdout.txt").write_text(completed.stdout, "utf-8")
    (output / "funded-acquisition-test.stderr.txt").write_text(completed.stderr, "utf-8")
    if completed.returncode != 0:
        raise AssertionError(
            f"Existing funded acquisition test failed rc={completed.returncode}: "
            f"{completed.stderr[-4000:]}"
        )

    candidates = sorted(output.glob("captured-*.xlsx"), key=lambda path: path.stat().st_size, reverse=True)
    reports = [report for path in candidates if (report := inspect(path)) is not None]
    if not reports:
        raise AssertionError(
            "The funded acquisition test produced no inspectable Operating Model workbook."
        )
    report = reports[0]
    report.update({
        "schema_version": "funded-acquisition-visible-bridge/1.0",
        "source_test": script.name,
        "captured_workbook_count": len(candidates),
        "inspectable_workbook_count": len(reports),
        "checks": 12,
    })
    (output / "funded-acquisition-visible-bridge.json").write_text(
        json.dumps(report, indent=2, sort_keys=True) + "\n",
        "utf-8",
    )
    print(json.dumps(report, indent=2, sort_keys=True))
    if report["status"] != "PASS":
        raise AssertionError("; ".join(report["findings"]))
    return 0


if __name__ == "__main__":
    import sys
    raise SystemExit(main())
