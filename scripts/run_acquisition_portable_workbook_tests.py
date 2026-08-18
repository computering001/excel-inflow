#!/usr/bin/env python3
"""Portable Off/On/Off/On acquisition workbook and cache-parity regression."""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import subprocess
import sys
from pathlib import Path

sys.dont_write_bytecode = True

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
ROOT = HERE.parent


def run(command: list[str], *, env: dict[str, str] | None = None) -> None:
    completed = subprocess.run(
        command,
        cwd=ROOT,
        text=True,
        capture_output=True,
        timeout=600,
        check=False,
        env=env,
    )
    if completed.returncode != 0:
        raise AssertionError(
            f"Command failed ({completed.returncode}): {' '.join(command)}\n"
            f"{completed.stdout[-8000:]}\n{completed.stderr[-8000:]}"
        )


def run_expect_failure(command: list[str]) -> str:
    completed = subprocess.run(
        command,
        cwd=ROOT,
        text=True,
        capture_output=True,
        timeout=600,
        check=False,
    )
    assert completed.returncode != 0, (
        f"Mutation unexpectedly passed: {' '.join(command)}"
    )
    return f"{completed.stdout}\n{completed.stderr}"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def build(case_payload: dict, target: Path) -> tuple[Path, dict, dict]:
    target.parent.mkdir(parents=True, exist_ok=True)
    case_path = target.parent / "case.json"
    case_path.write_text(json.dumps(case_payload, indent=2) + "\n", encoding="utf-8")
    run(
        [
            "node",
            str(HERE / "build_dynamic_model.mjs"),
            str(case_path),
            "--out",
            str(target),
            "--plan-only",
        ]
    )
    plan_path = Path(f"{target}.plan.json")
    run(
        [sys.executable, "-m", "emit", "build", str(plan_path), "--out", str(target)],
        env={
            **os.environ,
            "PYTHONDONTWRITEBYTECODE": "1",
            "PYTHONPATH": str(HERE),
        },
    )
    run(["node", str(HERE / "validate_cache_parity.mjs"), str(target)])
    oracle_path = Path(f"{target}.semantic-oracle.json")
    run(
        [
            sys.executable,
            str(HERE / "verify" / "workbook_semantic_oracle.py"),
            "--xlsx",
            str(target),
            "--contract",
            f"{target}.workbook-proof-contract.json",
            "--out",
            str(oracle_path),
        ]
    )
    return (
        plan_path,
        json.loads(plan_path.read_text(encoding="utf-8")),
        json.loads(Path(f"{target}.row-map.json").read_text(encoding="utf-8")),
    )


def operating_cells(plan: dict) -> dict:
    return next(
        sheet["cells"]
        for sheet in plan["workbook"]["sheets"]
        if sheet["name"] == "Operating Model"
    )


def statement_row(row_map: dict, role: str) -> int:
    for section in row_map["statement_rows"].values():
        for definition in section:
            if definition.get("semantic_role") == role:
                return int(definition["row"])
    raise AssertionError(f"Missing statement role {role}")


def assert_no_excel_errors(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    errors: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "e" or (
                    isinstance(cell.value, str) and cell.value.startswith("#")
                ):
                    errors.append(f"{sheet.title}!{cell.coordinate}={cell.value}")
    assert not errors, f"Cached Excel errors remain: {errors[:20]}"


def workbook_snapshot(workbook_path: Path, *, data_only: bool) -> dict:
    workbook = load_workbook(workbook_path, data_only=data_only, read_only=True)
    snapshot: dict[str, dict[str, object]] = {}
    for sheet in workbook.worksheets:
        cells: dict[str, object] = {}
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not data_only and not (
                    isinstance(value, str) and value.startswith("=")
                ):
                    continue
                if data_only and value is None:
                    continue
                if isinstance(value, (dt.date, dt.datetime, dt.time)):
                    value = value.isoformat()
                cells[cell.coordinate] = value
        snapshot[sheet.title] = cells
    return snapshot


def semantic_identity(row_map: dict) -> dict:
    return {
        key: row_map.get(key)
        for key in (
            "controls",
            "statement_rows",
            "instrument_rows",
            "debt_summary_rows",
            "interest_rows",
            "waterfall_rows",
        )
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("case")
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    source_path = Path(args.case).resolve()
    source = json.loads(source_path.read_text(encoding="utf-8"))
    source_execution_profile = source.get("execution_profile")
    # These state mutations are a forensic regression over a maintained
    # production-shaped fixture. They must not claim production evidence after
    # the acquisition control has been changed.
    source["execution_profile"] = "reference_parity"
    output = Path(args.out).resolve()
    output.mkdir(parents=True, exist_ok=True)

    states = []
    for sequence, enabled in enumerate((0, 1, 0, 1), start=1):
        payload = json.loads(json.dumps(source))
        payload.setdefault("acquisition", {})["enabled"] = enabled
        workbook_path = output / f"state-{sequence}" / "model.xlsx"
        plan_path, plan, row_map = build(payload, workbook_path)
        assert_no_excel_errors(workbook_path)
        states.append(
            (
                enabled,
                workbook_path,
                plan_path,
                plan,
                row_map,
                workbook_snapshot(workbook_path, data_only=False),
                workbook_snapshot(workbook_path, data_only=True),
                semantic_identity(row_map),
            )
        )

    for first, repeated in ((states[0], states[2]), (states[1], states[3])):
        assert first[0] == repeated[0]
        assert sha256(first[2]) == sha256(repeated[2]), (
            f"Acquisition state {first[0]} did not restore deterministically"
        )
        assert first[5] == repeated[5], (
            f"Acquisition state {first[0]} did not restore formula text"
        )
        assert first[6] == repeated[6], (
            f"Acquisition state {first[0]} did not restore cached values"
        )
        assert first[7] == repeated[7], (
            f"Acquisition state {first[0]} did not restore semantic identities"
        )

    # Formula text is state-independent: only the visible control input and its
    # cached consequences change. This catches a builder that emits a different
    # formula surface for On and Off even if each individual state recalculates.
    for state in states[1:]:
        assert states[0][5] == state[5], (
            "Acquisition control state changed workbook formula text"
        )
        assert states[0][7] == state[7], (
            "Acquisition control state changed workbook semantic identities"
        )

    off_cells = operating_cells(states[0][3])
    on_cells = operating_cells(states[1][3])
    row_map = states[1][4]
    controls = row_map["controls"]
    consideration_row = statement_row(row_map, "acquisitions_net_of_cash")
    change_in_debt_row = statement_row(row_map, "change_in_debt")
    opening_cash_row = statement_row(row_map, "opening_cash")
    ending_cash_row = statement_row(row_map, "ending_cash")
    pre_tax_income_row = statement_row(row_map, "pre_tax_income")
    tax_expense_row = statement_row(row_map, "tax_expense")
    acquisition_debt_row = int(row_map["debt_summary_rows"]["acquisition_debt"])
    debt_total_row = int(row_map["debt_summary_rows"]["total_change_in_debt"])
    waterfall_proceeds_row = int(row_map["waterfall_rows"]["non_rcf_debt_proceeds"])
    cash_before_rcf_row = int(row_map["waterfall_rows"]["cash_before_rcf"])

    close_year = int(source["acquisition"]["close_year"])
    forecast_years = [
        int(str(period["date"])[:4])
        for period in source["periods"]
        if period["status"] == "forecast"
    ]
    close_index = forecast_years.index(close_year)
    adjustment_column = ("N", "O", "P")[close_index]
    standalone_column = ("J", "K", "L")[close_index]
    pro_forma_column = ("S", "T", "U")[close_index]

    assert off_cells[f"{adjustment_column}{consideration_row}"].get("v") == 0
    assert on_cells[f"{adjustment_column}{consideration_row}"].get("v") == -float(
        source["acquisition"]["transaction_enterprise_value"]
    )
    assert off_cells[f"{adjustment_column}{change_in_debt_row}"].get("v") == 0
    assert on_cells[f"{adjustment_column}{acquisition_debt_row}"].get("v") == float(
        source["acquisition"]["acquisition_debt_amount"]
    )
    assert f"{adjustment_column}{debt_total_row}" in on_cells[
        f"{adjustment_column}{change_in_debt_row}"
    ].get("f", "")
    assert "$P$8" in on_cells[f"{adjustment_column}{waterfall_proceeds_row}"].get(
        "f", ""
    )
    assert f"{adjustment_column}{waterfall_proceeds_row}" in on_cells[
        f"{adjustment_column}{cash_before_rcf_row}"
    ].get("f", "")
    assert on_cells[f"{adjustment_column}{debt_total_row}"].get("f") == (
        f"IF($P$4=0,0,{pro_forma_column}{debt_total_row}-"
        f"{standalone_column}{debt_total_row})"
    )
    assert on_cells[f"{pro_forma_column}{change_in_debt_row}"].get("f") == (
        f"{pro_forma_column}{debt_total_row}"
    )
    assert abs(
        on_cells[f"{pro_forma_column}{debt_total_row}"].get("v")
        - on_cells[f"{standalone_column}{debt_total_row}"].get("v")
        - on_cells[f"{adjustment_column}{debt_total_row}"].get("v")
    ) < 1e-8

    ending_cash_definition = next(
        definition
        for definition in source["statement_structure"]["cash_flow"]
        if definition.get("semantic_role") == "ending_cash"
    )
    filed_ending_cash = float(
        ending_cash_definition.get("reported_historical_values", ending_cash_definition.get("values"))[2]
    )
    assert on_cells[f"I{ending_cash_row}"].get("v") == filed_ending_cash
    for column, prior_column in (("J", "I"), ("S", "R")):
        address = f"{column}{opening_cash_row}"
        assert on_cells[address].get("f") == f"{prior_column}{ending_cash_row}"
        assert on_cells[address].get("v") == filed_ending_cash

    mutated_plan = json.loads(json.dumps(states[1][3]))
    mutated_cells = operating_cells(mutated_plan)
    mutated_address = f"J{opening_cash_row}"
    mutated_cells[mutated_address]["v"] = filed_ending_cash + 1
    mutation_plan_path = output / "opening-cash-cache-mutation.plan.json"
    mutation_workbook_path = output / "opening-cash-cache-mutation.xlsx"
    mutation_plan_path.write_text(
        json.dumps(mutated_plan, indent=2) + "\n", encoding="utf-8"
    )
    run(
        [
            sys.executable,
            "-m",
            "emit",
            "build",
            str(mutation_plan_path),
            "--out",
            str(mutation_workbook_path),
        ],
        env={
            **os.environ,
            "PYTHONDONTWRITEBYTECODE": "1",
            "PYTHONPATH": str(HERE),
        },
    )
    mutation_output = run_expect_failure(
        [
            "node",
            str(HERE / "validate_cache_parity.mjs"),
            str(mutation_workbook_path),
        ]
    )
    assert mutated_address in mutation_output, (
        "Opening-cash cache mutation did not surface at its changed formula cell"
    )

    amount_basis = (
        f"({source['issuer']['reporting_currency']}, {source['issuer']['units']})"
    )
    for control_id in (
        "transaction_enterprise_value",
        "target_ebitda",
        "acquisition_debt_amount",
    ):
        label = str(on_cells[f"N{controls[control_id]}"].get("v", ""))
        assert label.endswith(amount_basis), (
            f"Acquisition amount control {control_id} does not state currency and units: {label}"
        )

    target_formula = str(on_cells[f"P{controls['target_ebitda']}"].get("f", ""))
    expected_target_formula = (
        f"P{controls['transaction_enterprise_value']}/"
        f"P{controls['entry_ev_to_ebitda']}"
    )
    assert target_formula == expected_target_formula
    assert "IFERROR(" not in target_formula
    assert "MAX(" not in target_formula

    operating_sheet = next(
        sheet
        for sheet in states[1][3]["workbook"]["sheets"]
        if sheet["name"] == "Operating Model"
    )
    validations = {
        record["sqref"]: record
        for record in operating_sheet.get("data_validations", [])
    }
    for control_id in (
        "transaction_enterprise_value",
        "entry_ev_to_ebitda",
    ):
        address = f"P{controls[control_id]}"
        validation = validations.get(address)
        assert validation is not None, (
            f"Acquisition valuation input {address} has no workbook-face validation"
        )
        assert validation.get("type") == "decimal"
        assert validation.get("operator") == "greaterThan"
        assert validation.get("formula1") == "0"
        assert validation.get("show_error_message") is True

    for column in ("N", "O", "P"):
        tax_formula = str(on_cells[f"{column}{tax_expense_row}"].get("f", ""))
        assert "MAX(" not in tax_formula, (
            f"Acquisition tax formula still floors negative PBT: {tax_formula}"
        )
        assert f"-{column}{pre_tax_income_row}*" in tax_formula, (
            f"Acquisition tax does not consume signed target PBT: {tax_formula}"
        )

    report = {
        "schema_version": "acquisition-portable-workbook-test/1.0",
        "status": "PASS",
        "source_fixture_sha256": sha256(source_path),
        "source_execution_profile": source_execution_profile,
        "scenario_execution_profile": source["execution_profile"],
        "states": [state[0] for state in states],
        "checks": 55,
        "close_year": close_year,
        "change_in_debt_row_id": "change_in_debt",
        "portable_formula_cache_parity": "PASS",
        "opening_cash_cache_mutation": "BLOCKED",
        "semantic_oracle": "PASS",
        "formula_text_unchanged_across_states": "PASS",
        "formula_cache_restoration": "PASS",
        "semantic_identity_restoration": "PASS",
        "native_excel": "NOT_AVAILABLE_IN_PORTABLE_TEST",
    }
    (output / "acquisition-portable-workbook-test.json").write_text(
        json.dumps(report, indent=2) + "\n", encoding="utf-8"
    )
    print(json.dumps(report, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
