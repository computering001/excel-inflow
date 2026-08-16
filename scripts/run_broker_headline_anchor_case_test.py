#!/usr/bin/env python3
"""Prove a usable delivered case consumes exactly one EBIT/EBITDA headline."""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore


def find_model_case(run_root: Path) -> Path:
    for path in sorted(run_root.rglob("*.json")):
        try:
            value=json.loads(path.read_text("utf-8"))
        except Exception:
            continue
        if not isinstance(value,dict): continue
        schema=str(value.get("schema_version") or "")
        if "model-case" in schema or ("statement_structure" in value and "periods" in value):
            return path
    raise FileNotFoundError(f"model case missing beneath {run_root}")


def strings(value: Any) -> list[str]:
    result=[]
    if isinstance(value,str): result.append(value)
    elif isinstance(value,dict):
        for key,entry in value.items(): result.append(str(key)); result.extend(strings(entry))
    elif isinstance(value,list):
        for entry in value: result.extend(strings(entry))
    return result


def broker_owned(row: dict[str,Any]) -> bool:
    tokens={item.casefold() for item in strings(row)}
    return any("broker_consensus" in item or "selected_broker" in item or "broker_link" in item for item in tokens)


def row_by_role(model_case: dict[str,Any], role: str) -> list[dict[str,Any]]:
    rows=[]
    for section in ["income_statement","cash_flow"]:
        for row in (model_case.get("statement_structure") or {}).get(section,[]):
            if row.get("semantic_role")==role: rows.append(row)
    return rows


def workbook_rows(workbook: Path) -> tuple[Any,dict[str,int]]:
    book=load_workbook(workbook,data_only=False,read_only=False,keep_links=True)
    sheet=book["Operating Model"]
    labels={}
    for row in range(1,sheet.max_row+1):
        for column in range(1,min(7,sheet.max_column)+1):
            value=sheet.cell(row,column).value
            if isinstance(value,str) and value.strip(): labels[value.strip().casefold()]=row; break
    return sheet,labels


def find_label(labels: dict[str,int], required: set[str]) -> tuple[str,int] | None:
    for label,row in labels.items():
        if required <= set(re.findall(r"[a-z0-9]+",label)): return label,row
    return None


def main() -> int:
    parser=argparse.ArgumentParser(description=__doc__)
    parser.add_argument("report")
    parser.add_argument("--out",required=True)
    args=parser.parse_args()
    report=json.loads(Path(args.report).read_text("utf-8"))
    usable=(report.get("canary_receipts") or {}).get("usable") or {}
    run_root=Path(usable["run_root"])
    workbook=Path(usable["workbook"])
    model_case=json.loads(find_model_case(run_root).read_text("utf-8"))
    ebit_rows=row_by_role(model_case,"ebit")+row_by_role(model_case,"operating_profit")
    ebitda_rows=row_by_role(model_case,"adjusted_ebitda")+row_by_role(model_case,"ebitda")
    da_rows=row_by_role(model_case,"depreciation_and_amortisation")
    ebit_broker=any(broker_owned(row) for row in ebit_rows)
    ebitda_broker=any(broker_owned(row) for row in ebitda_rows)
    assert ebit_broker ^ ebitda_broker, {
        "ebit_broker":ebit_broker,"ebitda_broker":ebitda_broker,
        "ebit_rows":ebit_rows,"ebitda_rows":ebitda_rows,
    }
    assert da_rows, "D&A bridge row is absent"
    sheet,labels=workbook_rows(workbook)
    ebit_label=find_label(labels,{"ebit"}) or find_label(labels,{"operating","profit"})
    ebitda_label=find_label(labels,{"adjusted","ebitda"}) or find_label(labels,{"ebitda"})
    da_label=find_label(labels,{"depreciation","amortisation"})
    assert ebit_label and ebitda_label and da_label, {"ebit":ebit_label,"ebitda":ebitda_label,"da":da_label}
    forecast_columns=range(10,13)  # J:L standalone forecasts.
    formulas={
        "ebit":[sheet.cell(ebit_label[1],column).value for column in forecast_columns],
        "ebitda":[sheet.cell(ebitda_label[1],column).value for column in forecast_columns],
        "da":[sheet.cell(da_label[1],column).value for column in forecast_columns],
    }
    def links(values: list[Any]) -> bool:
        return any(isinstance(value,str) and value.startswith("=") and "Brokers" in value for value in values)
    assert links(formulas["ebit"]) ^ links(formulas["ebitda"]), formulas
    derived_values=formulas["ebitda"] if links(formulas["ebit"]) else formulas["ebit"]
    assert all(isinstance(value,str) and value.startswith("=") for value in derived_values), derived_values
    receipt={
        "schema_version":"broker-headline-anchor-case-test/1.0",
        "status":"PASS",
        "model_case":str(find_model_case(run_root)),
        "workbook":str(workbook),
        "selected_headline":"ebit" if ebit_broker else "adjusted_ebitda",
        "ebit_formulas":formulas["ebit"],
        "ebitda_formulas":formulas["ebitda"],
        "da_formulas":formulas["da"],
    }
    target=Path(args.out); target.parent.mkdir(parents=True,exist_ok=True)
    target.write_text(json.dumps(receipt,indent=2,sort_keys=True,default=str)+"\n","utf-8")
    print(json.dumps({"status":"PASS","selected_headline":receipt["selected_headline"]},sort_keys=True))
    return 0


if __name__=="__main__":
    raise SystemExit(main())
