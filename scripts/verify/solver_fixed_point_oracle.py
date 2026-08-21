#!/usr/bin/env python3
"""Independently recompute the solver's fixed point, the RCF sweep and the debt roll-forward.

WHY THIS EXISTS
---------------
Nothing in this repository recomputed the solver's answer.  The incumbent
finance proof (``scripts/verify/finance_proof.py``) is genuinely independent of
the solver's *code*, but it is not independent of the solver's *answer* on the
three quantities that matter most:

  * the RCF sweep.  ``finance_proof.py:626`` and ``:630`` READ ``rcf_draw`` and
    ``rcf_repayment`` out of the workbook and then check only that
    ``ending = opening + draw - repayment`` (``:652``-``:656``).  A draw that
    contradicts the period's cash need passes: both legs are observations, and
    an identity between two observations is a consistency check, not a
    recomputation.  ``cash_before_rcf`` is likewise read (``:756``), never
    derived, so the sweep's INPUT is taken on trust as well.
  * the circularity.  ``finance_proof.py:698``-``:706`` computes RCF interest
    and the commitment fee from ``opening_rcf_native`` and
    ``ending_rcf_native`` -- both derived from workbook reads.  The
    interest -> cash -> revolver -> interest loop is therefore evaluated at the
    solver's own converged point; the loop is never iterated, so a converged
    WRONG fixed point is indistinguishable from a converged right one.
  * convergence itself.  ``converged``, ``iterations``, ``residual`` and
    ``convergence_tolerance`` are published by the solver and read by nobody:
    no file in ``scripts/verify`` mentions them.

This module closes those three holes.  Its expectation comes from the case's
own declared inputs and from finance, and NEVER from the solved artifact:

  1. it iterates the cash / interest / revolver circularity to its own fixed
     point, to a tolerance an order tighter than the solver's declared one,
     and compares the whole converged state period by period;
  2. it derives the revolver draw and repayment from CASH NEED -- the period's
     cash before the revolver against the declared minimum-cash floor and the
     declared facility capacity -- rather than reading them;
  3. it derives each instrument's roll-forward from the instrument's own
     declared terms and compares the seven movement terms and the ending
     balance against the solver's typed instrument state and against the P4.3
     typed schedule shadow.

NON-SELF-CONFIRMATION
---------------------
Every expected value this oracle produces is a pure function of the case.  The
report publishes ``expectation_digest``: a sha256 over the whole sorted
expectation set.  Perturb any solved number in the artifact and the digest is
byte-identical while findings appear -- which is the machine-checkable form of
"this oracle does not read the answer as its own expectation", and is what
``scripts/run_solver_fixed_point_oracle_tests.py`` asserts.  Handed a solution
file in place of a case (or vice versa) the oracle REFUSES with
``SFP_CASE_NOT_A_CASE`` / ``SFP_SOLUTION_NOT_A_SOLUTION`` rather than passing
vacuously on an empty comparison set.

INDEPENDENCE
------------
Standard library plus openpyxl (only when a workbook is supplied, as a second
observation of the same solved numbers).  Nothing from ``scripts/lib``, the
builder, the case/forecast compilers, the row planner, the solver, the emitter
or the render pipeline is imported -- see ``FORBIDDEN_PRODUCTION_IMPORTS``,
which the shared AST scan in ``oracle_independence.py`` applies to every file
in this directory.  This oracle never launches a subprocess: production code
PRODUCES the artifact it judges (the harness does that), it never computes the
expected answer.

WHAT IS NOT SILENCE
-------------------
``declared_gaps`` carries every quantity this oracle cannot recompute from
declared inputs, each with the reason, and the affected comparisons are named
rather than dropped without trace.  ``observations`` carries typed statements
about behaviour that is neither a break of an identity nor derivable from the
case -- most importantly the cash-tax forecast authority, where the case's own
``forecast_calculation`` and the shipped forecast-authority waterfall disagree.

THE CANDIDATE CASH-FLOW CONSTRUCTIONS
-------------------------------------
Operating cash flow is recomputed under every construction the case's declared
data admits, and the reconciling one is recorded:

  * ``accrual_link`` -- cash taxes = the declared accrual charge
    (``link(tax_expense)``), inside the generic indirect CFO;
  * ``latest_reported_carry_forward`` -- cash taxes held flat at the latest
    reported figure, the waterfall's disclosed fallback;
  * ``declared_graph`` (1.1) -- the case's OWN cash-flow statement resolved
    row by row exactly as the emitted workbook evaluates it: declared
    ``sum``/``link``/``negate`` rules, the forecast-authority waterfall's
    rejection of the legacy cash-tax link in favour of a carry-forward chain,
    the classifier's adoption of the finance captions, and the B3
    cash-interest convention on any row calculated against the interest
    schedule -- with the oracle's own solved circular quantities bound in as
    overrides, mirroring the production statementOverrides binding.  Offered
    only when the case declares a derivable ``cash_from_operations`` graph;
    a graph that cannot be evaluated independently withdraws the branch
    rather than guessing.
"""

from __future__ import annotations

import argparse
import calendar
import datetime as dt
import hashlib
import json
import math
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

# Production-side module tokens no oracle in this directory may import; the
# classification rule lives in oracle_independence.py.
FORBIDDEN_PRODUCTION_IMPORTS: list[str] = [
    "build_dynamic_model",
    "case_compiler",
    "emit",
    "forecast_candidate_compiler",
    "generated",
    "render",
    "row_plan",
    "scripts.lib",
    "solver",
]

ORACLE_VERSION = "solver-fixed-point-oracle/1.1"
# 1.1 -- the candidate cash-flow construction set gained its third member:
# ``declared_graph``, which evaluates the case's OWN declared cash-flow
# presentation (per-row link/hold-flat semantics, the rejected legacy
# cash-tax link, the classified finance captions) with this oracle's solved
# circular quantities bound in, matching the structure the emitted workbook
# actually carries.  Every expectation remains a pure function of the case.

# ---------------------------------------------------------------------------
# Declared tolerance
# ---------------------------------------------------------------------------
#
# Two independently converged fixed points of the same contraction differ by at
# most the sum of the two convergence residuals amplified by 1/(1 - L), where L
# is the loop gain of the circularity.  Here L is bounded by the cash yield and
# the revolver all-in rate acting on half a period's average balance, net of
# tax -- under 0.06 on every case in the corpus -- so the amplification is
# under 1.07.  The solver publishes convergence_tolerance 1e-8 and lands at a
# residual near 6e-9; this oracle iterates to ORACLE_CONVERGENCE_TOLERANCE
# (1e-12), four orders tighter, so essentially the whole gap is the solver's
# own declared slack: 1.07 * (1e-8 + 1e-12) ~ 1.1e-8.
#
# The second contributor is IEEE754 accumulation.  The longest derivation here
# is roughly 250 flops on magnitudes up to 1e4, giving 250 * 2^-53 * 1e4 ~
# 2.8e-10 absolute.
#
# ABS_TOLERANCE is therefore set at 1e-6: about ninety times the sum of both
# bounds, and far below any economically meaningful amount in a model whose
# units are millions (1e-6 of a million is one dollar).  REL_TOLERANCE widens
# it proportionally for large balances only.  The tolerance has teeth by
# mutation, not by assertion: the suite proves a 1e-5 perturbation of a solved
# value is caught, and that a 1e-3 relative tolerance would let it through.
ABS_TOLERANCE = 1e-6
REL_TOLERANCE = 1e-9
ORACLE_CONVERGENCE_TOLERANCE = 1e-12
ORACLE_MAX_ITERATIONS = 400

TOLERANCE_JUSTIFICATION = (
    "1e-6 absolute / 1e-9 relative. Dominated by the solver's own declared "
    "convergence_tolerance (1e-8) amplified by 1/(1-L) with loop gain L<0.06, "
    "i.e. ~1.1e-8; this oracle converges to 1e-12 so it contributes nothing. "
    "IEEE754 accumulation over ~250 flops at magnitudes up to 1e4 adds ~2.8e-10. "
    "1e-6 is ~90x the sum of both bounds and, in units of millions, one dollar."
)

# The only states the P4.3 typed schedule shadow may lawfully claim.
LAWFUL_TYPED_STATES = ("derived_number", "not_applicable", "unresolved")

# Case keys that make a document a v2 model case rather than a solved artifact.
CASE_REQUIRED_KEYS = ("issuer", "periods", "controls", "instruments", "cash_policy")
# Keys that make a document the solver's answer rather than its input.
SOLUTION_REQUIRED_KEYS = ("forecast", "converged", "iterations", "residual")


# ---------------------------------------------------------------------------
# Numeric helpers
# ---------------------------------------------------------------------------


def number(value: Any, default: float = 0.0) -> float:
    if value is None or value == "" or isinstance(value, bool):
        return float(default)
    try:
        result = float(value)
    except (TypeError, ValueError):
        return float(default)
    if not math.isfinite(result):
        return float(default)
    return result


def series(value: Any, length: int, default: float = 0.0) -> List[float]:
    if isinstance(value, list):
        return [number(value[index], default) if index < len(value) else float(default)
                for index in range(length)]
    return [float(default)] * length


def close_enough(actual: Any, expected: Any) -> bool:
    left = number(actual)
    right = number(expected)
    allowance = max(ABS_TOLERANCE, REL_TOLERANCE * max(1.0, abs(left), abs(right)))
    return abs(left - right) <= allowance


def mean(values: Sequence[float]) -> float:
    if not values:
        raise ValueError("mean of an empty series")
    return sum(values) / float(len(values))


# ---------------------------------------------------------------------------
# Calendar / day-count, derived here rather than imported
# ---------------------------------------------------------------------------


def iso_date(value: str) -> dt.date:
    return dt.date.fromisoformat(str(value)[:10])


def one_year_earlier(value: dt.date) -> dt.date:
    day = min(value.day, calendar.monthrange(value.year - 1, value.month)[1])
    return value.replace(year=value.year - 1, day=day)


def period_bounds(case: dict, period_index: int) -> Tuple[dt.date, dt.date]:
    """Inclusive [start, end] of a case period.

    A period begins the day after the previous period ends; the first period in
    the case begins the day after the same date one year earlier.  This is the
    only day-count convention the case declares (period end dates and a fiscal
    year end), so it is the one an independent recomputation must use.
    """
    period_end = iso_date(case["periods"][period_index]["date"])
    if period_index > 0:
        period_start = iso_date(case["periods"][period_index - 1]["date"]) + dt.timedelta(days=1)
    else:
        period_start = one_year_earlier(period_end) + dt.timedelta(days=1)
    return period_start, period_end


def fraction_between(start: dt.date, end: dt.date,
                     period_start: dt.date, period_end: dt.date) -> float:
    start = max(start, period_start)
    end = min(end, period_end)
    if start > end:
        return 0.0
    total = (period_end - period_start).days + 1
    elapsed = (end - start).days + 1
    if total <= 0:
        return 0.0
    return max(0.0, min(1.0, elapsed / float(total)))


def movement_fraction(date_value: Any, period_start: dt.date, period_end: dt.date,
                      active_end: dt.date, fallback: float) -> float:
    """Time-weighting of a dated principal movement; mid-period when undated."""
    if date_value in (None, ""):
        return fallback
    try:
        moment = iso_date(str(date_value))
    except (TypeError, ValueError):
        return fallback
    return fraction_between(moment, active_end, period_start, period_end)


# ---------------------------------------------------------------------------
# Declared-input resolution
# ---------------------------------------------------------------------------


class Refusal(Exception):
    """The oracle cannot judge this pair of documents at all."""

    def __init__(self, code: str, detail: str) -> None:
        super().__init__("%s: %s" % (code, detail))
        self.code = code
        self.detail = detail


def period_split(case: dict) -> Tuple[int, int]:
    statuses = [period.get("status") for period in case.get("periods") or []]
    historical = sum(1 for status in statuses if status == "historical")
    forecast = sum(1 for status in statuses if status == "forecast")
    if forecast <= 0:
        raise Refusal("SFP_CASE_HAS_NO_FORECAST", "the case declares no forecast period")
    if historical + forecast != len(statuses):
        raise Refusal(
            "SFP_UNCLASSIFIED_PERIOD",
            "a period carries neither historical nor forecast status",
        )
    return historical, forecast


def fx_rate(case: dict, currency: Optional[str], period_index: int, kind: str) -> float:
    """Reporting per native, for the case's declared quote direction."""
    reporting = case["issuer"]["reporting_currency"]
    if not currency or currency == reporting:
        return 1.0
    table = (case.get("fx") or {}).get(currency)
    if not table:
        raise Refusal(
            "SFP_FX_UNDECLARED",
            "instrument currency %s carries no declared FX table" % currency,
        )
    key = "average_rates" if kind == "average" else "period_end_rates"
    raw = number((table.get(key) or [None] * (period_index + 1))[period_index])
    if raw <= 0:
        raise Refusal("SFP_FX_UNDECLARED", "non-positive %s FX rate for %s" % (kind, currency))
    return raw if table.get("quote") == "reporting_per_native" else 1.0 / raw


def balance_currency(case: dict, instrument: dict) -> str:
    """Currency of the instrument's AMOUNTS, which legal denomination need not be."""
    if instrument.get("balance_basis") == "reporting_currency_carrying_value":
        return case["issuer"]["reporting_currency"]
    return instrument.get("currency") or case["issuer"]["reporting_currency"]


def all_in_rate(instrument: dict, index: int, forecast_count: int) -> float:
    if instrument.get("rate_type") != "floating":
        return series(instrument.get("coupon_or_all_in_rate"), forecast_count)[index]
    benchmark = series(instrument.get("benchmark_rate"), forecast_count)[index]
    floor = series(instrument.get("benchmark_floor"), forecast_count)[index]
    return max(floor, benchmark) + number(instrument.get("spread_bps")) / 10000.0


def non_cash_components(instrument: dict, index: int, forecast_count: int) -> Tuple[float, float]:
    components = instrument.get("non_cash_movement_components")
    if isinstance(components, dict):
        return (
            series(components.get("fair_value"), forecast_count)[index],
            series(components.get("other"), forecast_count)[index],
        )
    return (0.0, series(instrument.get("other_non_cash_movement"), forecast_count)[index])


def solved_pik(weighted_base: float, rate: float, active_fraction: float) -> float:
    """PIK accretes on the average balance, so it appears on both sides.

    ``pik = (base + pik * active / 2) * rate`` has the closed form below.  This
    is derived here, not assumed: a PIK coupon paid in kind increases the
    balance it accrues on, and the model's own convention (visible in every
    other average-balance term) weights an in-period movement by half the
    period.
    """
    if rate <= 0.0:
        return 0.0
    denominator = 1.0 - rate * active_fraction / 2.0
    if denominator <= 0.0:
        raise Refusal("SFP_PIK_RATE_UNBOUNDED", "a PIK rate at or above 200% has no fixed point")
    return max(0.0, weighted_base * rate / denominator)


def minimum_cash(case: dict) -> float:
    """The declared minimum-liquidity floor.

    An explicit override wins.  Absent one, the floor is the LOWEST historical
    year-end cash balance the case reports -- the only floor the case's own data
    declares, and the conservative reading of it.
    """
    policy = case.get("cash_policy") or {}
    override = policy.get("minimum_cash_override")
    if override is not None:
        return number(override)
    historical = policy.get("historical_year_end_cash")
    if isinstance(historical, list) and historical:
        return min(number(value) for value in historical)
    return 0.0


# ---------------------------------------------------------------------------
# Forecast drivers: which declared authority owns each operating line
# ---------------------------------------------------------------------------

# operating_metrics id -> broker_pack.metrics id.  Only the pairs the shipped
# case schema actually spells differently need an entry.
BROKER_METRIC_ALIASES = {
    "dividends_and_buybacks": "dividends",
}

DRIVER_METRICS = (
    "revenue",
    "adjusted_ebitda",
    "ebit",
    "depreciation_and_amortisation",
    "change_in_working_capital",
    "capex",
    "dividends_and_buybacks",
    "other_non_cash",
    "other_investing",
    "recurring_disclosed_adjustments",
)


FORECAST_WATERFALL_SELECTOR = "Forecast Waterfall"


def included_houses(entry: dict, forecast_index: int) -> List[str]:
    """The houses the case's own declared membership admits for one period.

    Membership is DECLARED DATA on the metric (``consensus_membership``): a
    contributor carries an overall status and a per-period status.  Absent a
    membership block every house in the pack contributes.
    """
    houses = sorted((entry.get("brokers") or {}).keys())
    membership = entry.get("consensus_membership")
    if not isinstance(membership, dict):
        return houses
    by_name = {
        contributor.get("house_name"): contributor
        for contributor in (membership.get("contributors") or [])
    }
    admitted: List[str] = []
    for house in houses:
        contributor = by_name.get(house)
        if contributor is None:
            continue
        if contributor.get("status") != "included":
            continue
        per_period = contributor.get("period_status") or []
        if forecast_index < len(per_period) and per_period[forecast_index] != "included":
            continue
        admitted.append(house)
    return admitted


def broker_series(case: dict, metric_id: str, forecast_count: int) -> Optional[List[float]]:
    """The broker-pack series the case's own ``broker_case`` control selects.

    The declared vocabulary treats ``Consensus`` and ``Model Consensus`` as the
    SAME authority: the model's own consensus, which is the arithmetic mean of
    the houses the declared membership admits for that period.  The pack's
    ``provider_consensus`` field is lineage, not a selectable authority -- it is
    used only when no house is admitted at all.  Any other value naming a house
    selects that house.
    """
    pack = (case.get("broker_pack") or {}).get("metrics") or {}
    entry = pack.get(metric_id)
    if not isinstance(entry, dict):
        return None
    selector = str((case.get("controls") or {}).get("broker_case") or "").strip()
    houses = entry.get("brokers") or {}
    if selector in houses:
        return series(houses[selector], forecast_count)
    values: List[float] = []
    for index in range(forecast_count):
        admitted = included_houses(entry, index)
        if admitted:
            values.append(mean([series(houses[house], forecast_count)[index]
                                for house in admitted]))
        elif entry.get("provider_consensus") is not None:
            values.append(series(entry.get("provider_consensus"), forecast_count)[index])
        else:
            return None
    return values


def provider_consensus_disagreement(case: dict, metric_id: str,
                                    selected: Sequence[float]) -> List[dict]:
    """Periods where the pack's published consensus is not the selected one."""
    entry = ((case.get("broker_pack") or {}).get("metrics") or {}).get(metric_id)
    if not isinstance(entry, dict) or entry.get("provider_consensus") is None:
        return []
    published = series(entry.get("provider_consensus"), len(selected))
    return [
        {
            "forecast_period": index + 1,
            "metric_id": metric_id,
            "provider_consensus": published[index],
            "selected_model_consensus": selected[index],
            "difference": abs(published[index] - selected[index]),
        }
        for index in range(len(selected))
        if not close_enough(published[index], selected[index])
    ]


def driver_series(case: dict, metric_id: str, forecast_count: int,
                  historical_count: int) -> Tuple[List[float], str, Optional[str]]:
    """Resolve one forecast driver from declared data.

    Returns ``(series, authority, ambiguity)``.  ``ambiguity`` is set when two
    declared authorities cover the same metric with different numbers, which is
    a case defect this oracle reports rather than silently resolving.
    """
    broker_id = BROKER_METRIC_ALIASES.get(metric_id, metric_id)
    from_broker = broker_series(case, broker_id, forecast_count)
    metric = (case.get("operating_metrics") or {}).get(metric_id) or {}
    from_supplied: Optional[List[float]] = None
    # Any declared forecast method that carries actual numbers for the forecast
    # periods IS a declared series: supplied_exact, company_assumption and any
    # other label all mean "these are the numbers".  ``not_applicable`` and the
    # formula methods leave the tail null and resolve to nothing here.
    values = metric.get("values")
    if isinstance(values, list) and len(values) >= historical_count + forecast_count:
        tail = values[historical_count:historical_count + forecast_count]
        if all(value is not None for value in tail):
            from_supplied = [number(value) for value in tail]
    if from_broker is not None:
        selector = str((case.get("controls") or {}).get("broker_case") or "")
        ambiguity = None
        if from_supplied is not None and not all(
            close_enough(left, right) for left, right in zip(from_broker, from_supplied)
        ):
            ambiguity = (
                "operating_metrics.%s declares %s (supplied_exact) while the selected "
                "broker authority for %s under broker_case %r is %s; the broker "
                "authority outranks the supplied series, so the supplied numbers are "
                "never published"
                % (metric_id, from_supplied, broker_id, selector, from_broker)
            )
        return from_broker, "broker_pack:%s" % selector, ambiguity
    if from_supplied is not None:
        return (
            from_supplied,
            "operating_metrics.%s" % (metric.get("forecast_method") or "declared_values"),
            None,
        )
    return [0.0] * forecast_count, "absent", None


# ---------------------------------------------------------------------------
# The declared income statement
# ---------------------------------------------------------------------------
#
# The case's ``statement_structure.income_statement`` IS the income statement.
# A role declared as an input row with forecast values is data; a role declared
# as a calculation is that calculation over its refs.  Only the two roles fed by
# the debt schedule (interest income and interest expense) are bound from the
# circularity, because the case cannot state them: they are what the fixed point
# solves for.


class IncomeStatement:
    """A resolver over the case's declared income-statement rows."""

    SUPPORTED_OPERATORS = ("sum", "link", "negate")

    def __init__(self, case: dict, historical_count: int, forecast_count: int) -> None:
        self.rows = {
            row.get("row_id"): row
            for row in ((case.get("statement_structure") or {}).get("income_statement") or [])
            if row.get("row_id")
        }
        self.by_role: Dict[str, str] = {}
        for row_id, row in self.rows.items():
            role = row.get("semantic_role")
            if role and role not in self.by_role:
                self.by_role[role] = row_id
        self.historical_count = historical_count
        self.forecast_count = forecast_count
        self.unsupported_operators: List[str] = []

    def row_for_role(self, role: str) -> Optional[dict]:
        row_id = self.by_role.get(role)
        return self.rows.get(row_id) if row_id else None

    def role_is_supplied_input(self, role: str, index: int) -> bool:
        """True when the role is DATA for this forecast period, not a formula."""
        row = self.row_for_role(role)
        if row is None or row.get("row_type") != "input":
            return False
        values = row.get("values")
        position = self.historical_count + index
        return (
            isinstance(values, list)
            and position < len(values)
            and values[position] is not None
        )

    def resolve_role(self, role: str, index: int,
                     bindings: Dict[str, float]) -> Optional[float]:
        row_id = self.by_role.get(role)
        if row_id is None:
            return None
        return self._resolve(row_id, index, bindings, set())

    def _resolve(self, row_id: str, index: int, bindings: Dict[str, float],
                 seen: set) -> Optional[float]:
        if row_id in seen:
            return None
        row = self.rows.get(row_id)
        if row is None:
            return None
        role = row.get("semantic_role")
        if role and role in bindings:
            return bindings[role]
        if row_id in bindings:
            return bindings[row_id]
        seen = seen | {row_id}
        if row.get("row_type") == "input":
            values = row.get("values")
            position = self.historical_count + index
            if isinstance(values, list) and position < len(values) and values[position] is not None:
                return number(values[position])
            return None
        if row.get("row_type") != "calculation":
            return None
        calculation = row.get("calculation") or {}
        operator = calculation.get("operator")
        refs = calculation.get("refs") or []
        if operator not in self.SUPPORTED_OPERATORS:
            if operator and operator not in self.unsupported_operators:
                self.unsupported_operators.append(operator)
            return None
        if operator == "sum":
            if not refs:
                return None
            parts = [self._resolve(ref, index, bindings, seen) for ref in refs]
            if any(part is None for part in parts):
                return None
            return sum(parts)
        if operator == "link":
            return self._resolve(refs[0], index, bindings, seen) if refs else None
        value = self._resolve(refs[0], index, bindings, seen) if refs else None
        return None if value is None else -value

    def profit_is_supplied(self) -> bool:
        """Is the period's profit DATA rather than a built figure?

        When the case supplies profit before tax or profit for the year as a
        hardcoded row, the model cannot decompose that figure into its cash and
        non-cash parts, so operating cash flow must be built indirectly FROM the
        supplied profit rather than from EBITDA down.  Which of the two routes
        applies is therefore a property of the declared statement, not a choice.
        """
        return any(
            self.role_is_supplied_input(role, index)
            for role in ("pre_tax_income", "net_income")
            for index in range(self.forecast_count)
        )


class DeclaredGraphWithdrawn(Exception):
    """The case's declared cash-flow graph cannot be evaluated independently."""


class DeclaredCashFlowGraph:
    """The case's OWN cash-flow statement, resolved as the workbook evaluates it.

    Where the generic constructions rebuild operating cash flow indirectly
    from the drivers, this graph walks the DECLARED presentation row by row:
    each row's forecast rule under the shipped authority waterfall, the
    classifier's adoption of the finance captions on unlabeled rows, and the
    disclosed rejection of a silent ``link(tax_expense)`` on the cash-tax line
    in favour of a carry-forward chain.  The oracle's own solved circular
    quantities -- the interest schedule, the profit block, the drivers -- are
    bound in as overrides keyed by semantic role, mirroring the production
    statement-overrides binding; every remaining number comes from the case
    alone.

    A graph that cannot be evaluated independently -- an unsupported
    operator, an undeclared reference, a circular reference -- withdraws its
    candidacy (:class:`DeclaredGraphWithdrawn`) rather than guessing.  Cases
    without a declared ``cash_from_operations`` row (every archetype) offer no
    branch at all.
    """

    # Mirror of classifyStatementLine's accepted outcomes for the four roles
    # the normaliser adopts inline on unlabeled cash-flow rows.
    CAPTION_ALIASES = {
        "income taxes paid": "cash_taxes",
        "finance costs paid": "cash_interest_paid",
        "finance income received": "cash_interest_received",
        "net finance costs add-back": "net_finance_addback",
    }

    SUPPORTED_OPERATORS = (
        "sum", "link", "subtract", "negate", "negate_sum", "ratio",
        "negated_ratio", "growth", "tax", "average",
        "prior_period", "prior_period_scaled_by",
    )

    def __init__(self, case: dict, historical_count: int, forecast_count: int) -> None:
        self.historical_count = historical_count
        self.forecast_count = forecast_count
        self.withdrawal: Optional[str] = None
        self.offered = False
        # The declared rows, income statement first so cash-flow links into it
        # resolve, with the classifier mirror applied: an unlabeled row whose
        # caption the production classifier adopts carries that role here too.
        self.rows: Dict[str, dict] = {}
        for section in ("income_statement", "cash_flow"):
            for raw in (case.get("statement_structure") or {}).get(section) or []:
                row = dict(raw)
                if not row.get("row_id"):
                    continue
                if not row.get("semantic_role"):
                    classified = self.CAPTION_ALIASES.get(self._caption(row.get("label")))
                    if classified == "cash_taxes":
                        self._reject_silent_tax_link(row, self.rows)
                    if not row.get("semantic_role"):
                        row["semantic_role"] = classified  # may stay None
                self.rows[row["row_id"]] = row
        if "cash_from_operations" not in self.rows:
            self.withdrawal = "the case declares no cash_from_operations row"
            return
        try:
            previous = self.reported_values()
            for index in range(self.forecast_count):
                self.resolve(index, {}, previous)
        except DeclaredGraphWithdrawn as reason:
            self.withdrawal = str(reason)
            return
        self.offered = True

    @staticmethod
    def _caption(label: Any) -> str:
        return " ".join(str(label or "").lower().replace("–", "-").split())

    def _reject_silent_tax_link(self, row: dict, declared: Dict[str, dict]) -> None:
        """The waterfall's disclosed rejection of a silent cash-tax link.

        A cash-flow row captioned as taxes paid whose declared forecast rule
        is a bare ``link(tax_expense)`` is re-owned by the carry-forward
        chain: the reported figures stand and every forecast column holds the
        latest reported figure flat via ``prior_period`` on the row itself.
        """
        rule = row.get("forecast_calculation") or {}
        refs = rule.get("refs") or []
        ref_row = declared.get(refs[0]) if refs else None
        is_tax_ref = (
            rule.get("operator") == "link"
            and len(refs) == 1
            and (
                refs[0] == "tax_expense"
                or (ref_row or {}).get("semantic_role") == "tax_expense"
            )
        )
        if not is_tax_ref:
            return
        row["semantic_role"] = "cash_taxes"
        original = list(row.get("values") or [])
        row.pop("forecast_calculation", None)
        row["values"] = original[:self.historical_count] + [None] * self.forecast_count
        row["forecast_period_calculations"] = [
            {"operator": "prior_period", "refs": [row["row_id"]]}
            for _ in range(self.forecast_count)
        ]
        row["forecast_treatment"] = "formula"

    def reported_values(self) -> Dict[str, float]:
        """The latest reported column per row -- what period 1 chains from."""
        position = max(0, self.historical_count - 1)
        state: Dict[str, float] = {}
        for row_id, row in self.rows.items():
            values = row.get("values") or []
            state[row_id] = number(values[position]) if position < len(values) else 0.0
        return state

    def forecast_rule(self, row: Optional[dict], forecast_index: int) -> Optional[dict]:
        """The authority-waterfall mirror: which rule owns this row this period."""
        if row is None:
            return None
        per_period = row.get("forecast_period_calculations")
        if isinstance(per_period, list):
            return per_period[forecast_index] if forecast_index < len(per_period) else None
        if row.get("forecast_calculation"):
            return row["forecast_calculation"]
        treatment = row.get("forecast_treatment")
        if treatment in ("broker", "hardcode", "zero", "uncalculated"):
            certified = (
                row.get("row_type") == "uncalculated"
                or bool(row.get("forecast_capture_parent_id"))
            )
            if treatment == "uncalculated" and not certified and row.get("calculation"):
                return row["calculation"]
            return None
        return row.get("calculation")

    def resolve(self, forecast_index: int, overrides: Dict[str, float],
                previous: Dict[str, float]) -> Dict[str, float]:
        """Resolve EVERY declared row for one forecast period.

        A role bound in ``overrides`` reads the bound quantity -- the oracle's
        own solved circular state; otherwise the row's waterfall rule
        evaluates over its refs and ``previous``, the prior period's resolved
        state; otherwise the declared values stand.
        """
        values: Dict[str, float] = {}

        def resolve_row(row_id: str, active: frozenset) -> float:
            if row_id in values:
                return values[row_id]
            if row_id in active:
                raise DeclaredGraphWithdrawn("circular reference through %s" % row_id)
            row = self.rows.get(row_id)
            if row is None:
                raise DeclaredGraphWithdrawn("reference to undeclared row %s" % row_id)
            active = active | {row_id}
            role = row.get("semantic_role")
            if role and role in overrides:
                values[row_id] = number(overrides[role])
                return values[row_id]
            rule = self.forecast_rule(row, forecast_index)
            if rule is not None:
                value = self._evaluate(rule, resolve_row, previous, row, active)
            else:
                declared = row.get("values") or []
                position = self.historical_count + forecast_index
                value = (
                    number(declared[position])
                    if position < len(declared) and declared[position] is not None
                    else 0.0
                )
            values[row_id] = float(value)
            return values[row_id]

        for row_id in self.rows:
            resolve_row(row_id, frozenset())
        return values

    def _evaluate(self, rule: dict, resolve, previous: Dict[str, float],
                  row: dict, active: frozenset) -> float:
        """One declared rule, in the workbook's own evaluation order."""
        operator = rule.get("operator")
        refs = rule.get("refs") or []
        if operator not in self.SUPPORTED_OPERATORS:
            raise DeclaredGraphWithdrawn(
                "unsupported operator %r on %s" % (operator, row.get("row_id"))
            )
        if operator == "prior_period":
            return number(previous.get(refs[0], 0.0))
        if operator == "prior_period_scaled_by":
            if len(refs) < 2:
                raise DeclaredGraphWithdrawn(
                    "%s needs two refs on %s" % (operator, row.get("row_id"))
                )
            current = resolve(refs[1], active)
            prior_denominator = number(previous.get(refs[1], 0.0))
            if prior_denominator == 0:
                return 0.0
            return number(previous.get(refs[0], 0.0)) * current / prior_denominator
        values = [resolve(ref, active) for ref in refs]
        if operator == "sum":
            return sum(values)
        if operator == "link":
            return values[0] if values else 0.0
        if operator == "subtract":
            return (values[0] if values else 0.0) - sum(values[1:])
        if operator == "negate":
            return -(values[0] if values else 0.0)
        if operator == "negate_sum":
            return -sum(values)
        if operator in ("ratio", "negated_ratio") and len(values) < 2:
            raise DeclaredGraphWithdrawn(
                "%s needs two refs on %s" % (operator, row.get("row_id"))
            )
        if operator == "ratio":
            return 0.0 if values[1] == 0 else values[0] / values[1]
        if operator == "negated_ratio":
            return 0.0 if values[1] == 0 else -values[0] / values[1]
        if operator == "growth":
            # Single-ref convention: the prior figure is the referenced row's
            # own prior-period resolved value.
            if not values:
                raise DeclaredGraphWithdrawn(
                    "%s needs a ref on %s" % (operator, row.get("row_id"))
                )
            prior = number(previous.get(refs[0], 0.0))
            return 0.0 if prior == 0 else values[0] / prior - 1.0
        if operator == "tax":
            return -(values[0] * values[1]) if values[0] > 0 else 0.0
        if operator == "average":
            return sum(values) / len(values) if values else 0.0
        raise DeclaredGraphWithdrawn(
            "unsupported operator %r on %s" % (operator, row.get("row_id"))
        )


def effective_tax_rates(case: dict, forecast_count: int) -> Tuple[List[float], str]:
    """The declared effective-tax-rate authority.

    The broker pack outranks the case's own forecast assumption block (the
    ordinary method priority: a contributed rate beats a modelled one).  With
    NO declared rate the model's declared loss/absent rule is a zero charge --
    recorded as an authority of ``absent`` so a silent 0% cannot be mistaken
    for a resolved 0%.
    """
    from_broker = broker_series(case, "effective_tax_rate", forecast_count)
    if from_broker is not None:
        selector = str((case.get("controls") or {}).get("broker_case") or "")
        return from_broker, "broker_pack:%s" % selector
    declared = (case.get("forecast_assumptions") or {}).get("effective_tax_rate")
    if isinstance(declared, list) and declared:
        return series(declared, forecast_count), "forecast_assumptions"
    return [0.0] * forecast_count, "absent"


def cash_flow_row(case: dict, row_id: str) -> dict:
    for row in ((case.get("statement_structure") or {}).get("cash_flow") or []):
        if row.get("row_id") == row_id:
            return row
    return {}


def declared_cash_tax_candidates(case: dict, forecast_count: int,
                                 historical_count: int) -> List[dict]:
    """The closed candidate set for forecast CASH taxes paid.

    The case declares ``income_taxes_paid`` with
    ``forecast_calculation: link(tax_expense)``.  The shipped forecast-authority
    waterfall may instead hold the latest reported cash-tax figure flat, which
    it discloses.  Both are derivable from declared data; WHICH one governs is
    not, so both are enumerated and the reconciling branch is recorded.  A case
    that declares a derivable cash-flow graph additionally carries the
    ``declared_graph`` branch (appended by :class:`Recomputation`), which
    resolves the issuer's own presentation instead of a generic indirect CFO.
    See ``declared_gaps``.
    """
    row = cash_flow_row(case, "income_taxes_paid")
    calculation = (row.get("forecast_calculation") or {})
    linked = (
        calculation.get("operator") == "link"
        and (calculation.get("refs") or [None])[0] == "tax_expense"
    )
    candidates: List[dict] = []
    if linked or not row:
        candidates.append({"branch": "accrual_link", "declared_by": "statement_structure.income_taxes_paid.forecast_calculation"})
    latest: Optional[float] = None
    values = row.get("values")
    if isinstance(values, list) and len(values) >= historical_count:
        history = [value for value in values[:historical_count] if value is not None]
        if history:
            latest = abs(number(history[-1]))
    if latest is None:
        metric = (case.get("operating_metrics") or {}).get("tax") or {}
        values = metric.get("values")
        if isinstance(values, list) and len(values) >= historical_count:
            history = [value for value in values[:historical_count] if value is not None]
            if history:
                latest = abs(number(history[-1]))
    if latest is not None:
        candidates.append({
            "branch": "latest_reported_carry_forward",
            "declared_by": "latest reported cash tax held flat",
            "amount": latest,
        })
    if not candidates:
        candidates.append({"branch": "accrual_link", "declared_by": "default"})
    return candidates


# ---------------------------------------------------------------------------
# The independent recomputation
# ---------------------------------------------------------------------------


class Recomputation:
    """Everything this oracle expects, derived from the case alone."""

    def __init__(self, case: dict) -> None:
        self.case = case
        self.historical_count, self.forecast_count = period_split(case)
        self.declared_gaps: List[dict] = []
        self.controls = case.get("controls") or {}
        self.circularity = int(number(self.controls.get("circularity")))
        self.maturities_roll = int(number(self.controls.get("debt_maturities_roll")))
        self.reporting = case["issuer"]["reporting_currency"]
        self.rcf_policy = case.get("rcf_policy") or {}
        self.cash_policy = case.get("cash_policy") or {}
        self.lease_policy = case.get("lease_policy") or {}
        if str(self.controls.get("broker_case") or "").strip() == FORECAST_WATERFALL_SELECTOR:
            raise Refusal(
                "SFP_FORECAST_WATERFALL_AUTHORITY_NOT_DERIVABLE",
                "broker_case %r routes every operating driver through the forecast "
                "candidate waterfall, whose ranking is not declared in the case; this "
                "oracle cannot derive the drivers independently and refuses rather than "
                "comparing against a guess" % FORECAST_WATERFALL_SELECTOR,
            )
        self._declare_gaps()
        self.provider_consensus_disagreements: List[dict] = []
        self.drivers: Dict[str, List[float]] = {}
        self.driver_authorities: Dict[str, str] = {}
        self.driver_ambiguities: List[str] = []
        for metric_id in DRIVER_METRICS:
            values, authority, ambiguity = driver_series(
                case, metric_id, self.forecast_count, self.historical_count
            )
            self.drivers[metric_id] = values
            self.driver_authorities[metric_id] = authority
            if ambiguity:
                self.driver_ambiguities.append(ambiguity)
            if authority.startswith("broker_pack:"):
                self.provider_consensus_disagreements.extend(
                    provider_consensus_disagreement(
                        case, BROKER_METRIC_ALIASES.get(metric_id, metric_id), values
                    )
                )
        # EBIT is not an independent driver: it IS adjusted EBITDA less the
        # depreciation and amortisation charge.  An oracle that adopted a
        # declared EBIT contradicting its own two components would be asserting
        # an internally inconsistent income statement, so the identity is the
        # expectation and a contradicting declaration becomes a reported
        # disagreement instead.
        declared_ebit = (
            list(self.drivers["ebit"]) if self.driver_authorities["ebit"] != "absent" else None
        )
        self.drivers["ebit"] = [
            self.drivers["adjusted_ebitda"][index]
            - self.drivers["depreciation_and_amortisation"][index]
            for index in range(self.forecast_count)
        ]
        self.discarded_ebit_authority: List[dict] = []
        if declared_ebit is not None:
            for index in range(self.forecast_count):
                if not close_enough(declared_ebit[index], self.drivers["ebit"][index]):
                    self.discarded_ebit_authority.append({
                        "forecast_period": index + 1,
                        "declared_ebit": declared_ebit[index],
                        "adjusted_ebitda_less_depreciation_and_amortisation":
                            self.drivers["ebit"][index],
                        "difference": abs(declared_ebit[index] - self.drivers["ebit"][index]),
                        "declared_by": self.driver_authorities["ebit"],
                    })
            self.driver_authorities["ebit"] = (
                "derived:adjusted_ebitda - depreciation_and_amortisation "
                "(declared authority %s not used)" % self.driver_authorities["ebit"]
            )
        else:
            self.driver_authorities["ebit"] = (
                "derived:adjusted_ebitda - depreciation_and_amortisation"
            )
        self.etr, self.etr_authority = effective_tax_rates(case, self.forecast_count)
        self.income_statement = IncomeStatement(
            case, self.historical_count, self.forecast_count
        )
        self.profit_supplied = self.income_statement.profit_is_supplied()
        if self.profit_supplied:
            self._gap(
                "supplied_profit_cash_decomposition",
                "the case supplies profit before tax or profit for the year as data, so "
                "the cash and non-cash parts of that figure are not declared. Operating "
                "cash flow is recomputed indirectly from the supplied profit plus "
                "depreciation and working capital; any other non-cash item inside the "
                "supplied profit is invisible to both the model and this oracle.",
                ["cash_from_operations"],
            )
        self.minimum_cash = minimum_cash(case)
        self.cash_tax_candidates = declared_cash_tax_candidates(
            case, self.forecast_count, self.historical_count
        )
        # The third construction, offered only where the case declares its own
        # derivable cash-flow presentation.  Cases without such a graph -- the
        # archetypes -- carry the two generic branches alone.
        self.cash_flow_graph = DeclaredCashFlowGraph(
            case, self.historical_count, self.forecast_count
        )
        if self.cash_flow_graph.offered:
            self.cash_tax_candidates.append({
                "branch": "declared_graph",
                "declared_by": "statement_structure.cash_flow.cash_from_operations",
            })

    # -- gap declaration ---------------------------------------------------

    def _gap(self, quantity: str, reason: str, affects: Sequence[str]) -> None:
        self.declared_gaps.append({
            "quantity": quantity,
            "reason": reason,
            "affected_comparisons": list(affects),
        })

    def _declare_gaps(self) -> None:
        self._gap(
            "cash_taxes_paid_forecast_authority",
            "the case declares forecast_calculation link(tax_expense) on income_taxes_paid, "
            "while the shipped forecast-authority waterfall may instead hold the latest "
            "reported cash tax flat and disclose it. Both constructions are derivable from "
            "declared data; which one governs is a waterfall decision the case does not "
            "declare, so this oracle recomputes the whole fixed point under EVERY declared "
            "construction -- the accrual link, the disclosed carry-forward and, where the "
            "case declares a derivable cash-flow presentation, that presentation itself -- "
            "and records which reconciles.",
            ["cash_from_operations", "cash_before_rcf", "rcf_draw", "rcf_repayment", "ending_cash"],
        )
        if number((self.case.get("acquisition") or {}).get("enabled")) == 1:
            self._gap(
                "pro_forma_acquisition_overlay",
                "the funded-acquisition overlay (pro-forma block: acquisition debt, its "
                "interest and the cash consideration) is a transaction overlay on top of the "
                "standalone fixed point. This oracle recomputes the STANDALONE block only; "
                "the pro-forma block is compared for convergence metadata only.",
                ["pro_forma.*"],
            )
        if (self.cash_policy.get("buckets") or []):
            self._gap(
                "explicit_cash_buckets",
                "the case declares explicit cash buckets with per-bucket forecast "
                "treatments and eligibility weights. This oracle recomputes the single "
                "balancing-cash aggregate; per-bucket balances are not recomputed.",
                ["cash_bucket_balances"],
            )
        if number((self.case.get("modules") or {}).get("historical_normalisation")) == 1:
            self._gap(
                "historical_normalisation",
                "historical normalisation restates the reported base this forecast starts "
                "from; the restatement rules are not declared in the case.",
                ["opening_cash", "opening_debt"],
            )

    # -- instrument roll-forward ------------------------------------------

    def instrument_roll_forward(self) -> List[Dict[str, dict]]:
        """Per forecast period, per instrument, the seven movement terms.

        The identity is DERIVED here, not copied.  A period's ending balance is
        whatever the opening balance becomes after each movement the instrument
        can lawfully make:

            available     = opening + issuance + fair_value + other_non_cash
            amortisation  = the scheduled amount, but never more than available
            base_ending   = available - amortisation
            pik           = accretion on the average balance (see solved_pik)
            pre_maturity  = base_ending + pik
            ending        = pre_maturity - maturity_repayment

        Substituting gives exactly seven terms:

            ending = opening + issuance + fair_value + other_non_cash + pik
                     - amortisation - maturity_repayment

        which holds unconditionally EXCEPT where a clamp binds: ``available``
        is floored at zero and ``amortisation`` is capped at ``available``.  The
        clamp conditions are recorded per instrument-period so a break can be
        attributed rather than argued about.
        """
        case = self.case
        instruments = sorted(
            [item for item in case.get("instruments") or []
             if item.get("instrument_id") != self.balancing_rcf_id()],
            key=lambda item: (number(item.get("display_order")), item.get("instrument_id") or ""),
        )
        balances = {
            item["instrument_id"]: number(item.get("opening_balance")) for item in instruments
        }
        out: List[Dict[str, dict]] = []
        for index in range(self.forecast_count):
            period_index = index + self.historical_count
            start, end = period_bounds(case, period_index)
            per_period: Dict[str, dict] = {}
            for instrument in instruments:
                instrument_id = instrument["instrument_id"]
                currency = balance_currency(case, instrument)
                opening = balances[instrument_id]
                issuance = series(instrument.get("new_issuance"), self.forecast_count)[index]
                fair_value, other_non_cash = non_cash_components(
                    instrument, index, self.forecast_count
                )
                unclamped_available = opening + issuance + fair_value + other_non_cash
                available = max(0.0, unclamped_available)
                scheduled = max(0.0, series(
                    instrument.get("scheduled_amortisation"), self.forecast_count
                )[index])
                amortisation = min(available, scheduled)
                base_ending = max(0.0, available - amortisation)

                maturity_raw = instrument.get("maturity_date")
                maturity = iso_date(maturity_raw) if maturity_raw else None
                matures = bool(
                    self.maturities_roll == 1 and maturity is not None and maturity <= end
                )
                active_end = min(maturity, end) if matures and maturity else end
                active_fraction = fraction_between(start, active_end, start, end)
                fallback = active_fraction / 2.0
                issuance_dates = instrument.get("new_issuance_dates") or [None] * self.forecast_count
                amortisation_dates = (
                    instrument.get("scheduled_amortisation_dates") or [None] * self.forecast_count
                )
                weighted_base = max(0.0, (
                    opening * active_fraction
                    + issuance * movement_fraction(
                        issuance_dates[index] if index < len(issuance_dates) else None,
                        start, end, active_end, fallback)
                    + (fair_value + other_non_cash) * fallback
                    - amortisation * movement_fraction(
                        amortisation_dates[index] if index < len(amortisation_dates) else None,
                        start, end, active_end, fallback)
                ))
                pik = (
                    solved_pik(
                        weighted_base,
                        series(instrument.get("pik_rate"), self.forecast_count)[index],
                        active_fraction,
                    )
                    if self.circularity == 1 else 0.0
                )
                pre_maturity = max(0.0, base_ending + pik)
                maturity_repayment = pre_maturity if matures else 0.0
                ending = pre_maturity - maturity_repayment

                average_fx = fx_rate(case, currency, period_index, "average")
                ending_fx = fx_rate(case, currency, period_index, "period_end")
                opening_fx = fx_rate(case, currency, period_index - 1, "period_end")
                cash_coupon = (
                    (weighted_base + pik * active_fraction / 2.0)
                    * all_in_rate(instrument, index, self.forecast_count)
                    * average_fx
                ) if self.circularity == 1 else 0.0

                seven_term = (
                    opening + issuance + fair_value + other_non_cash + pik
                    - amortisation - maturity_repayment
                )
                per_period[instrument_id] = {
                    "opening_native": opening,
                    "issuance_native": issuance,
                    "fair_value_movement_native": fair_value,
                    "other_non_cash_movement_native": other_non_cash,
                    "pik_interest_native": pik,
                    "amortisation_native": amortisation,
                    "maturity_repayment_native": maturity_repayment,
                    "ending_native": ending,
                    "seven_term_ending_native": seven_term,
                    "ending_reporting": ending * ending_fx,
                    "opening_reporting": opening * opening_fx,
                    "cash_coupon_interest_reporting": cash_coupon,
                    "pik_interest_reporting": pik * average_fx,
                    "interest_reporting": cash_coupon + pik * average_fx,
                    "cash_repayment_native": amortisation + maturity_repayment,
                    "cash_repayment_reporting": (amortisation + maturity_repayment) * average_fx,
                    "issuance_reporting": issuance * average_fx,
                    "matures": matures,
                    "available_clamped": unclamped_available < 0.0,
                    "amortisation_clamped": scheduled > available,
                    "currency": currency,
                    "include_in_gross_debt": instrument.get("include_in_gross_debt") is not False,
                    "include_in_net_debt": instrument.get("include_in_net_debt") is not False,
                    "class": instrument.get("class"),
                }
                balances[instrument_id] = ending
            out.append(per_period)
        return out

    def balancing_rcf_id(self) -> Optional[str]:
        policy = self.rcf_policy
        if policy.get("mode") == "none":
            return None
        return policy.get("instrument_id")

    def rcf_instrument(self) -> Optional[dict]:
        target = self.balancing_rcf_id()
        if not target:
            return None
        for item in self.case.get("instruments") or []:
            if item.get("instrument_id") == target:
                return item
        return None

    # -- lease -------------------------------------------------------------

    def lease_schedule(self) -> List[dict]:
        policy = self.lease_policy
        mode = policy.get("mode")
        basis = policy.get("interest_basis") or ("none" if mode == "exclude" else "total_liability")
        historical = policy.get("historical_liabilities")
        balance = (
            number(historical[-1]) if isinstance(historical, list) and historical
            else number(policy.get("opening_liability"))
        )
        supplied_interest_bearing = series(
            policy.get("historical_interest_bearing_liabilities"), max(1, self.historical_count)
        )
        interest_balance = (
            supplied_interest_bearing[-1] if basis == "separately_supplied" else balance
        )
        out: List[dict] = []
        for index in range(self.forecast_count):
            rate = (
                series(policy.get("effective_rate"), self.forecast_count)[index]
                if self.circularity == 1 else 0.0
            )
            principal = 0.0 if mode == "exclude" else series(
                policy.get("principal_repayment"), self.forecast_count
            )[index]
            additions = series(policy.get("additions"), self.forecast_count)[index]
            other = series(policy.get("other_movements"), self.forecast_count)[index]
            if mode == "exclude":
                ending = 0.0
            elif mode == "sourced_balance":
                ending = number((policy.get("forecast_liabilities") or [0.0] * self.forecast_count)[index])
            elif mode == "flat_replacement":
                ending = balance
            elif basis == "total_liability":
                denominator = 1.0 - rate / 2.0
                if denominator <= 0.0:
                    raise Refusal("SFP_LEASE_RATE_UNBOUNDED",
                                  "a lease effective rate at or above 200% has no fixed point")
                ending = max(0.0, (
                    balance + additions + other - principal + interest_balance * rate / 2.0
                ) / denominator)
            else:
                provisional = 0.0 if basis == "none" else (
                    interest_balance
                    + number((policy.get("forecast_interest_bearing_liabilities")
                              or [0.0] * self.forecast_count)[index])
                ) / 2.0 * rate
                ending = max(0.0, balance + additions + provisional + other - principal)
            ending_interest_bearing = (
                0.0 if basis == "none"
                else number((policy.get("forecast_interest_bearing_liabilities")
                             or [0.0] * self.forecast_count)[index])
                if basis == "separately_supplied" else ending
            )
            interest = 0.0 if basis == "none" or self.circularity != 1 else (
                (interest_balance + ending_interest_bearing) / 2.0 * rate
            )
            out.append({
                "opening": balance,
                "ending": ending,
                "principal": principal,
                "interest": interest,
                "opening_interest_bearing": interest_balance,
                "ending_interest_bearing": ending_interest_bearing,
                "included_in_gross_debt": policy.get("include_in_gross_debt") is not False and mode != "exclude",
                "included_in_net_debt": policy.get("include_in_net_debt") is not False and mode != "exclude",
            })
            balance = ending
            interest_balance = ending_interest_bearing
        return out

    # -- the fixed point ---------------------------------------------------

    def fixed_point(self, cash_tax_branch: dict) -> dict:
        """Iterate the cash / interest / revolver circularity to convergence.

        The loop is: the revolver balance sets revolver interest and the
        commitment fee; cash sets interest income; both set pre-tax income and
        therefore tax; those set operating cash flow, which sets the cash
        available before the revolver, which sets the draw or repayment, which
        sets the revolver balance and the cash balance again.

        Gauss-Seidel over the whole forecast, seeded from the declared opening
        balances with the circular terms at zero.
        """
        case = self.case
        rolls = self.instrument_roll_forward()
        leases = self.lease_schedule()
        rcf = self.rcf_instrument()
        rcf_currency = balance_currency(case, rcf) if rcf else None
        capacity_native = number(
            self.rcf_policy.get("capacity")
            if self.rcf_policy.get("capacity") is not None
            else (rcf or {}).get("facility_capacity")
        )
        opening_draw_native = number(
            self.rcf_policy.get("opening_draw")
            if self.rcf_policy.get("opening_draw") is not None
            else (rcf or {}).get("opening_balance")
        )
        fee_convention = self.rcf_policy.get("commitment_fee_convention")
        fee_value = number(self.rcf_policy.get("commitment_fee_value"))
        eligible_pct = number(self.cash_policy.get("eligible_cash_percentage"), 1.0)
        yields = series(self.cash_policy.get("cash_yield"), self.forecast_count)
        other_interest = series(case.get("other_interest"), self.forecast_count)
        non_cash_interest = series(case.get("non_cash_interest"), self.forecast_count)
        fx_on_cash = series(
            (case.get("forecast_assumptions") or {}).get("fx_effect_on_cash"), self.forecast_count
        )
        opening_cash_declared = number(self.cash_policy.get("opening_cash"))

        # The declared_graph branch resolves the case's own cash-flow
        # presentation instead of the generic indirect constructions below.
        using_graph = (
            cash_tax_branch["branch"] == "declared_graph"
            and self.cash_flow_graph is not None
            and self.cash_flow_graph.offered
        )
        graph_overrides: Dict[str, float] = {}

        # State the loop iterates on.
        ending_cash = [opening_cash_declared] * self.forecast_count
        ending_rcf_native = [opening_draw_native] * self.forecast_count

        periods: List[dict] = []
        iterations = 0
        residual = float("inf")
        for iterations in range(1, ORACLE_MAX_ITERATIONS + 1):
            periods = []
            previous_cash = opening_cash_declared
            previous_rcf_native = opening_draw_native
            worst = 0.0
            graph_previous = (
                self.cash_flow_graph.reported_values() if using_graph else {}
            )
            for index in range(self.forecast_count):
                period_index = index + self.historical_count
                roll = rolls[index]
                lease = leases[index]
                average_fx = fx_rate(case, rcf_currency, period_index, "average") if rcf else 1.0
                ending_fx = fx_rate(case, rcf_currency, period_index, "period_end") if rcf else 1.0
                opening_fx = (
                    fx_rate(case, rcf_currency, period_index - 1, "period_end") if rcf else 1.0
                )

                # --- circular terms, evaluated at the current iterate -----
                drawn_average_native = (previous_rcf_native + ending_rcf_native[index]) / 2.0
                rcf_rate = all_in_rate(rcf, index, self.forecast_count) if rcf else 0.0
                rcf_interest = (
                    drawn_average_native * rcf_rate * average_fx
                    if rcf and self.circularity == 1 else 0.0
                )
                commitment_fee = 0.0
                if rcf and self.circularity == 1 and fee_convention == "bps_on_undrawn":
                    commitment_fee = (
                        max(0.0, capacity_native - drawn_average_native)
                        * fee_value / 10000.0 * average_fx
                    )
                interest_income = (
                    (previous_cash + ending_cash[index]) / 2.0 * eligible_pct * yields[index]
                    if self.circularity == 1 else 0.0
                )

                # --- income statement -------------------------------------
                instrument_interest = sum(
                    entry["interest_reporting"] for entry in roll.values()
                )
                non_cash_instrument_interest = sum(
                    entry["pik_interest_reporting"] for entry in roll.values()
                )
                gross_interest = (
                    instrument_interest + rcf_interest + commitment_fee + lease["interest"]
                    + other_interest[index] + non_cash_interest[index]
                )
                net_interest = gross_interest - interest_income
                adjusted_ebitda = self.drivers["adjusted_ebitda"][index]

                # The declared income statement, with only the two
                # schedule-fed roles bound from the circularity.
                bindings = {
                    "revenue": self.drivers["revenue"][index],
                    "adjusted_ebitda": adjusted_ebitda,
                    "interest_income": interest_income,
                    "interest_expense": -gross_interest,
                }
                statement = self.income_statement
                ebit = statement.resolve_role("ebit", index, bindings)
                if ebit is None:
                    ebit = self.drivers["ebit"][index]
                bindings["ebit"] = ebit
                pre_tax_income = statement.resolve_role("pre_tax_income", index, bindings)
                if pre_tax_income is None:
                    pre_tax_income = ebit - net_interest
                bindings["pre_tax_income"] = pre_tax_income
                if statement.role_is_supplied_input("tax_expense", index):
                    # A supplied tax charge is data; the expense is filed negative.
                    tax = -number(statement.resolve_role("tax_expense", index, bindings))
                else:
                    tax = self.etr[index] * pre_tax_income if pre_tax_income > 0.0 else 0.0
                if statement.role_is_supplied_input("net_income", index):
                    net_income = number(statement.resolve_role("net_income", index, bindings))
                else:
                    net_income = pre_tax_income - tax

                if using_graph:
                    # The circular quantities this oracle has just resolved,
                    # bound into the presentation exactly as the production
                    # statement overrides bind them.  The two finance captions
                    # carry the B3 convention wherever their rows are
                    # calculated against the interest schedule.
                    graph_overrides = {
                        "revenue": self.drivers["revenue"][index],
                        "adjusted_ebitda": adjusted_ebitda,
                        "depreciation_and_amortisation":
                            self.drivers["depreciation_and_amortisation"][index],
                        "cash_flow_da": self.drivers["depreciation_and_amortisation"][index],
                        "ebit": ebit,
                        "interest_income": interest_income,
                        "interest_expense": -gross_interest,
                        "net_finance_addback": net_interest,
                        "pre_tax_income": pre_tax_income,
                        "tax_expense": -tax,
                        "net_income": net_income,
                        "change_in_working_capital":
                            self.drivers["change_in_working_capital"][index],
                        "non_cash_interest_addback":
                            non_cash_interest[index] + non_cash_instrument_interest,
                        "other_non_cash": self.drivers["other_non_cash"][index],
                        "capex": self.drivers["capex"][index],
                        "other_investing": self.drivers["other_investing"][index],
                    }
                    costs_row = self.cash_flow_graph.rows.get("finance_costs_paid")
                    if self.cash_flow_graph.forecast_rule(costs_row, index) is not None:
                        graph_overrides["cash_interest_paid"] = -(
                            gross_interest - lease["interest"]
                            - non_cash_interest[index] - non_cash_instrument_interest
                        )
                    income_row = self.cash_flow_graph.rows.get("finance_income_received")
                    if self.cash_flow_graph.forecast_rule(income_row, index) is not None:
                        graph_overrides["cash_interest_received"] = interest_income

                # --- cash taxes: the declared branch under test -----------
                if using_graph:
                    graph_values = self.cash_flow_graph.resolve(
                        index, graph_overrides, graph_previous
                    )
                    # The presentation files its payments as outflows; the
                    # recorded charge keeps the positive-outflow convention of
                    # the generic branches.
                    cash_taxes_paid = -number(graph_values.get("income_taxes_paid"))
                    cash_from_operations = number(
                        graph_values.get("cash_from_operations")
                    )
                    graph_previous.update(graph_values)
                else:
                    if cash_tax_branch["branch"] == "accrual_link":
                        cash_taxes_paid = tax
                    else:
                        cash_taxes_paid = number(cash_tax_branch.get("amount"))

                    cash_interest_paid = (
                        gross_interest - non_cash_interest[index] - non_cash_instrument_interest
                    )
                    if self.profit_supplied:
                        # Indirect from the supplied profit: the only non-cash item
                        # the case declares alongside it is the depreciation charge.
                        cash_from_operations = (
                            net_income
                            + self.drivers["depreciation_and_amortisation"][index]
                            + self.drivers["change_in_working_capital"][index]
                            + self.drivers["other_non_cash"][index]
                        )
                    else:
                        # Anything the declared statement puts between EBITDA and
                        # profit before tax that is neither the depreciation charge
                        # nor net interest -- an impairment inside operating profit,
                        # a share of associates, a discontinued result -- is a
                        # movement the statement declares and the cash flow carries
                        # at its declared sign.  It is recovered as a residual of
                        # the declared statement rather than by matching row names,
                        # so a new line cannot be missed by not being enumerated.
                        # A line that is genuinely non-cash is reversed by the
                        # case's own other_non_cash add-back.
                        statement_extra = (
                            pre_tax_income
                            - (adjusted_ebitda
                               - self.drivers["depreciation_and_amortisation"][index])
                            - interest_income
                            + gross_interest
                        )
                        cash_from_operations = (
                            adjusted_ebitda
                            + self.drivers["change_in_working_capital"][index]
                            + self.drivers["other_non_cash"][index]
                            + statement_extra
                            + interest_income
                            - cash_interest_paid
                            - cash_taxes_paid
                        )
                cash_from_investing = (
                    -self.drivers["capex"][index] + self.drivers["other_investing"][index]
                )
                non_debt_financing = -self.drivers["dividends_and_buybacks"][index]

                cash_before_debt = (
                    previous_cash + cash_from_operations + cash_from_investing
                    + fx_on_cash[index] + non_debt_financing
                )
                non_rcf_issuance = sum(entry["issuance_reporting"] for entry in roll.values())
                non_rcf_repayment = sum(
                    entry["cash_repayment_reporting"] for entry in roll.values()
                )
                mandatory_repayment = non_rcf_repayment + lease["principal"]
                pre_rcf_debt_cash_flow = non_rcf_issuance - mandatory_repayment
                cash_before_rcf = cash_before_debt + pre_rcf_debt_cash_flow

                # --- the revolver sweep, DERIVED FROM CASH NEED -----------
                surplus = cash_before_rcf - self.minimum_cash
                headroom_native = max(0.0, capacity_native - previous_rcf_native)
                if rcf:
                    draw_native = min(max(0.0, -surplus) / average_fx, headroom_native)
                    repayment_native = (
                        0.0 if draw_native > 0.0
                        else min(max(0.0, surplus) / average_fx, previous_rcf_native)
                    )
                else:
                    draw_native = 0.0
                    repayment_native = 0.0
                new_ending_rcf_native = previous_rcf_native + draw_native - repayment_native
                draw_reporting = draw_native * average_fx
                repayment_reporting = repayment_native * average_fx
                new_ending_cash = cash_before_rcf + draw_reporting - repayment_reporting
                liquidity_shortfall = max(0.0, self.minimum_cash - new_ending_cash)

                worst = max(
                    worst,
                    abs(new_ending_cash - ending_cash[index]),
                    abs(new_ending_rcf_native - ending_rcf_native[index]),
                )
                ending_cash[index] = new_ending_cash
                ending_rcf_native[index] = new_ending_rcf_native

                gross_debt = sum(
                    entry["ending_reporting"] for entry in roll.values()
                    if entry["include_in_gross_debt"]
                ) + new_ending_rcf_native * ending_fx + (
                    lease["ending"] if lease["included_in_gross_debt"] else 0.0
                )
                eligible_cash = new_ending_cash * eligible_pct
                net_debt = sum(
                    entry["ending_reporting"] for entry in roll.values()
                    if entry["include_in_net_debt"]
                ) + new_ending_rcf_native * ending_fx + (
                    lease["ending"] if lease["included_in_net_debt"] else 0.0
                ) - eligible_cash

                periods.append({
                    "period": case["periods"][period_index]["date"],
                    "revenue": self.drivers["revenue"][index],
                    "adjusted_ebitda": adjusted_ebitda,
                    "depreciation_and_amortisation": self.drivers["depreciation_and_amortisation"][index],
                    "ebit": ebit,
                    "instrument_interest": instrument_interest,
                    "non_cash_instrument_interest": non_cash_instrument_interest,
                    "rcf_interest": rcf_interest,
                    "rcf_commitment_fee": commitment_fee,
                    "lease_interest": lease["interest"],
                    "other_interest": other_interest[index],
                    "non_cash_interest": non_cash_interest[index],
                    "gross_interest": gross_interest,
                    "interest_income": interest_income,
                    "net_interest": net_interest,
                    "pre_tax_income": pre_tax_income,
                    "effective_tax_rate": self.etr[index],
                    "tax": tax,
                    "net_income": net_income,
                    "cash_taxes_paid": cash_taxes_paid,
                    "cash_from_operations": cash_from_operations,
                    "cash_from_investing": cash_from_investing,
                    "non_debt_financing": non_debt_financing,
                    "cash_before_debt": cash_before_debt,
                    "non_rcf_issuance": non_rcf_issuance,
                    "non_rcf_repayment": non_rcf_repayment,
                    "mandatory_repayment": mandatory_repayment,
                    "pre_rcf_debt_cash_flow": pre_rcf_debt_cash_flow,
                    "cash_before_rcf": cash_before_rcf,
                    "minimum_cash": self.minimum_cash,
                    "cash_surplus_or_deficit": surplus,
                    "rcf_opening_native": previous_rcf_native,
                    "rcf_capacity_native": capacity_native,
                    "rcf_headroom_native": headroom_native,
                    "rcf_draw_native": draw_native,
                    "rcf_repayment_native": repayment_native,
                    "rcf_ending_native": new_ending_rcf_native,
                    "rcf_draw": draw_reporting,
                    "rcf_repayment": repayment_reporting,
                    "ending_rcf": new_ending_rcf_native * ending_fx,
                    "opening_cash": previous_cash,
                    "ending_cash": new_ending_cash,
                    "liquidity_shortfall": liquidity_shortfall,
                    "ending_lease": lease["ending"],
                    "lease_principal": lease["principal"],
                    "gross_debt": gross_debt,
                    "eligible_cash": eligible_cash,
                    "net_debt": net_debt,
                    "rcf_opening_fx": opening_fx,
                    "rcf_average_fx": average_fx,
                    "rcf_ending_fx": ending_fx,
                })
                previous_cash = new_ending_cash
                previous_rcf_native = new_ending_rcf_native
            residual = worst
            if residual <= ORACLE_CONVERGENCE_TOLERANCE:
                break
        return {
            "branch": cash_tax_branch,
            "periods": periods,
            "instrument_roll_forward": rolls,
            "lease": leases,
            "iterations": iterations,
            "residual": residual,
            "converged": residual <= ORACLE_CONVERGENCE_TOLERANCE,
        }


# ---------------------------------------------------------------------------
# Document classification -- the refusal that stops vacuous passes
# ---------------------------------------------------------------------------


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def classify_case(document: Any) -> dict:
    if not isinstance(document, dict):
        raise Refusal("SFP_CASE_NOT_A_CASE", "the case document is not an object")
    missing = [key for key in CASE_REQUIRED_KEYS if key not in document]
    if missing:
        raise Refusal(
            "SFP_CASE_NOT_A_CASE",
            "a model case declares %s; this document is missing %s"
            % (", ".join(CASE_REQUIRED_KEYS), ", ".join(missing)),
        )
    if all(key in document for key in SOLUTION_REQUIRED_KEYS):
        raise Refusal(
            "SFP_CASE_NOT_A_CASE",
            "this document carries the solver's own answer (%s) and is a solution, not a case"
            % ", ".join(SOLUTION_REQUIRED_KEYS),
        )
    return document


def classify_solution(document: Any, case: Optional[dict] = None) -> dict:
    """Accept either the solver's bare answer or the build's wrapped sidecar.

    This oracle recomputes the STANDALONE block (see ``declared_gaps``: the
    funded-acquisition overlay is not recomputed).  A wrapped sidecar carries
    that block explicitly.  A BARE solve of a case whose acquisition module is
    enabled carries the overlay folded into the one block it returns, and is
    refused rather than compared against a standalone expectation.
    """
    if not isinstance(document, dict):
        raise Refusal("SFP_SOLUTION_NOT_A_SOLUTION", "the solution document is not an object")
    candidate = document
    wrapped = False
    if "forecast" not in candidate and isinstance(candidate.get("standalone"), dict):
        candidate = candidate["standalone"]
        wrapped = True
    missing = [key for key in SOLUTION_REQUIRED_KEYS if key not in candidate]
    if missing:
        raise Refusal(
            "SFP_SOLUTION_NOT_A_SOLUTION",
            "a solved artifact declares %s; this document is missing %s"
            % (", ".join(SOLUTION_REQUIRED_KEYS), ", ".join(missing)),
        )
    if not isinstance(candidate.get("forecast"), list) or not candidate["forecast"]:
        raise Refusal("SFP_SOLUTION_NOT_A_SOLUTION", "the solved artifact carries no forecast period")
    if (
        not wrapped
        and case is not None
        and number((case.get("acquisition") or {}).get("enabled")) == 1
    ):
        raise Refusal(
            "SFP_ACQUISITION_OVERLAY_NOT_RECOMPUTED",
            "the case enables the funded-acquisition module, and this artifact is a "
            "single solved block with the overlay folded in. This oracle recomputes the "
            "standalone block only, so it refuses rather than comparing a standalone "
            "expectation against a pro-forma answer. Supply the build's wrapped "
            "solution sidecar, which carries the standalone block separately.",
        )
    return candidate


# ---------------------------------------------------------------------------
# Comparison
# ---------------------------------------------------------------------------

# Solved period fields compared against the independent fixed point, grouped by
# the domain each one evidences.
CONVERGENCE_FIELDS = (
    "revenue", "adjusted_ebitda", "depreciation_and_amortisation", "ebit",
    "instrument_interest", "non_cash_instrument_interest", "rcf_interest",
    "rcf_commitment_fee", "lease_interest", "other_interest", "non_cash_interest",
    "gross_interest", "interest_income", "net_interest", "pre_tax_income", "tax",
    "net_income", "cash_from_operations", "cash_from_investing",
    "non_debt_financing", "cash_before_debt", "non_rcf_issuance",
    "non_rcf_repayment", "mandatory_repayment", "pre_rcf_debt_cash_flow",
    "cash_before_rcf", "minimum_cash", "ending_cash", "ending_lease",
    "lease_principal", "gross_debt", "eligible_cash", "net_debt",
)

RCF_FIELDS = (
    "rcf_opening_native", "rcf_capacity_native", "rcf_draw_native",
    "rcf_repayment_native", "rcf_ending_native", "rcf_draw", "rcf_repayment",
    "ending_rcf", "liquidity_shortfall",
)

ROLL_FORWARD_TERMS = (
    "opening_native", "issuance_native", "fair_value_movement_native",
    "other_non_cash_movement_native", "pik_interest_native",
    "amortisation_native", "maturity_repayment_native", "ending_native",
)


# Every domain this oracle declares.  They are seeded into the report so its
# shape does not depend on what the corpus happens to contain: a domain that
# compared nothing says so, rather than vanishing.
DECLARED_DOMAINS = (
    "convergence",
    "rcf_draw_repay",
    "debt_roll_forward",
    "opening_debt",
    "effective_tax_rate",
    "typed_states",
    "workbook_cross_read",
)


class Comparison:
    def __init__(self, case: dict, solution: dict) -> None:
        self.case = case
        self.solution = solution
        self.findings: List[dict] = []
        self.observations: List[dict] = []
        self.comparisons = 0
        self.domains: Dict[str, Dict[str, int]] = {
            domain: {"compared": 0, "findings": 0} for domain in DECLARED_DOMAINS
        }
        self.expectations: List[Tuple[str, float]] = []

    def _domain(self, domain: str, outcome: str) -> None:
        bucket = self.domains.setdefault(domain, {"compared": 0, "findings": 0})
        bucket["compared"] += 1
        if outcome == "finding":
            bucket["findings"] += 1

    def record_expectation(self, key: str, value: Any) -> None:
        """One number in the expectation set behind ``expectation_digest``.

        The digest is built from the COMPLETE recomputation -- every field of
        every period of every declared branch -- and never from the individual
        comparisons, which are conditional on what the artifact happens to
        carry.  That makes the digest a pure function of the case: it is the
        machine-checkable form of "this oracle does not read the answer".
        """
        self.expectations.append((key, float(number(value))))

    def expect(self, code: str, domain: str, period: int, metric: str,
               actual: Any, expected: Any, **context: Any) -> bool:
        self.comparisons += 1
        if close_enough(actual, expected):
            self._domain(domain, "pass")
            return True
        self._domain(domain, "finding")
        self.findings.append({
            "code": code,
            "domain": domain,
            "forecast_period": period + 1,
            "metric": metric,
            "actual": actual,
            "expected": expected,
            "difference": abs(number(actual) - number(expected)),
            **context,
        })
        return False

    def identity(self, code: str, domain: str, period: int, metric: str,
                 left: Any, right: Any, **context: Any) -> bool:
        """Compare two quantities the ARTIFACT states, not an expectation.

        A roll-forward identity over the solver's own typed terms is a real and
        valuable check -- it catches an ending balance that its own movements do
        not produce -- but neither side of it is independent of the artifact, so
        it is deliberately kept OUT of the expectation digest.  Mixing it in
        would make the digest move with the artifact and destroy the one
        property that proves this oracle does not read the answer.
        """
        self.comparisons += 1
        if close_enough(left, right):
            self._domain(domain, "pass")
            return True
        self._domain(domain, "finding")
        self.findings.append({
            "code": code,
            "domain": domain,
            "forecast_period": period + 1,
            "metric": metric,
            "actual": left,
            "expected": right,
            "difference": abs(number(left) - number(right)),
            "expectation_source": "the artifact's own terms (internal identity)",
            **context,
        })
        return False

    def require(self, code: str, domain: str, period: int, metric: str,
                condition: bool, **context: Any) -> bool:
        self.comparisons += 1
        if condition:
            self._domain(domain, "pass")
            return True
        self._domain(domain, "finding")
        self.findings.append({
            "code": code,
            "domain": domain,
            "forecast_period": period + 1,
            "metric": metric,
            **context,
        })
        return False

    def observe(self, code: str, detail: str, **context: Any) -> None:
        self.observations.append({"code": code, "detail": detail, **context})


def choose_branch(recomputation: Recomputation,
                  solved: List[dict]) -> Tuple[dict, List[dict], float]:
    """Run the fixed point under EVERY declared cash-tax branch; pick the one
    that reconciles.

    The candidate set is closed and derived from the case alone, and the
    report's expectation digest covers every branch, so the choice cannot
    launder the answer into the expectation: a solved value that matches no
    branch still produces findings (against the closest branch, so they are
    informative rather than arbitrary).

    A branch whose own recomputation fails -- it refuses, or it never
    converges -- is not evidence about the others: it scores infinity and is
    chosen only when nothing else survived, in which case the convergence
    refusal below still fires.
    """
    runs: List[dict] = []
    failure: Optional[Refusal] = None
    for candidate in recomputation.cash_tax_candidates:
        try:
            runs.append(recomputation.fixed_point(candidate))
        except Refusal as refusal:
            if failure is None:
                failure = refusal
    if not runs and failure is not None:
        raise failure
    scored: List[Tuple[float, dict]] = []
    for run in runs:
        if not run["converged"]:
            scored.append((float("inf"), run))
            continue
        worst = 0.0
        # Scored on the FIRST forecast period only.  The branches differ there
        # in full, and a later period's own divergence must not be allowed to
        # decide which branch is under test.
        if run["periods"] and solved:
            for field in ("cash_before_rcf", "ending_cash", "rcf_draw", "rcf_repayment"):
                worst = max(worst, abs(
                    number(solved[0].get(field)) - number(run["periods"][0].get(field))
                ))
        scored.append((worst, run))
    scored.sort(key=lambda item: item[0])
    return scored[0][1], runs, scored[0][0]


def compare(case: dict, solution_document: Any, workbook_path: Optional[Path] = None) -> dict:
    solution = classify_solution(solution_document, case)
    recomputation = Recomputation(case)
    solved = solution["forecast"]
    comparison = Comparison(case, solution)

    if len(solved) != recomputation.forecast_count:
        raise Refusal(
            "SFP_PERIOD_COUNT_MISMATCH",
            "the case declares %d forecast periods; the artifact carries %d"
            % (recomputation.forecast_count, len(solved)),
        )

    chosen, runs, branch_score = choose_branch(recomputation, solved)
    if not chosen["converged"]:
        raise Refusal(
            "SFP_ORACLE_DID_NOT_CONVERGE",
            "the independent recomputation did not reach %g after %d iterations (residual %g)"
            % (ORACLE_CONVERGENCE_TOLERANCE, chosen["iterations"], chosen["residual"]),
        )

    # Every branch's expectation enters the digest, so the digest is a function
    # of the case alone even though only one branch is compared.
    for run in runs:
        for index, expected in enumerate(run["periods"]):
            for field in sorted(expected):
                value = expected[field]
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    comparison.record_expectation(
                        "branch:%s|%d|%s" % (run["branch"]["branch"], index, field), value
                    )

    expected_periods = chosen["periods"]

    # No declared cash-tax branch reconciling the first forecast period is a
    # finding in its own right: the comparison that follows is then against the
    # closest of a set none of whose members the artifact matches.
    comparison.require(
        "SFP_NO_CASH_TAX_BRANCH_RECONCILES", "convergence", 0, "cash_tax_branch",
        branch_score <= max(ABS_TOLERANCE, REL_TOLERANCE * 1e6),
        actual=branch_score,
        expected="a declared cash-tax branch reconciling forecast period 1",
        branches_evaluated=[candidate["branch"]
                            for candidate in recomputation.cash_tax_candidates],
        closest_branch=chosen["branch"]["branch"],
    )

    # ---- domain: convergence ------------------------------------------
    declared_tolerance = number(solution.get("convergence_tolerance"), ABS_TOLERANCE)
    comparison.require(
        "SFP_NOT_CONVERGED", "convergence", 0, "solver_converged",
        solution.get("converged") is True,
        actual=solution.get("converged"), expected=True,
    )
    comparison.require(
        "SFP_RESIDUAL_ABOVE_DECLARED_TOLERANCE", "convergence", 0, "solver_residual",
        number(solution.get("residual")) <= declared_tolerance,
        actual=solution.get("residual"), expected="<= %g" % declared_tolerance,
    )
    comparison.require(
        "SFP_ITERATION_COUNT_NOT_POSITIVE", "convergence", 0, "solver_iterations",
        number(solution.get("iterations")) >= 1,
        actual=solution.get("iterations"), expected=">= 1",
    )
    comparison.require(
        "SFP_ORACLE_ITERATION_COUNT_IMPLAUSIBLE", "convergence", 0, "oracle_iterations",
        1 <= chosen["iterations"] <= ORACLE_MAX_ITERATIONS,
        actual=chosen["iterations"], expected="1..%d" % ORACLE_MAX_ITERATIONS,
    )

    for index, expected in enumerate(expected_periods):
        actual = solved[index]
        comparison.require(
            "SFP_PERIOD_LABEL_MISMATCH", "convergence", index, "period",
            str(actual.get("period")) == str(expected["period"]),
            actual=actual.get("period"), expected=expected["period"],
        )
        for field in CONVERGENCE_FIELDS:
            if field not in actual:
                continue
            comparison.expect(
                "SFP_FIXED_POINT_VALUE_MISMATCH", "convergence", index, field,
                actual.get(field), expected[field],
            )

    # ---- domain: effective tax rate ------------------------------------
    for index, expected in enumerate(expected_periods):
        actual = solved[index]
        statement = actual.get("statement_values") or {}
        if "effective_tax_rate" in statement:
            comparison.expect(
                "SFP_ETR_MISMATCH", "effective_tax_rate", index, "effective_tax_rate",
                statement.get("effective_tax_rate"), expected["effective_tax_rate"],
            )
        comparison.expect(
            "SFP_TAX_CHARGE_MISMATCH", "effective_tax_rate", index, "tax",
            actual.get("tax"), expected["tax"],
        )
        # The declared loss rule: no charge on a loss, never a negative charge.
        comparison.require(
            "SFP_TAX_CHARGE_ON_LOSS", "effective_tax_rate", index, "tax_loss_rule",
            not (number(actual.get("pre_tax_income")) <= 0.0
                 and abs(number(actual.get("tax"))) > ABS_TOLERANCE),
            actual=actual.get("tax"), expected=0.0,
            pre_tax_income=actual.get("pre_tax_income"),
        )
    if recomputation.etr_authority == "absent":
        comparison.observe(
            "SFP_ETR_AUTHORITY_ABSENT",
            "no declared effective-tax-rate authority (no broker pack entry, no "
            "forecast assumption); the model's declared rule is a zero charge, which "
            "this oracle expects rather than infers.",
        )

    # ---- domain: rcf draw / repayment from cash need --------------------
    for index, expected in enumerate(expected_periods):
        actual = solved[index]
        for field in RCF_FIELDS:
            if field not in actual:
                continue
            code = (
                "SFP_RCF_DRAW_CONTRADICTS_CASH_NEED"
                if "draw" in field else
                "SFP_RCF_REPAYMENT_CONTRADICTS_CASH_SURPLUS"
                if "repayment" in field else
                "SFP_RCF_SWEEP_VALUE_MISMATCH"
            )
            comparison.expect(
                code, "rcf_draw_repay", index, field,
                actual.get(field), expected[field],
                cash_before_rcf=actual.get("cash_before_rcf"),
                minimum_cash=actual.get("minimum_cash"),
                cash_surplus_or_deficit=expected["cash_surplus_or_deficit"],
            )
        draw = number(actual.get("rcf_draw_native"))
        repayment = number(actual.get("rcf_repayment_native"))
        comparison.require(
            "SFP_RCF_DRAW_AND_REPAYMENT_BOTH_POSITIVE", "rcf_draw_repay", index,
            "exclusive_legs",
            not (draw > ABS_TOLERANCE and repayment > ABS_TOLERANCE),
            actual={"draw": draw, "repayment": repayment},
            expected="at most one positive leg",
        )
        comparison.require(
            "SFP_RCF_CAPACITY_BREACHED", "rcf_draw_repay", index, "capacity",
            -ABS_TOLERANCE <= number(actual.get("rcf_ending_native"))
            <= number(actual.get("rcf_capacity_native")) + ABS_TOLERANCE,
            actual=actual.get("rcf_ending_native"),
            expected="0..%s" % actual.get("rcf_capacity_native"),
        )
        comparison.identity(
            "SFP_RCF_ROLL_FORWARD_BREAK", "rcf_draw_repay", index, "rcf_roll_forward",
            number(actual.get("rcf_ending_native")),
            number(actual.get("rcf_opening_native")) + draw - repayment,
        )
        # A draw while cash sits above the floor, or a repayment while cash sits
        # below it, is a contradiction the identity above cannot see.
        surplus = expected["cash_surplus_or_deficit"]
        comparison.require(
            "SFP_RCF_DRAW_WITHOUT_CASH_NEED", "rcf_draw_repay", index, "draw_needs_deficit",
            not (draw > ABS_TOLERANCE and surplus > ABS_TOLERANCE),
            actual={"draw": draw, "surplus": surplus},
            expected="a draw requires a cash deficit against the minimum-cash floor",
        )
        comparison.require(
            "SFP_RCF_REPAYMENT_WITHOUT_CASH_SURPLUS", "rcf_draw_repay", index,
            "repayment_needs_surplus",
            not (repayment > ABS_TOLERANCE and surplus < -ABS_TOLERANCE),
            actual={"repayment": repayment, "surplus": surplus},
            expected="a repayment requires cash above the minimum-cash floor",
        )
        comparison.expect(
            "SFP_LIQUIDITY_SHORTFALL_MISSTATED", "rcf_draw_repay", index,
            "liquidity_shortfall",
            actual.get("liquidity_shortfall"), expected["liquidity_shortfall"],
        )

    # ---- domain: debt roll-forward -------------------------------------
    for index, expected in enumerate(expected_periods):
        actual = solved[index]
        rolls = chosen["instrument_roll_forward"][index]
        results = {
            entry.get("instrument_id"): entry
            for entry in (actual.get("instrument_results") or [])
        }
        for instrument_id, terms in rolls.items():
            solved_entry = results.get(instrument_id)
            if solved_entry is None:
                comparison.require(
                    "SFP_INSTRUMENT_ABSENT_FROM_SOLUTION", "debt_roll_forward", index,
                    instrument_id, False,
                    expected="a typed instrument state for every declared instrument",
                )
                continue
            for term in ROLL_FORWARD_TERMS:
                if term not in solved_entry:
                    continue
                comparison.expect(
                    "SFP_ROLL_FORWARD_TERM_MISMATCH", "debt_roll_forward", index,
                    "%s.%s" % (instrument_id, term),
                    solved_entry.get(term), terms[term],
                    available_clamped=terms["available_clamped"],
                    amortisation_clamped=terms["amortisation_clamped"],
                )
            # The seven-term identity, on the SOLVER's own typed terms.
            solver_seven_term = (
                number(solved_entry.get("opening_native"))
                + number(solved_entry.get("issuance_native"))
                + number(solved_entry.get("fair_value_movement_native"))
                + number(solved_entry.get("other_non_cash_movement_native"))
                + number(solved_entry.get("pik_interest_native"))
                - number(solved_entry.get("amortisation_native"))
                - number(solved_entry.get("maturity_repayment_native"))
            )
            comparison.identity(
                "SFP_ROLL_FORWARD_BREAK", "debt_roll_forward", index,
                "%s.seven_term_identity" % instrument_id,
                number(solved_entry.get("ending_native")), solver_seven_term,
                note="opening + issuance + fair_value + other_non_cash + pik "
                     "- amortisation - maturity",
            )
            comparison.expect(
                "SFP_ROLL_FORWARD_ENDING_MISMATCH", "debt_roll_forward", index,
                "%s.ending_reporting" % instrument_id,
                solved_entry.get("ending_reporting"), terms["ending_reporting"],
            )
            comparison.expect(
                "SFP_INSTRUMENT_INTEREST_MISMATCH", "debt_roll_forward", index,
                "%s.interest_reporting" % instrument_id,
                solved_entry.get("interest_reporting"), terms["interest_reporting"],
            )

        # ---- domain: opening debt -------------------------------------
        for instrument_id, terms in rolls.items():
            solved_entry = results.get(instrument_id)
            if solved_entry is None:
                continue
            if index == 0:
                comparison.expect(
                    "SFP_OPENING_BALANCE_NOT_THE_DECLARED_ONE", "opening_debt", index,
                    "%s.opening_native" % instrument_id,
                    solved_entry.get("opening_native"), terms["opening_native"],
                    note="the first forecast period must open on the case's declared "
                         "opening_balance",
                )
            else:
                previous = {
                    entry.get("instrument_id"): entry
                    for entry in (solved[index - 1].get("instrument_results") or [])
                }.get(instrument_id)
                if previous is not None:
                    comparison.identity(
                        "SFP_OPENING_BALANCE_NOT_PRIOR_ENDING", "opening_debt", index,
                        "%s.opening_native" % instrument_id,
                        solved_entry.get("opening_native"), previous.get("ending_native"),
                    )
        if index == 0:
            comparison.expect(
                "SFP_OPENING_CASH_NOT_THE_DECLARED_ONE", "opening_debt", index,
                "opening_cash",
                actual.get("statement_values", {}).get("opening_cash",
                                                       expected["opening_cash"]),
                number((case.get("cash_policy") or {}).get("opening_cash")),
            )
            comparison.expect(
                "SFP_OPENING_RCF_NOT_THE_DECLARED_ONE", "opening_debt", index,
                "rcf_opening_native",
                actual.get("rcf_opening_native"), expected["rcf_opening_native"],
            )

    # ---- domain: typed schedule shadow (P4.3) --------------------------
    for index, expected in enumerate(expected_periods):
        actual = solved[index]
        shadow = actual.get("typed_states")
        if shadow is None:
            comparison.require(
                "SFP_TYPED_STATE_SHADOW_ABSENT", "typed_states", index, "typed_states",
                False, expected="a typed schedule shadow on every forecast period",
            )
            continue
        rcf_shadow = (shadow.get("rcf") or {})
        for key, field in (
            ("opening_balance", "rcf_opening_native"),
            ("draw", "rcf_draw_native"),
            ("repayment", "rcf_repayment_native"),
            ("ending_balance", "rcf_ending_native"),
        ):
            typed = rcf_shadow.get(key)
            if not isinstance(typed, dict):
                continue
            state = typed.get("state")
            comparison.require(
                "SFP_TYPED_STATE_ILLEGAL", "typed_states", index, "rcf.%s.state" % key,
                state in LAWFUL_TYPED_STATES,
                actual=state, expected=list(LAWFUL_TYPED_STATES),
            )
            if state == "derived_number":
                comparison.expect(
                    "SFP_TYPED_STATE_VALUE_MISMATCH", "typed_states", index,
                    "rcf.%s" % key, typed.get("value"), expected[field],
                )
            else:
                comparison.require(
                    "SFP_TYPED_STATE_UNRESOLVED_WITH_A_VALUE", "typed_states", index,
                    "rcf.%s" % key, typed.get("value") is None,
                    actual=typed.get("value"), expected=None,
                )
        cash_shadow = (shadow.get("cash") or {}).get("ending_cash")
        if isinstance(cash_shadow, dict) and cash_shadow.get("state") == "derived_number":
            comparison.expect(
                "SFP_TYPED_STATE_VALUE_MISMATCH", "typed_states", index,
                "cash.ending_cash", cash_shadow.get("value"), expected["ending_cash"],
            )

    # ---- second observation: the emitted workbook ----------------------
    workbook_metrics: Dict[str, Any] = {}
    if workbook_path is not None:
        workbook_metrics = compare_workbook(comparison, workbook_path, expected_periods)

    # ---- case-level observations ---------------------------------------
    for ambiguity in recomputation.driver_ambiguities:
        comparison.observe("SFP_DRIVER_AUTHORITY_AMBIGUOUS", ambiguity)
    if recomputation.provider_consensus_disagreements:
        comparison.observe(
            "SFP_PROVIDER_CONSENSUS_IS_NOT_THE_SELECTED_CONSENSUS",
            "the broker pack publishes a provider consensus that differs from the "
            "consensus the model selects and publishes (the mean of the admitted "
            "houses). broker_case 'Consensus' selects the MODEL consensus, so the "
            "pack's published consensus figure is never the number in the workbook.",
            periods=recomputation.provider_consensus_disagreements[:24],
            count=len(recomputation.provider_consensus_disagreements),
        )
    for discarded in recomputation.discarded_ebit_authority:
        comparison.observe(
            "SFP_DECLARED_EBIT_AUTHORITY_DISCARDED",
            "the case declares an EBIT authority whose value contradicts adjusted "
            "EBITDA less depreciation and amortisation; the model publishes the "
            "derived figure, so the declared EBIT is a discarded authority and the "
            "published EBIT is a number no declared source states.",
            **discarded,
        )
    declared_link = any(
        candidate["branch"] == "accrual_link" for candidate in recomputation.cash_tax_candidates
    )
    if declared_link and chosen["branch"]["branch"] != "accrual_link":
        if chosen["branch"]["branch"] == "declared_graph":
            per_period = (
                "%.6g" % number(chosen["periods"][0].get("cash_taxes_paid"))
                if chosen["periods"] else "its declared carry-forward"
            )
            detail = (
                "the case declares forecast_calculation link(tax_expense) on "
                "income_taxes_paid, but the artifact reconciles only under the case's "
                "own declared cash-flow presentation (the declared_graph branch), which "
                "resolves the cash-tax line to %s per period from the presentation's "
                "carry-forward chain rather than from the accrual charge." % per_period
            )
        else:
            detail = (
                "the case declares forecast_calculation link(tax_expense) on "
                "income_taxes_paid, but the artifact reconciles only under the "
                "latest-reported-cash-tax carry-forward branch (%s per period). The "
                "declared link would change operating cash flow, and therefore the "
                "revolver sweep, in every forecast period."
                % chosen["branch"].get("amount")
            )
        comparison.observe(
            "SFP_CASH_TAX_AUTHORITY_OVERRIDES_DECLARED_LINK",
            detail,
            reconciling_branch=chosen["branch"]["branch"],
            declared_branch="accrual_link",
        )

    findings_by_code: Dict[str, int] = {}
    for finding in comparison.findings:
        findings_by_code[finding["code"]] = findings_by_code.get(finding["code"], 0) + 1

    digest = hashlib.sha256(
        "\n".join(
            "%s=%.17g" % (key, value)
            for key, value in sorted(comparison.expectations)
        ).encode("utf-8")
    ).hexdigest()

    return {
        "oracle": ORACLE_VERSION,
        "case_id": case.get("case_id"),
        "verdict": "PASS" if not comparison.findings else "FAIL",
        "comparisons": comparison.comparisons,
        "findings": comparison.findings[:200],
        "findings_total": len(comparison.findings),
        "findings_by_code": findings_by_code,
        "observations": comparison.observations,
        "declared_gaps": recomputation.declared_gaps,
        "domains": comparison.domains,
        "expectation_digest": digest,
        "expectation_count": len(comparison.expectations),
        "tolerance": {
            "absolute": ABS_TOLERANCE,
            "relative": REL_TOLERANCE,
            "oracle_convergence_tolerance": ORACLE_CONVERGENCE_TOLERANCE,
            "justification": TOLERANCE_JUSTIFICATION,
        },
        "recomputation": {
            "forecast_periods": recomputation.forecast_count,
            "historical_periods": recomputation.historical_count,
            "circularity": recomputation.circularity,
            "debt_maturities_roll": recomputation.maturities_roll,
            "minimum_cash": recomputation.minimum_cash,
            "driver_authorities": recomputation.driver_authorities,
            "effective_tax_rate_authority": recomputation.etr_authority,
            "cash_tax_branches_evaluated": [
                candidate["branch"] for candidate in recomputation.cash_tax_candidates
            ],
            "cash_tax_branch_reconciling": chosen["branch"]["branch"],
            "declared_cash_flow_graph": {
                "offered": bool(recomputation.cash_flow_graph.offered),
                "withdrawal": recomputation.cash_flow_graph.withdrawal,
            },
            "non_revolver_instrument_count": len(
                chosen["instrument_roll_forward"][0] if chosen["instrument_roll_forward"] else {}
            ),
            "oracle_iterations": chosen["iterations"],
            "oracle_residual": chosen["residual"],
        },
        "solver_convergence": {
            "converged": solution.get("converged"),
            "iterations": solution.get("iterations"),
            "residual": solution.get("residual"),
            "convergence_tolerance": solution.get("convergence_tolerance"),
        },
        "workbook": workbook_metrics,
        "independence": {
            "forbidden_production_imports": list(FORBIDDEN_PRODUCTION_IMPORTS),
            "subprocesses_launched": 0,
            "expectation_source": "the model case only",
        },
    }


def compare_workbook(comparison: Comparison, workbook_path: Path,
                     expected_periods: List[dict]) -> Dict[str, Any]:
    """Read the same solved numbers a second time, out of the emitted workbook.

    The row map supplies ADDRESSES only; every expectation still comes from the
    independent recomputation.  A disagreement between the workbook's cached
    values and the solution sidecar is itself a typed finding -- two
    observations of one answer must agree.
    """
    from openpyxl import load_workbook  # local: only needed for this second read

    row_map_path = Path(str(workbook_path) + ".row-map.json")
    if not row_map_path.exists():
        comparison.observe(
            "SFP_WORKBOOK_ROW_MAP_ABSENT",
            "no row map beside %s, so the workbook cross-read is skipped"
            % workbook_path.name,
        )
        return {"read": False}
    row_map = load_json(row_map_path)
    columns = ((row_map.get("columns") or {}).get("forecast") or [])
    waterfall = row_map.get("waterfall_rows") or {}
    rows_by_id = row_map.get("rows_by_id") or {}
    book = load_workbook(str(workbook_path), data_only=True)
    sheet = book[book.sheetnames[0]]

    def cell(column: str, row: Any) -> Any:
        if row is None:
            return None
        return sheet["%s%d" % (column, int(row))].value

    channels = (
        ("cash_before_rcf", waterfall.get("cash_before_rcf"), "cash_before_rcf", 1.0),
        ("minimum_cash", waterfall.get("minimum_cash"), "minimum_cash", 1.0),
        ("rcf_draw", waterfall.get("rcf_draw_waterfall"), "rcf_draw", 1.0),
        ("rcf_repayment", waterfall.get("rcf_repayment_waterfall"), "rcf_repayment", 1.0),
        ("ending_rcf", waterfall.get("ending_rcf"), "ending_rcf", 1.0),
        ("opening_rcf", waterfall.get("opening_rcf"), "ending_rcf", 1.0),
        ("liquidity_shortfall", waterfall.get("liquidity_shortfall"),
         "liquidity_shortfall", 1.0),
        ("ending_cash", rows_by_id.get("ending_cash"), "ending_cash", 1.0),
    )
    read = 0
    for index, column in enumerate(columns[:len(expected_periods)]):
        expected = expected_periods[index]
        for name, row, field, sign in channels:
            if row is None:
                continue
            if name == "opening_rcf":
                # The revolver opens on the prior period's close; period one
                # opens on the declared opening draw.
                target = (
                    expected["rcf_opening_native"] * expected["rcf_opening_fx"]
                    if index == 0
                    else expected_periods[index - 1]["ending_rcf"]
                )
            else:
                target = expected[field] * sign
            read += 1
            comparison.expect(
                "SFP_WORKBOOK_VALUE_MISMATCH", "workbook_cross_read", index, name,
                cell(column, row), target,
                address="%s%s" % (column, row),
            )
    return {"read": True, "cells_compared": read, "sheet": book.sheetnames[0]}


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def build_report(case_path: Path, solution_path: Path,
                 workbook_path: Optional[Path]) -> dict:
    try:
        case = classify_case(load_json(case_path))
        return compare(case, load_json(solution_path), workbook_path)
    except Refusal as refusal:
        return {
            "oracle": ORACLE_VERSION,
            "verdict": "REFUSED",
            "refusal": {"code": refusal.code, "detail": refusal.detail},
            "comparisons": 0,
            "findings": [],
            "findings_total": 0,
            "findings_by_code": {},
            "observations": [],
            "declared_gaps": [],
            "domains": {},
            "tolerance": {
                "absolute": ABS_TOLERANCE,
                "relative": REL_TOLERANCE,
                "justification": TOLERANCE_JUSTIFICATION,
            },
            "independence": {
                "forbidden_production_imports": list(FORBIDDEN_PRODUCTION_IMPORTS),
                "subprocesses_launched": 0,
                "expectation_source": "the model case only",
            },
        }


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("case", help="the v2 model case (the only source of expectations)")
    parser.add_argument("solution", help="the solved artifact under test")
    parser.add_argument("--workbook", default=None,
                        help="optional emitted .xlsx, read as a second observation")
    parser.add_argument("--out", default=None, help="write the JSON report here")
    arguments = parser.parse_args(list(argv) if argv is not None else None)

    report = build_report(
        Path(arguments.case),
        Path(arguments.solution),
        Path(arguments.workbook) if arguments.workbook else None,
    )
    text = json.dumps(report, indent=2, sort_keys=False) + "\n"
    if arguments.out:
        Path(arguments.out).write_text(text, encoding="utf-8")
    else:
        sys.stdout.write(text)
    return 0 if report["verdict"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
