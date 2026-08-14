"""
N10 EMIT — place the plan into a workbook with openpyxl.

This stage writes everything openpyxl can express natively. What it cannot
express is left to `patch`, and the split is stated per property rather than
assumed:

  written here   values, formulas, complete cell styles, column widths, row
                 heights, row outline levels, outline summary side, frozen
                 pane, gridline visibility, conditional-format rules and their
                 dxfs, data validation, page margins, calc properties, legacy
                 comment shells

  left to patch  the CACHED VALUE of every formula cell (openpyxl holds a
                 formula or a value, never both), and the modern threaded
                 comment parts (openpyxl writes legacy notes only)

openpyxl is cell-oriented and the plan is cell-level, so the range-to-loop
translation the port note warned about — `getRange("G5:U5").format.fill = x`
becoming a loop, at roughly two hundred sites — does not appear at all. The
emitter resolved the ranges; the renderer never sees one.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.properties import CalcProperties

from .patch import audit_package, read_package
from .styles import StyleTable, build_differential


def _apply_calc_properties(workbook, record):
    if not record:
        return
    kwargs = {}
    if record.get("calc_id") is not None:
        kwargs["calcId"] = int(record["calc_id"])
    if record.get("calc_mode") is not None:
        kwargs["calcMode"] = record["calc_mode"]
    if record.get("full_calc_on_load") is not None:
        kwargs["fullCalcOnLoad"] = record["full_calc_on_load"]
    if record.get("force_full_calc") is not None:
        kwargs["forceFullCalc"] = record["force_full_calc"]
    # Iteration is not decoration. The interest <-> debt balance <-> cash sweep
    # circularity is declared and expected; a workbook opened without iteration
    # enabled reports the model's own design as an error.
    if record.get("iterate") is not None:
        kwargs["iterate"] = record["iterate"]
    if record.get("iterate_count") is not None:
        kwargs["iterateCount"] = int(record["iterate_count"])
    if record.get("iterate_delta") is not None:
        kwargs["iterateDelta"] = float(record["iterate_delta"])
    workbook.calculation = CalcProperties(**kwargs)


def _apply_sheet_chrome(sheet, spec):
    sheet.sheet_view.showGridLines = spec.get("show_gridlines", True)

    height = spec.get("default_row_height")
    if height is not None:
        sheet.sheet_format.defaultRowHeight = height

    outline = spec.get("outline") or {}
    if "summary_below" in outline:
        sheet.sheet_properties.outlinePr.summaryBelow = outline["summary_below"]
    if "summary_right" in outline:
        sheet.sheet_properties.outlinePr.summaryRight = outline["summary_right"]
    if outline.get("outline_level_row") is not None:
        sheet.sheet_format.outlineLevelRow = outline["outline_level_row"]
    if outline.get("outline_level_col") is not None:
        sheet.sheet_format.outlineLevelCol = outline["outline_level_col"]

    # LibreOffice preserves a frozen pane, so this is written natively. The Node
    # emitter patches it at package level only because the legacy workbook library writer
    # accepted freezePanes.freezeRows/freezeColumns and emitted no <pane> at
    # all — an API defect, not a format constraint.
    freeze = spec.get("freeze_pane")
    if freeze:
        sheet.freeze_panes = freeze["top_left_cell"]

    margins = spec.get("page_margins")
    if margins:
        for key, value in margins.items():
            setattr(sheet.page_margins, key, value)

    setup = spec.get("page_setup")
    if setup:
        # Explicit page scaling is one of the three things LibreOffice does not
        # preserve, so it is written here AND re-asserted by the terminal patch.
        if setup.get("orientation"):
            sheet.page_setup.orientation = setup["orientation"]
        if setup.get("scale") is not None:
            sheet.page_setup.scale = setup["scale"]
        if setup.get("paper_size") is not None:
            sheet.page_setup.paperSize = setup["paper_size"]
        if setup.get("fit_to_width") is not None:
            sheet.page_setup.fitToWidth = setup["fit_to_width"]
        if setup.get("fit_to_height") is not None:
            sheet.page_setup.fitToHeight = setup["fit_to_height"]


def _apply_columns(sheet, columns):
    for record in columns:
        for index in range(record["min"], record["max"] + 1):
            dimension = sheet.column_dimensions[get_column_letter(index)]
            if record.get("width") is not None:
                dimension.width = record["width"]
            dimension.hidden = bool(record.get("hidden", False))
            # customWidth is derived by openpyxl from `width is not None`, not
            # set: it is a read-only property. Nothing to assign, and asserting
            # it would be asserting openpyxl's own bookkeeping.
            if record.get("outline_level"):
                dimension.outlineLevel = record["outline_level"]


def _apply_rows(sheet, rows):
    for record in rows:
        dimension = sheet.row_dimensions[record["row"]]
        if record.get("height") is not None:
            # customHeight, like customWidth, is derived from the height being
            # set at all and is read-only.
            dimension.height = record["height"]
        if record.get("hidden"):
            dimension.hidden = True
        # Grouping: summaryBelow=False (set on the sheet) puts the collapse
        # control on the consolidated line, which sits ABOVE its constituents.
        if record.get("outline_level"):
            dimension.outlineLevel = record["outline_level"]
        if record.get("collapsed"):
            dimension.collapsed = True


def _cell_value(record):
    """
    The literal to hand openpyxl.

    A formula wins: openpyxl stores a formula OR a value, never both, and the
    cached value is restored by the patch stage from the plan. Writing the
    cached value here instead would turn a live model into a page of frozen
    numbers, which is the worse of the two failures by a distance.
    """
    if "f" in record:
        formula = record["f"]
        return formula if formula.startswith("=") else f"={formula}"
    if "v" not in record:
        return None
    return record["v"]


def _apply_cells(sheet, cells, table):
    for address, record in cells.items():
        cell = sheet[address]
        value = _cell_value(record)
        if value is not None:
            cell.value = value
        index = record.get("s")
        if index is not None:
            # ONE assignment of a COMPLETE style. See styles.StyleTable.apply.
            table.apply(cell, index)


def _apply_conditional_formats(sheet, blocks, differentials):
    for block in blocks:
        for record in block["rules"]:
            index = record.get("dxf")
            rule = Rule(
                type=record["type"],
                dxf=differentials[index] if index is not None else None,
                priority=record.get("priority", 1),
                operator=record.get("operator"),
                formula=list(record.get("formulas", [])),
                stopIfTrue=record.get("stop_if_true") or None,
            )
            sheet.conditional_formatting.add(block["sqref"], rule)


def _apply_data_validations(sheet, validations):
    for record in validations:
        validation = DataValidation(
            type=record["type"],
            operator=record.get("operator"),
            formula1=record.get("formula1"),
            formula2=record.get("formula2"),
            allowBlank=record.get("allow_blank", False),
            showErrorMessage=record.get("show_error_message", False),
            showInputMessage=record.get("show_input_message", False),
            showDropDown=False,
        )
        validation.sqref = record["sqref"]
        sheet.add_data_validation(validation)


def _apply_comments(sheet, comments, author):
    """
    Legacy comment shells only.

    openpyxl writes xl/commentsN.xml plus the VML shape part; it has no notion
    of a modern threaded comment. The threaded parts are added by `patch`, and
    the gap is stated rather than papered over: a renderer that silently
    downgrades threaded comments to notes has changed a channel without telling
    anyone.
    """
    for record in comments:
        text = record["text"]
        comment = Comment(text, record.get("author") or author or "User")
        # openpyxl's default box is small enough to hide most of a provenance
        # note behind its own edge.
        comment.width = 340
        comment.height = 20 + 14 * (1 + text.count("\n") + len(text) // 60)
        sheet[record["cell"]].comment = comment


def _apply_document_properties(workbook, plan):
    """
    Pin the document properties.

    openpyxl stamps `created` and `modified` with the wall clock, so two builds
    of the same plan differ in docProps/core.xml and nothing else — which is
    enough to fail the determinism double-build (L4: build twice, byte-compare)
    for a reason that has nothing to do with the model. The timestamp is taken
    from the plan, so it is a property of the PLAN rather than of when the
    renderer happened to run.
    """
    from datetime import datetime

    stamp = (plan.get("generator") or {}).get("generated_at")
    when = None
    if stamp:
        try:
            when = datetime.strptime(stamp[:19], "%Y-%m-%dT%H:%M:%S")
        except ValueError:
            when = None
    if when is None:
        when = datetime(1980, 1, 1)
    workbook.properties.created = when
    workbook.properties.modified = when
    workbook.properties.creator = "excel-inflow/emit"
    workbook.properties.lastModifiedBy = "excel-inflow/emit"


def render_plan(plan, path):
    """Render `plan` to `path`. Returns a dict of what was placed."""
    workbook_spec = plan["workbook"]

    workbook = Workbook()
    workbook.remove(workbook.active)
    _apply_document_properties(workbook, plan)

    _apply_calc_properties(workbook, workbook_spec.get("calc_properties"))

    table = StyleTable(workbook_spec["styles"])
    differentials = [build_differential(record) for record in workbook_spec.get("differential_styles", [])]

    default_font = workbook_spec.get("default_font")
    if default_font:
        # Font 0 is the workbook default. N14 asserts the rendered font set is
        # exactly {Carlito}, and openpyxl seeds font 0 with its own Calibri 11 —
        # which would put a second family in the file that not one cell uses,
        # and fail the gate for a reason no cell can explain.
        #
        # It has to be replaced in place rather than appended: font 0 is
        # referenced by the default named style, and appending a Carlito would
        # leave the Calibri behind. `_fonts` is an IndexedList keeping a
        # value->index map alongside the list, so the map is rebuilt after the
        # substitution or every later lookup resolves against a stale key.
        from .styles import build_font

        fonts = workbook._fonts
        list.__setitem__(fonts, 0, build_font(default_font))
        fonts._rebuild_dict()

    placed = {"sheets": []}
    for spec in workbook_spec["sheets"]:
        sheet = workbook.create_sheet(title=spec["name"])
        _apply_sheet_chrome(sheet, spec)
        _apply_columns(sheet, spec.get("columns", []))
        _apply_cells(sheet, spec["cells"], table)
        _apply_rows(sheet, spec.get("rows", []))
        _apply_conditional_formats(sheet, spec.get("conditional_formats", []), differentials)
        _apply_data_validations(sheet, spec.get("data_validations", []))
        _apply_comments(sheet, spec.get("comments", []), workbook_spec.get("comment_author"))
        placed["sheets"].append(
            {
                "name": spec["name"],
                "cells": len(spec["cells"]),
                "formulas": sum(1 for cell in spec["cells"].values() if "f" in cell),
                "conditional_formats": len(spec.get("conditional_formats", [])),
                "data_validations": len(spec.get("data_validations", [])),
                "comments": len(spec.get("comments", [])),
                "images": len(spec.get("images", [])),
            }
        )

    workbook.save(path)
    placed["styles"] = len(table)
    placed["differential_styles"] = len(differentials)
    # AUDIT WHAT WAS WRITTEN, HERE TOO. `patch` audits the terminal package and
    # is the only stage that sees what a reader opens — but the stages are
    # separately runnable on purpose, and a `render` run that is never patched
    # would otherwise ship unaudited. Auditing at both ends also localises the
    # blame: a package that is coherent here and incoherent after `patch` was
    # broken by `patch`.
    placed["package_audit"] = audit_package(read_package(path))
    return placed
