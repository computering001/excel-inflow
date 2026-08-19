#!/usr/bin/env python3
"""Independent checks over emitted Excel Inflow artifacts.

This module consumes only the public model case, emitted workbook, forecast
receipt, workbook proof contract and source-arithmetic artifact.  It imports no
compiler, forecast compiler, row planner, solver or workbook emitter.
"""

from __future__ import annotations

import json
import hashlib
import math
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .oracle_independence import production_imports


FORECAST_COLUMNS = ("J", "K", "L")
ADJUSTMENT_COLUMNS = ("N", "O", "P")
BROKER_COLUMNS = ("D", "E", "F")
BLUE = "0000FF"
BLACK = "000000"
GREEN = "008000"
# Production-side module tokens this oracle (and every scripts/verify oracle)
# must never import; the classification rule lives in oracle_independence.
FORBIDDEN_PRODUCTION_IMPORTS: list[str] = [
    "build_dynamic_model",
    "case_compiler",
    "emit",
    "forecast_candidate_compiler",
    "generated",
    "render",
    "row_plan",
    "scripts.lib",
    "solver",
]


def _finite(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value))


def _colour(cell) -> str | None:
    colour = cell.font.color
    if colour is None or colour.type != "rgb" or not colour.rgb:
        return None
    return str(colour.rgb)[-6:].upper()


def _normal_formula(value: Any) -> str:
    return str(value or "").replace("$", "").replace("'", "").replace(" ", "").upper()


def _formula_refs(value: Any) -> set[str]:
    formula = _normal_formula(value)
    refs: set[str] = set()
    for column, row, end_column, end_row in re.findall(r"(?<![A-Z0-9_])([A-Z]{1,3})(\d+)(?::([A-Z]{1,3})(\d+))?", formula):
        if not end_column:
            refs.add(f"{column}{row}")
            continue
        if column == end_column:
            refs.update(f"{column}{number}" for number in range(int(row), int(end_row) + 1))
        else:
            refs.add(f"{column}{row}")
            refs.add(f"{end_column}{end_row}")
    return refs


def _labels(sheet) -> dict[str, list[int]]:
    result: dict[str, list[int]] = defaultdict(list)
    for row in range(1, sheet.max_row + 1):
        value = sheet[f"B{row}"].value
        if isinstance(value, str):
            result[value].append(row)
    return result


def _one(labels: dict[str, list[int]], label: str) -> int:
    rows = labels.get(label) or []
    if len(rows) != 1:
        raise KeyError(f"expected one row labelled {label!r}, found {rows}")
    return rows[0]


def _first(labels: dict[str, list[int]], label: str) -> int:
    rows = labels.get(label) or []
    if not rows:
        raise KeyError(f"expected a row labelled {label!r}")
    return rows[0]


def _finding(findings: list[dict], code: str, **context: Any) -> None:
    findings.append({"code": code, **context})


def inspect_consensus(workbook_path: Path, findings: list[dict]) -> dict:
    formulas = load_workbook(workbook_path, data_only=False, read_only=False)
    brokers = formulas["Brokers"]
    operating = formulas["Operating Model"]
    labels = _labels(brokers)
    operating_formula_cells = [
        cell for row in operating.iter_rows() for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    model_rows = [
        row for label, rows in labels.items() if label.endswith(" — Model Consensus") for row in rows
    ]
    checked = 0
    green_links_checked = 0
    for model_row in model_rows:
        label = str(brokers[f"B{model_row}"].value)
        metric = label[: -len(" — Model Consensus")]
        selected_rows = labels.get(f"{metric} — Selected Forecast") or []
        if len(selected_rows) != 1:
            _finding(findings, "CONSENSUS_SELECTED_ROW", metric=metric)
            continue
        selected_row = selected_rows[0]
        provider_row = model_row + 1 if brokers[f"B{model_row + 1}"].value == "Provider Consensus" else None
        for column in BROKER_COLUMNS:
            model_cell = brokers[f"{column}{model_row}"]
            selected_cell = brokers[f"{column}{selected_row}"]
            refs = _formula_refs(model_cell.value)
            named_refs = {
                ref for ref in refs
                if ref.startswith(column) and isinstance(brokers[ref].value, (int, float))
            }
            expected_named_refs: set[str] = set()
            for candidate_row in range(model_row - 1, 0, -1):
                candidate_label = brokers[f"B{candidate_row}"].value
                candidate = brokers[f"{column}{candidate_row}"]
                if not isinstance(candidate_label, str) or not _finite(candidate.value):
                    break
                if candidate_label in {
                    "Provider Consensus", "High", "Low", "Contributor Count", "Excluded Count"
                }:
                    break
                expected_named_refs.add(candidate.coordinate)
            if not str(model_cell.value or "").upper().startswith("=IFERROR(AVERAGE("):
                _finding(findings, "CONSENSUS_MODEL_NOT_FORMULA", cell=model_cell.coordinate)
            if _colour(model_cell) != BLACK:
                _finding(findings, "CONSENSUS_MODEL_NOT_BLACK", cell=model_cell.coordinate)
            if not named_refs:
                _finding(findings, "CONSENSUS_MEMBERSHIP_EMPTY", cell=model_cell.coordinate)
            if named_refs != expected_named_refs:
                _finding(
                    findings, "CONSENSUS_MEMBERSHIP_MISMATCH", cell=model_cell.coordinate,
                    expected=sorted(expected_named_refs), actual=sorted(named_refs),
                )
            for ref in sorted(named_refs):
                ref_row = brokers[ref].row
                ref_label = brokers[f"B{ref_row}"].value
                if ref_row >= model_row or ref_label in {
                    "Provider Consensus", "High", "Low", "Contributor Count", "Excluded Count"
                }:
                    _finding(findings, "CONSENSUS_MEMBERSHIP_INVALID", cell=model_cell.coordinate, member=ref)
                if _colour(brokers[ref]) != BLUE or not _finite(brokers[ref].value):
                    _finding(findings, "CONSENSUS_NAMED_SOURCE_NOT_BLUE", cell=ref)
            if provider_row:
                provider_cell = brokers[f"{column}{provider_row}"]
                if _colour(provider_cell) != BLUE or not _finite(provider_cell.value):
                    _finding(findings, "CONSENSUS_PROVIDER_NOT_BLUE", cell=provider_cell.coordinate)
                if not provider_cell.comment or "Provider consensus source:" not in provider_cell.comment.text:
                    _finding(findings, "CONSENSUS_PROVIDER_LINEAGE", cell=provider_cell.coordinate)
            selected_formula = _normal_formula(selected_cell.value)
            if f"{column}{model_row}" not in _formula_refs(selected_cell.value):
                _finding(findings, "CONSENSUS_SELECTED_MISSING_MODEL", cell=selected_cell.coordinate)
            if provider_row and f"{column}{provider_row}" not in _formula_refs(selected_cell.value):
                _finding(findings, "CONSENSUS_SELECTED_MISSING_PROVIDER", cell=selected_cell.coordinate)
            if not selected_formula.startswith("=IF("):
                _finding(findings, "CONSENSUS_SELECTED_NOT_FORMULA", cell=selected_cell.coordinate)
            target = f"BROKERS!{column}{selected_row}"
            links = [
                cell for cell in operating_formula_cells
                if target in _normal_formula(cell.value)
            ]
            for link in links:
                if _colour(link) != GREEN:
                    _finding(findings, "CONSENSUS_CROSS_SHEET_LINK_NOT_GREEN", cell=link.coordinate)
                green_links_checked += 1
            checked += 1
    if not checked:
        _finding(findings, "CONSENSUS_BLOCKS_ABSENT")
    if not green_links_checked:
        _finding(findings, "CONSENSUS_CROSS_SHEET_LINK_MISSING")
    return {
        "consensus_periods_checked": checked,
        "consensus_metrics_checked": len(model_rows),
        "consensus_green_links_checked": green_links_checked,
    }


def inspect_forecast_ownership(receipt_path: Path, findings: list[dict]) -> dict:
    entries = json.loads(receipt_path.read_text(encoding="utf-8"))
    keys: set[tuple[str, int]] = set()
    for entry in entries:
        key = (str(entry.get("display_id")), int(entry.get("forecast_index", -1)))
        if key in keys:
            _finding(findings, "FORECAST_OWNERSHIP_DUPLICATE_WRITER", display_id=key[0], forecast_index=key[1])
        keys.add(key)
        if entry.get("status") != "resolved":
            _finding(findings, "FORECAST_OWNERSHIP_UNRESOLVED", display_id=key[0], forecast_index=key[1])
        if entry.get("captured_by") and entry.get("method") != "not_separately_forecast":
            _finding(findings, "FORECAST_CAPTURE_LIVE_WRITER", display_id=key[0], forecast_index=key[1])
    return {"forecast_ownership_entries_checked": len(entries)}


def inspect_protected_cash_and_rcf(workbook_path: Path, contract_path: Path, findings: list[dict]) -> dict:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    contract = json.loads(contract_path.read_text(encoding="utf-8"))
    sheet = workbook[contract["statement_sheet"]]
    checked = 0
    for identity in contract.get("protected_formula_identities") or []:
        for column in identity.get("columns") or []:
            owner = f"{column}{identity['owner_row']}"
            required = {f"{column}{row}" for row in identity.get("member_rows") or []}
            if not required.issubset(_formula_refs(sheet[owner].value)):
                _finding(findings, "PROTECTED_CASH_IDENTITY", concept=identity.get("concept_id"), cell=owner)
            checked += 1
    captures = contract.get("semantic_scope_captures") or []
    rcf = {entry.get("child_label") for entry in captures if str(entry.get("child_label", "")).startswith("RCF ")}
    if rcf != {"RCF draw", "RCF repayment"}:
        _finding(findings, "RCF_CAPTURE_CONTRACT", captured=sorted(rcf))
    financing = next((entry for entry in contract.get("protected_formula_identities") or [] if entry.get("concept_id") == "cash_from_financing"), None)
    change_row = next((entry.get("parent_row") for entry in captures if entry.get("child_label") == "RCF draw"), None)
    if not financing or change_row not in (financing.get("member_rows") or []):
        _finding(findings, "RCF_FINANCING_INCLUSION")
    return {"protected_cash_periods_checked": checked, "rcf_captures_checked": len(rcf)}


def inspect_debt_and_interest(workbook_path: Path, model_case_path: Path, findings: list[dict]) -> dict:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    sheet = workbook["Operating Model"]
    model_case = json.loads(model_case_path.read_text(encoding="utf-8"))
    labels = _labels(sheet)
    group_ranges = {
        "bond_fixed": (_first(labels, "Bonds"), _first(labels, "Total Bonds")),
        "term_loan_floating": (_first(labels, "Bank Debt"), _first(labels, "Total Bank Debt")),
        "securitisation_floating": (_first(labels, "Bank Debt"), _first(labels, "Total Bank Debt")),
        "commercial_paper_floating": (_first(labels, "Other Debt"), _first(labels, "Total Other Debt")),
        "manual_all_in": (_first(labels, "Other Debt"), _first(labels, "Total Other Debt")),
        "rcf_floating": (_first(labels, "Bank Debt"), _first(labels, "Total Bank Debt")),
    }
    interest_heading = _one(labels, "7. INTEREST SCHEDULE")
    balance_by_label = {}
    for row in range(1, interest_heading):
        label = sheet[f"B{row}"].value
        if isinstance(label, str) and _finite(sheet[f"D{row}"].value):
            balance_by_label.setdefault(label, row)
    checked = 0
    for row in range(interest_heading + 1, sheet.max_row + 1):
        label = sheet[f"B{row}"].value
        if not isinstance(label, str) or label not in balance_by_label:
            continue
        balance_row = balance_by_label[label]
        for index, column in enumerate(FORECAST_COLUMNS):
            prior_column = ("D", "J", "K")[index]
            refs = _formula_refs(sheet[f"{column}{row}"].value)
            if f"{prior_column}{balance_row}" not in refs or f"D{row}" not in refs:
                _finding(findings, "INTEREST_LINEAGE", instrument=label, cell=f"{column}{row}")
            checked += 1
    if not checked:
        _finding(findings, "INTEREST_LINEAGE_ABSENT")

    classified = 0
    for instrument in model_case.get("instruments") or []:
        debt_class = instrument.get("class")
        if debt_class not in group_ranges or debt_class == "rcf_floating":
            continue
        candidates = [
            row for row in range(1, sheet.max_row + 1)
            if sheet[f"C{row}"].value == instrument.get("currency")
            and _finite(sheet[f"D{row}"].value)
            and abs(float(sheet[f"D{row}"].value) - float(instrument.get("opening_balance", 0))) < 1e-7
        ]
        start, end = group_ranges[debt_class]
        if len(candidates) != 1 or not (start < candidates[0] < end):
            _finding(findings, "DEBT_CLASSIFICATION", instrument_id=instrument.get("instrument_id"), debt_class=debt_class)
        classified += 1
    return {"interest_periods_checked": checked, "debt_instruments_classified": classified}


def inspect_acquisition_and_lease(workbook_path: Path, model_case_path: Path, findings: list[dict]) -> dict:
    formulas = load_workbook(workbook_path, data_only=False, read_only=False)
    values = load_workbook(workbook_path, data_only=True, read_only=False)
    sheet, cached = formulas["Operating Model"], values["Operating Model"]
    labels = _labels(sheet)
    model_case = json.loads(model_case_path.read_text(encoding="utf-8"))
    acquisition = model_case.get("acquisition") or {}
    years = [int(str(period["date"])[:4]) for period in model_case.get("periods") or [] if period.get("status") == "forecast"]
    close_index = years.index(int(acquisition["close_year"]))
    column = ADJUSTMENT_COLUMNS[close_index]
    consideration_row = _one(labels, "Acquisitions, net of cash acquired")
    debt_row = _one(labels, "Acquisition term debt")
    interest_row = _one(labels, "Acquisition debt interest")
    expected = {
        f"{column}{consideration_row}": -float(acquisition["transaction_enterprise_value"]),
        f"{column}{debt_row}": float(acquisition["acquisition_debt_amount"]),
    }
    for address, expected_value in expected.items():
        actual = cached[address].value
        if not _finite(actual) or abs(float(actual) - expected_value) > 1e-6:
            _finding(findings, "ACQUISITION_CASH_DEBT", cell=address, expected=expected_value, actual=actual)
    interest_address = f"{column}{interest_row}"
    interest_value = cached[interest_address].value
    interest_formula = _formula_refs(sheet[interest_address].value)
    if float(acquisition["acquisition_debt_amount"]) > 0 and (not _finite(interest_value) or float(interest_value) >= 0):
        _finding(findings, "ACQUISITION_INTEREST", cell=interest_address, actual=interest_value)
    if "P8" not in interest_formula or "P9" not in interest_formula:
        _finding(findings, "ACQUISITION_INTEREST_LINEAGE", cell=interest_address)

    ending_row = _one(labels, "Lease liabilities — after assumed new lease additions")
    principal_row = _one(labels, "Lease principal repayment assumption")
    additions_row = _one(labels, "New lease additions assumption")
    other_row = _one(labels, "Other lease liability movements assumption")
    lease_interest_row = _one(labels, "Lease interest")
    lease_checked = 0
    for index, forecast_column in enumerate(FORECAST_COLUMNS):
        prior_column = ("I", "J", "K")[index]
        ending = cached[f"{forecast_column}{ending_row}"].value
        prior = cached[f"{prior_column}{ending_row}"].value
        principal = cached[f"{forecast_column}{principal_row}"].value
        additions = cached[f"{forecast_column}{additions_row}"].value
        other = cached[f"{forecast_column}{other_row}"].value
        interest = cached[f"{forecast_column}{lease_interest_row}"].value
        numbers = (ending, prior, principal, additions, other, interest)
        if not all(_finite(value) for value in numbers) or abs(float(ending) - (float(prior) + float(additions) + float(other) - float(principal) - float(interest))) > 1e-5:
            _finding(findings, "LEASE_ROLL_FORWARD", cell=f"{forecast_column}{ending_row}")
        ending_refs = _formula_refs(sheet[f"{forecast_column}{ending_row}"].value)
        required = {
            f"{prior_column}{ending_row}", f"{forecast_column}{principal_row}",
            f"{forecast_column}{additions_row}", f"{forecast_column}{other_row}",
        }
        if not required.issubset(ending_refs):
            _finding(findings, "LEASE_ROLL_FORWARD_LINEAGE", cell=f"{forecast_column}{ending_row}")
        lease_checked += 1
    return {"acquisition_periods_checked": 1, "lease_periods_checked": lease_checked}


def inspect_source_arithmetic(artifact_path: Path, findings: list[dict]) -> dict:
    artifact = json.loads(artifact_path.read_text(encoding="utf-8"))
    declared = artifact.get("artifact_sha256")
    body = {key: value for key, value in artifact.items() if key != "artifact_sha256"}
    expected = hashlib.sha256(
        (json.dumps(body, sort_keys=True, separators=(",", ":")) + "\n").encode()
    ).hexdigest()
    if declared != expected or not artifact.get("source_model_case_sha256"):
        _finding(findings, "SOURCE_ARITHMETIC_SEAL")
    rows = {str(row["source_line_id"]): row for row in artifact.get("rows") or []}
    dimensions = ("numeric_type", "reporting_currency", "units", "sign_convention")
    families: dict[str, list[dict]] = defaultdict(list)
    for row in rows.values():
        parent_id = row.get("parent_source_line_id")
        if parent_id:
            families[str(parent_id)].append(row)
    for parent_id, children in families.items():
        parent = rows.get(parent_id)
        if parent is None:
            _finding(findings, "SOURCE_ARITHMETIC_PARENT_ABSENT", parent_id=parent_id)
            continue
        for child in children:
            for dimension in dimensions:
                if child.get(dimension) != parent.get(dimension):
                    _finding(findings, "SOURCE_ARITHMETIC_DIMENSION", parent_id=parent_id, child_id=child.get("source_line_id"), dimension=dimension)
        for period in range(3):
            child_total = sum(float(child["values"][period]) for child in children)
            if abs(child_total - float(parent["values"][period])) > 1e-7:
                _finding(findings, "SOURCE_ARITHMETIC_VALUE", parent_id=parent_id, period=period)
    if not families:
        _finding(findings, "SOURCE_ARITHMETIC_FAMILY_ABSENT")
    return {"source_arithmetic_families_checked": len(families)}


def inspect_emitted_candidate(
    *, workbook_path: Path, model_case_path: Path, forecast_receipt_path: Path,
    proof_contract_path: Path, source_arithmetic_path: Path,
) -> dict:
    findings: list[dict] = []
    coverage = {}
    coverage.update(inspect_consensus(workbook_path, findings))
    coverage.update(inspect_forecast_ownership(forecast_receipt_path, findings))
    coverage.update(inspect_protected_cash_and_rcf(workbook_path, proof_contract_path, findings))
    coverage.update(inspect_debt_and_interest(workbook_path, model_case_path, findings))
    coverage.update(inspect_acquisition_and_lease(workbook_path, model_case_path, findings))
    coverage.update(inspect_source_arithmetic(source_arithmetic_path, findings))
    return {
        "schema_version": "emitted-candidate-independent-oracle/1.0",
        "status": "PASS" if not findings else "BLOCK",
        "coverage": coverage,
        "findings": findings,
        "total_violations": len(findings),
        "production_imports": production_imports(Path(__file__), FORBIDDEN_PRODUCTION_IMPORTS),
        "artifacts_consumed": [
            "emitted_workbook", "public_model_case", "forecast_receipt",
            "workbook_proof_contract", "source_arithmetic_artifact",
        ],
    }
