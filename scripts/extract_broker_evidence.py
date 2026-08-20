#!/usr/bin/env python3
"""Extract broker documents into a hash-bound, cell-level evidence bundle.

The extractor is deliberately conservative. Native PDF/XLSX/CSV structure is
captured deterministically. Pages whose numeric content cannot be reconstructed
confidently are rendered and emitted as explicit vision tasks; they are never
silently treated as complete.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import math
import os
import re
import signal
import sys
import time
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

from broker_terminal_recovery import compile_broker_demand_contract, normalized_label


VERSION = "broker-evidence-extractor/1.1"
PDF_LANE_MIN_SECONDS = 2.0
PDF_LANE_MAX_SECONDS = 15.0
MAX_REGION_CROPS_PER_PAGE = int(os.environ.get("BROKER_MAX_REGION_CROPS_PER_PAGE", "12"))
NUMERIC_RE = re.compile(
    # Both boundaries exclude letters *and digits*.  The old right boundary
    # excluded letters only, so the engine could backtrack through ``2026E``
    # and count the prefix ``202`` as a missing table value.  That turned a
    # complete ruled forecast table into unnecessary OCR work.
    r"(?<![A-Za-z0-9])(?:\(?[-+]?[$€£¥]?\s*(?:\d{1,3}(?:[, ]\d{3})+|\d+)(?:\.\d+)?%?\)?|[-+]?\d+(?:\.\d+)?x)(?![A-Za-z0-9])",
)
SUPPORTED = {
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroEnabled.12",
    "text/csv",
    "text/plain",
    "text/markdown",
    "text/html",
    "application/json",
    "image/png",
    "image/jpeg",
}


def demand_targets_for_surface(
    text: str,
    tables: list[dict[str, Any]],
    demand_contract: dict[str, Any],
    *,
    opaque_image: bool = False,
) -> list[dict[str, Any]]:
    """Return only model-demand concepts plausibly visible on one surface.

    Native text and table labels are discovery evidence only.  An opaque image
    cannot be searched safely, so it receives the same bounded target list,
    but never a whole-page table-transcription instruction.
    """
    targets = list(demand_contract.get("targets") or [])
    if opaque_image:
        return targets
    haystack = f" {normalized_label(text)} "
    table_labels = {
        normalized_label(cell.get("raw_text"))
        for table in tables
        for row in table.get("rows") or []
        for cell in row[:2]
        if normalized_label(cell.get("raw_text"))
    }
    selected = []
    for target in targets:
        aliases = {
            normalized_label(value)
            for value in target.get("discovery_aliases") or []
            if normalized_label(value)
        }
        if any(alias in table_labels or f" {alias} " in haystack for alias in aliases):
            selected.append(target)
    return selected


def selected_cell_instruction(targets: list[dict[str, Any]]) -> str:
    concepts = ", ".join(
        f"{item['metric_id']} ({'/'.join(item.get('demand_labels') or [])})"
        for item in targets
    )
    periods = ", ".join(
        str(value) for value in ((targets[0].get("forecast_periods") or []) if targets else [])
    )
    return (
        f"Recover only the requested model-demand cells: {concepts}. Required periods: {periods}. "
        "For each visible requested concept, return its exact printed row label, exact period headers, "
        "units and only the requested period values in a small grid with cell bboxes. Preserve blanks, "
        "dashes, parentheses, signs and footnote markers. Do not transcribe unrelated rows, tables, "
        "valuation data, narrative numbers or a whole-page numeric inventory. If none of the requested "
        "concepts is visible, return no tables and set surface_disposition to verified_non_tabular with "
        "non_tabular_reason 'no requested model-demand concept visible'. Each pass decides independently."
    )


def normalise_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def pdf_lane_timeout_budget(
    *, page_count: int, word_count: int, target_count: int, lane_id: str
) -> float:
    """Give legitimate native table discovery time in proportion to evidence.

    The former fixed 0.5 second alarm routinely abandoned valid ruled tables
    on dense or multi-page reports. The budget remains finite and per-lane, but
    grows with document size, page density and the bounded model-demand target
    set. More expensive discovery strategies receive a small explicit factor.
    """
    lane_factor = {"lines": 1.0, "lines_strict": 1.2, "text": 1.4}.get(lane_id, 1.0)
    evidence_seconds = (
        1.5
        + min(max(0, page_count), 200) / 80.0
        + min(max(0, word_count), 3_000) / 600.0
        + min(max(0, target_count), 20) * 0.20
    ) * lane_factor
    return min(PDF_LANE_MAX_SECONDS, max(PDF_LANE_MIN_SECONDS, evidence_seconds))


def pdf_document_table_budget(*, page_count: int, target_count: int) -> float:
    return min(300.0, max(30.0, page_count * 2.5 + min(max(0, target_count), 20) * 3.0))


def table_discovery_lane_summary(
    lane_id: str, *, candidate_count: int = 0, error: Exception | None = None
) -> dict[str, Any]:
    """Return only the evidence-contract surface for one discovery lane.

    Timeout budgets and elapsed time are controller telemetry. They do not
    belong in the immutable extraction bundle, whose closed schema deliberately
    admits only discovery outcome and an optional diagnostic message.
    """
    if error is None:
        return {
            "lane_id": lane_id,
            "status": "pass" if candidate_count else "empty",
            "candidate_count": candidate_count,
        }
    prefix = "timeout: " if isinstance(error, TimeoutError) else ""
    return {
        "lane_id": lane_id,
        "status": "error",
        "candidate_count": 0,
        "message": f"{prefix}{error}",
    }


def bounded_table_find(page: Any, kwargs: dict[str, Any], timeout_seconds: float) -> Any:
    """Bound pathological PDF table discovery and fall through to rendering."""
    if not hasattr(signal, "setitimer") or timeout_seconds <= 0:
        return page.find_tables(**kwargs)

    def timed_out(_signum: int, _frame: Any) -> None:
        raise TimeoutError(
            f"PDF table lane exceeded {timeout_seconds:g}s evidence-sized budget"
        )

    previous = signal.signal(signal.SIGALRM, timed_out)
    try:
        signal.setitimer(signal.ITIMER_REAL, timeout_seconds)
        return page.find_tables(**kwargs)
    finally:
        signal.setitimer(signal.ITIMER_REAL, 0)
        signal.signal(signal.SIGALRM, previous)


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_json(value: Any) -> bytes:
    return (json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False) + "\n").encode("utf-8")


def safe_id(value: str) -> str:
    cleaned = re.sub(r"[^a-z0-9_.-]+", "-", str(value).lower()).strip("-.")
    return cleaned or "item"


def normalise_numeric_token(value: Any) -> str | None:
    if value is None or isinstance(value, bool):
        return None
    text = str(value).strip()
    if not text or text in {"-", "–", "—"}:
        return None
    match = NUMERIC_RE.fullmatch(text)
    if not match:
        return None
    negative = text.startswith("(") and text.endswith(")")
    percent = "%" in text
    multiple = text.lower().endswith("x")
    cleaned = text.replace(",", "").replace(" ", "")
    cleaned = re.sub(r"[$€£¥%xX()]", "", cleaned)
    try:
        number = float(cleaned)
    except ValueError:
        return text
    if negative:
        number = -abs(number)
    if number == 0:
        number = 0.0
    formatted = str(int(number)) if float(number).is_integer() else format(number, ".15g")
    if percent:
        formatted += "%"
    if multiple:
        formatted += "x"
    return formatted


def numeric_tokens(text: str) -> list[str]:
    result: list[str] = []
    for match in NUMERIC_RE.finditer(text or ""):
        token = normalise_numeric_token(match.group(0))
        if token is not None:
            result.append(token)
    return result


def counter_difference(left: Iterable[str], right: Iterable[str]) -> list[str]:
    difference = Counter(left) - Counter(right)
    return sorted(token for token, count in difference.items() for _ in range(count))


def value_kind(value: Any, formula: str | None = None) -> str:
    if formula:
        return "formula"
    if value is None:
        return "blank"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (dt.date, dt.datetime, dt.time)):
        return "date"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "number"
    return "text"


def serialise_value(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return str(value)
    return value


class ArtifactWriter:
    def __init__(self, root: Path, document_id: str):
        self.root = root
        self.document_id = safe_id(document_id)
        self.document_root = root / "artifacts" / self.document_id
        self.document_root.mkdir(parents=True, exist_ok=True)
        self.records: list[dict[str, Any]] = []

    def write(self, name: str, kind: str, data: bytes) -> str:
        path = self.document_root / name
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(data)
        relative = path.relative_to(self.root).as_posix()
        artifact_id = f"{self.document_id}.{safe_id(name)}"
        self.records.append({
            "artifact_id": artifact_id,
            "kind": kind,
            "path": relative,
            "sha256": sha256_bytes(data),
        })
        return artifact_id

    def write_json(self, name: str, kind: str, value: Any) -> str:
        return self.write(name, kind, canonical_json(value))

    def write_text(self, name: str, kind: str, value: str) -> str:
        return self.write(name, kind, value.encode("utf-8"))


def cell_record(row: int, column: int, raw: Any, source_ref: str, *, formula: str | None = None,
                number_format: str | None = None, bbox: list[float] | None = None,
                confidence: float | None = 1.0) -> dict[str, Any]:
    value = serialise_value(raw)
    return {
        "row": row,
        "column": column,
        "raw_text": None if raw is None else str(serialise_value(raw)),
        "value": value,
        "value_kind": value_kind(raw, formula),
        "number_format": number_format,
        "formula": formula,
        "bbox": bbox,
        "source_ref": source_ref,
        "confidence": confidence,
    }


def table_tokens(table: dict[str, Any]) -> list[str]:
    tokens: list[str] = []
    for row in table["rows"]:
        for cell in row:
            token = normalise_numeric_token(cell.get("raw_text"))
            if token is not None:
                tokens.append(token)
    return tokens


def pdf_words_in_bbox(words: list[list[Any]], bbox: tuple[float, float, float, float]) -> list[str]:
    x0, y0, x1, y1 = bbox
    values: list[str] = []
    for word in words:
        wx0, wy0, wx1, wy1 = map(float, word[:4])
        cx, cy = (wx0 + wx1) / 2, (wy0 + wy1) / 2
        if x0 <= cx <= x1 and y0 <= cy <= y1:
            values.extend(numeric_tokens(str(word[4])))
    return values


def bbox_iou(left: list[float] | tuple[float, ...] | None,
             right: list[float] | tuple[float, ...] | None) -> float:
    """Return intersection-over-union for two page-space rectangles."""
    if not left or not right or len(left) != 4 or len(right) != 4:
        return 0.0
    lx0, ly0, lx1, ly1 = map(float, left)
    rx0, ry0, rx1, ry1 = map(float, right)
    width = max(0.0, min(lx1, rx1) - max(lx0, rx0))
    height = max(0.0, min(ly1, ry1) - max(ly0, ry0))
    intersection = width * height
    union = max(0.0, (lx1 - lx0) * (ly1 - ly0)) + max(0.0, (rx1 - rx0) * (ry1 - ry0)) - intersection
    return 0.0 if union <= 0 else intersection / union


def word_numeric_records(words: list[list[Any]]) -> list[dict[str, Any]]:
    """Create the source-owned whole-surface numeric denominator.

    These records are generated before table discovery.  Later table or vision
    lanes may account for them, but may never create or remove them.
    """
    records: list[dict[str, Any]] = []
    for word_index, word in enumerate(words):
        raw = str(word[4])
        for token_index, token in enumerate(numeric_tokens(raw)):
            records.append({
                "token_id": f"w{word_index + 1}.n{token_index + 1}",
                "token": token,
                "raw_text": raw,
                "bbox": [float(value) for value in word[:4]],
                "block": int(word[5]),
                "line": int(word[6]),
                "word": int(word[7]),
            })
    return records


STRICT_PERIOD_RE = re.compile(
    r"^(?:"
    r"(?:fy|cy)\s*(?:19|20)?\d{2}[eaf]?"
    r"|(?:19|20)\d{2}(?:\s*/\s*\d{2,4})?[eaf]?"
    r"|(?:q[1-4]|[1-4]q|h[12]|[12]h)\s*(?:fy|cy)?\s*(?:19|20)?\d{2}[eaf]?"
    r"|\d{1,2}\s*/\s*\d{2}[eaf]?"
    r"|[a-z]{3,9}[-\s]\d{2,4}[eaf]?"
    r"|ltm|ttm|ntm"
    r")$",
    re.IGNORECASE,
)


def reconstruct_columnar_tables(words: list[list[Any]],
                                excluded_bboxes: list[list[float]]) -> list[dict[str, Any]]:
    """Rebuild unruled financial tables from the text layer's own geometry.

    Broker reports are born-digital: the publisher draws almost no ruling
    lines, so the ruled discovery lanes find nothing, and a fifty-row Key
    financials appendix used to fall through to an LLM vision pass - the least
    deterministic tool in the box doing the most deterministic part of the
    job. But the text layer carries exact glyph coordinates, and a financial
    table is the easiest structure-recognition case there is: numbers
    right-aligned in tidy column bands, period headings across the top, labels
    flush left. That is x-coordinate clustering, not perception.

    The lane claims the same native, deterministic authority class as the
    ruled lanes - single-read ownership - because it is the same evidence
    (the page's own text) read by a different geometry. What EARNS that
    authority is self-verification: a reconstruction is emitted only when its
    own structure checks pass, and it demotes itself silently otherwise,
    leaving the page to the vision path. It never touches regions the ruled
    lanes already own.
    """
    ROW_TOLERANCE = 3.0      # points: words within this y-band share a row
    COLUMN_GAP = 12.0        # points: a wider gap between right edges opens a new band
    MIN_DATA_ROWS = 2        # a grid needs a header row AND data
    MIN_COLUMNS = 2

    def excluded(box: list[float]) -> bool:
        return any(bbox_iou(box, other) >= 0.25 for other in excluded_bboxes)

    # ------------------------------------------------------------- rows
    rows: list[dict[str, Any]] = []
    for word in sorted(words, key=lambda item: ((item[1] + item[3]) / 2.0, item[0])):
        box = [float(value) for value in word[:4]]
        if excluded(box):
            continue
        centre = (box[1] + box[3]) / 2.0
        text = str(word[4]).strip()
        if not text:
            continue
        if rows and abs(centre - rows[-1]["centre"]) <= ROW_TOLERANCE:
            rows[-1]["words"].append((box, text))
            rows[-1]["centre"] = (rows[-1]["centre"] + centre) / 2.0
        else:
            rows.append({"centre": centre, "words": [(box, text)]})

    def is_numeric(text: str) -> bool:
        return normalise_numeric_token(text) is not None

    def is_period(text: str) -> bool:
        return bool(STRICT_PERIOD_RE.fullmatch(re.sub(r"\s+", " ", text.strip())))

    # A table row carries at least two numeric or period tokens; a run of them
    # bounded by non-table rows is one candidate region.
    def row_kind(row: dict[str, Any]) -> str:
        values = sum(1 for _, text in row["words"] if is_numeric(text))
        periods = sum(1 for _, text in row["words"] if is_period(text))
        if periods >= 2 and periods >= values:
            return "header"
        if values >= 2:
            return "data"
        return "other"

    regions: list[list[dict[str, Any]]] = []
    current: list[dict[str, Any]] = []
    gap = 0
    for row in rows:
        kind = row_kind(row)
        if kind in ("header", "data"):
            current.append({**row, "kind": kind})
            gap = 0
        elif current:
            # One quiet row (a section title inside the table) is tolerated;
            # a second closes the region.
            gap += 1
            if gap >= 2:
                regions.append(current)
                current = []
                gap = 0
    if current:
        regions.append(current)

    tables: list[dict[str, Any]] = []
    for region in regions:
        header_rows = [row for row in region if row["kind"] == "header"]
        data_rows = [row for row in region if row["kind"] == "data"]
        if not header_rows or len(data_rows) < MIN_DATA_ROWS:
            continue
        header = header_rows[0]

        # -------------------------------------------------------- columns
        # Column bands from the RIGHT edges of the value/period tokens: a
        # financial column is right-aligned, so right edges stack. Bands are
        # built over the whole region, then each token must land in exactly
        # one band - that is the self-check that earns ownership.
        edges = sorted(
            box[2]
            for row in region
            for box, text in row["words"]
            if is_numeric(text) or is_period(text)
        )
        if not edges:
            continue
        bands: list[list[float]] = [[edges[0], edges[0]]]
        for edge in edges[1:]:
            if edge - bands[-1][1] <= COLUMN_GAP:
                bands[-1][1] = edge
            else:
                bands.append([edge, edge])
        if len(bands) < MIN_COLUMNS:
            continue

        def band_of(box: list[float]) -> int | None:
            hits = [
                index for index, (low, high) in enumerate(bands)
                if low - COLUMN_GAP / 2 <= box[2] <= high + COLUMN_GAP / 2
            ]
            return hits[0] if len(hits) == 1 else None

        # ---------------------------------------------- self-verification
        assignable = True
        for row in region:
            for box, text in row["words"]:
                if (is_numeric(text) or is_period(text)) and band_of(box) is None:
                    assignable = False
        if not assignable:
            continue

        grid_rows: list[list[str | None]] = []
        region_box = [
            min(box[0] for row in region for box, _ in row["words"]),
            min(box[1] for row in region for box, _ in row["words"]),
            max(box[2] for row in region for box, _ in row["words"]),
            max(box[3] for row in region for box, _ in row["words"]),
        ]
        for row in region:
            cells: list[str | None] = [None] * (1 + len(bands))
            label_parts: list[str] = []
            collision = False
            for box, text in sorted(row["words"], key=lambda item: item[0][0]):
                if is_numeric(text) or is_period(text):
                    band = band_of(box)
                    if cells[1 + band] is not None:
                        collision = True
                    cells[1 + band] = text
                else:
                    label_parts.append(text)
            if collision:
                assignable = False
                break
            cells[0] = " ".join(label_parts) or None
            grid_rows.append(cells)
        if not assignable or len(grid_rows) < 1 + MIN_DATA_ROWS:
            continue
        labelled = sum(1 for cells in grid_rows if cells[0] and not is_period(cells[0]))
        if labelled < MIN_DATA_ROWS:
            continue
        tables.append({
            "region_bbox": region_box,
            "rows": grid_rows,
            "column_count": 1 + len(bands),
        })
    return tables


def record_inside_any_bbox(record: dict[str, Any], bboxes: list[list[float]]) -> bool:
    x0, y0, x1, y1 = map(float, record["bbox"])
    cx, cy = (x0 + x1) / 2.0, (y0 + y1) / 2.0
    return any(float(box[0]) <= cx <= float(box[2]) and float(box[1]) <= cy <= float(box[3]) for box in bboxes)


def uncovered_numeric_regions(records: list[dict[str, Any]], bboxes: list[list[float]]) -> list[dict[str, Any]]:
    """Cluster source numbers not covered by a bounded table rectangle.

    A single narrative sentence often contains two or three numbers.  That is
    not a missing table.  An uncovered line is material only when it belongs
    to a nearby multi-line numeric grid with at least two aligned numeric
    columns.  Broad native-text table candidates are handled separately as
    discovery hints and cannot turn whole-page prose into a table denominator.
    """
    uncovered = [record for record in records if not record_inside_any_bbox(record, bboxes)]
    by_line: dict[tuple[int, int], list[dict[str, Any]]] = {}
    for record in uncovered:
        by_line.setdefault((record["block"], record["line"]), []).append(record)
    line_items = []
    for key, items in by_line.items():
        items.sort(key=lambda item: (item["bbox"][0], item["token_id"]))
        line_items.append((key, items, sum((item["bbox"][1] + item["bbox"][3]) / 2.0 for item in items) / len(items)))

    def aligned(left: list[dict[str, Any]], right: list[dict[str, Any]]) -> bool:
        if len(left) < 2 or len(right) < 2:
            return False
        left_x = [(item["bbox"][0] + item["bbox"][2]) / 2.0 for item in left]
        right_x = [(item["bbox"][0] + item["bbox"][2]) / 2.0 for item in right]
        matches = sum(1 for x in left_x if any(abs(x - other) <= 24.0 for other in right_x))
        return matches >= 2

    regions: list[dict[str, Any]] = []
    for region_index, ((block, line), items, y_mid) in enumerate(sorted(line_items), start=1):
        material = any(
            other_key != (block, line)
            and abs(other_y - y_mid) <= 54.0
            and aligned(items, other_items)
            for other_key, other_items, other_y in line_items
        )
        bbox = [
            min(item["bbox"][0] for item in items),
            min(item["bbox"][1] for item in items),
            max(item["bbox"][2] for item in items),
            max(item["bbox"][3] for item in items),
        ]
        regions.append({
            "region_id": f"uncovered-{region_index}",
            "block": block,
            "line": line,
            "bbox": bbox,
            "token_ids": [item["token_id"] for item in items],
            "tokens": [item["token"] for item in items],
            "material": material,
            "disposition": "requires_vision" if material else "evidence_only_narrative_or_singleton",
        })
    return regions


def table_context(words: list[list[Any]], bbox: list[float]) -> tuple[str | None, str | None, list[str]]:
    """Preserve nearby caption, unit and footnote text without interpreting metrics."""
    x0, y0, x1, y1 = map(float, bbox)
    above: list[tuple[float, str]] = []
    below: list[tuple[float, str]] = []
    for word in words:
        wx0, wy0, wx1, wy1 = map(float, word[:4])
        if wx1 < x0 - 20 or wx0 > x1 + 20:
            continue
        text = str(word[4]).strip()
        if not text:
            continue
        if 0 <= y0 - wy1 <= 42:
            above.append((wy0, text))
        elif 0 <= wy0 - y1 <= 48:
            below.append((wy0, text))
    caption = " ".join(text for _, text in sorted(above)) or None
    footnote_text = " ".join(text for _, text in sorted(below))
    footnotes = [footnote_text] if footnote_text else []
    unit_match = re.search(r"(?:in\s+)?(?:[$€£¥]|usd|eur|gbp)?\s*(?:m|mn|mm|bn|billions?|millions?|thousands?|%|x)\b", caption or "", re.I)
    units = unit_match.group(0).strip() if unit_match else None
    return caption, units, footnotes


def native_cell_bbox(native_table: Any, row_index: int, column_index: int) -> list[float] | None:
    try:
        row = native_table.rows[row_index - 1]
        bbox = row.cells[column_index - 1]
        return None if bbox is None else [float(value) for value in bbox]
    except Exception:
        return None


def extract_pdf(path: Path, descriptor: dict[str, Any], writer: ArtifactWriter,
                render_dpi: int, demand_contract: dict[str, Any],
                vision_house_id: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str], list[str], str]:
    try:
        import fitz  # type: ignore
    except Exception as error:  # pragma: no cover - environment-specific
        raise RuntimeError(f"PyMuPDF is required for PDF extraction: {error}") from error

    document = fitz.open(path)
    if document.needs_pass:
        password = descriptor.get("password") or ""
        if not document.authenticate(password):
            raise RuntimeError("PDF is encrypted and the supplied password did not authenticate.")

    surfaces: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    source_table_tokens: list[str] = []
    captured_table_tokens: list[str] = []
    unresolved = False
    seen_images: set[int] = set()
    table_budget_seconds = pdf_document_table_budget(
        page_count=document.page_count,
        target_count=len(demand_contract.get("targets") or []),
    )
    table_budget_started = time.monotonic()

    for page_index in range(document.page_count):
        page = document.load_page(page_index)
        surface_id = f"{descriptor['document_id']}.p{page_index + 1}"
        fallback_text = page.get_text("text", sort=True) or ""
        selected_targets = demand_targets_for_surface(
            fallback_text,
            [],
            demand_contract,
            opaque_image=not bool(fallback_text.strip()),
        )
        if descriptor.get("house_id") != vision_house_id:
            selected_targets = []
        text = page.get_text("text", sort=True) or ""
        words = page.get_text("words", sort=True) or []
        text_ref = writer.write_text(f"pages/page-{page_index + 1:04d}.txt", "native_text", text)
        word_payload = [
            {"bbox": [float(v) for v in word[:4]], "text": str(word[4]), "block": int(word[5]), "line": int(word[6]), "word": int(word[7])}
            for word in words
        ]
        words_ref = writer.write_json(f"pages/page-{page_index + 1:04d}.words.json", "word_geometry", word_payload)

        # Whole-page records are an immutable discovery census, not the table
        # denominator.  Narrative, valuation and disclosure prose must remain
        # available as evidence without being forced into broker table sheets.
        source_numeric_records = word_numeric_records(words)

        discovery_specs = [
            ("lines", {}),
            ("lines_strict", {"vertical_strategy": "lines_strict", "horizontal_strategy": "lines_strict"}),
            ("text", {"vertical_strategy": "text", "horizontal_strategy": "text", "min_words_vertical": 2, "min_words_horizontal": 1}),
        ]
        page_tables: list[tuple[str, Any]] = []
        lane_summaries: list[dict[str, Any]] = []
        for lane_id, kwargs in discovery_specs:
            lane_timeout_seconds = pdf_lane_timeout_budget(
                page_count=document.page_count,
                word_count=len(words),
                target_count=len(selected_targets),
                lane_id=lane_id,
            )
            try:
                if time.monotonic() - table_budget_started > table_budget_seconds:
                    raise TimeoutError(
                        f"document table-discovery budget {table_budget_seconds:g}s exhausted"
                    )
                finder = bounded_table_find(page, kwargs, lane_timeout_seconds)
                found = list(getattr(finder, "tables", []) or [])
                lane_summaries.append(table_discovery_lane_summary(
                    lane_id, candidate_count=len(found)
                ))
                page_tables.extend((lane_id, item) for item in found)
            except Exception as error:
                lane_summaries.append(table_discovery_lane_summary(
                    lane_id, error=error
                ))

        table_bboxes: list[list[float]] = []
        text_discovery_bboxes: list[list[float]] = []
        lane_tables: dict[str, list[dict[str, Any]]] = {lane_id: [] for lane_id, _ in discovery_specs}
        for table_index, (lane_id, native_table) in enumerate(page_tables, start=1):
            extracted = native_table.extract() or []
            if not extracted:
                continue
            table_id = f"{surface_id}.{lane_id}-t{table_index}"
            bbox = [float(value) for value in native_table.bbox]
            if lane_id == "text":
                text_discovery_bboxes.append(bbox)
            else:
                table_bboxes.append(bbox)
            rows: list[list[dict[str, Any]]] = []
            for row_index, row in enumerate(extracted, start=1):
                cells: list[dict[str, Any]] = []
                for column_index, value in enumerate(row, start=1):
                    cells.append(cell_record(
                        row_index,
                        column_index,
                        value,
                        f"{path.name}#page={page_index + 1};lane={lane_id};table={table_index};r={row_index};c={column_index}",
                        bbox=native_cell_bbox(native_table, row_index, column_index),
                        confidence=0.98 if lane_id != "text" else 0.90,
                    ))
                rows.append(cells)
            caption, units, footnotes = table_context(words, bbox)
            table = {
                "table_id": table_id,
                "surface_id": surface_id,
                "source_location": f"page {page_index + 1}, {lane_id} table {table_index}",
                "title": caption,
                "units": units,
                "footnotes": footnotes,
                "bbox": bbox,
                "extraction_method": f"native_pdf_{lane_id}",
                "discovery_lane": lane_id,
                "authority_role": "discovery_only" if lane_id == "text" else "bounded_native",
                "confidence": 0.98 if lane_id != "text" else 0.90,
                "rows": rows,
            }
            tables.append(table)
            lane_tables[lane_id].append(table)
            writer.write_json(f"tables/{safe_id(table_id)}.json", "table_json", table)

        # LANE 0 — columnar reconstruction of unruled tables from the text
        # layer's own geometry. Runs after the ruled lanes and only outside
        # the regions they already bounded, so it can never compete with a
        # ruled result; where its self-verification fails, it emits nothing
        # and the page falls to the vision path exactly as before.
        columnar_grids = reconstruct_columnar_tables(words, table_bboxes)
        columnar_tokens: list[str] = []
        for grid_index, grid in enumerate(columnar_grids, start=1):
            table_id = f"{surface_id}.columnar-t{grid_index}"
            grid_rows: list[list[dict[str, Any]]] = []
            for row_index, values in enumerate(grid["rows"], start=1):
                cells = []
                for column_index, value in enumerate(values, start=1):
                    cells.append(cell_record(
                        row_index,
                        column_index,
                        value,
                        f"{path.name}#page={page_index + 1};lane=columnar;table={grid_index};r={row_index};c={column_index}",
                        confidence=0.95,
                    ))
                grid_rows.append(cells)
            caption, units, footnotes = table_context(words, grid["region_bbox"])
            table = {
                "table_id": table_id,
                "surface_id": surface_id,
                "source_location": f"page {page_index + 1}, columnar table {grid_index}",
                "title": caption,
                "units": units,
                "footnotes": footnotes,
                "bbox": [float(value) for value in grid["region_bbox"]],
                "extraction_method": "native_pdf_columnar",
                "discovery_lane": "columnar",
                "authority_role": "bounded_native",
                "confidence": 0.95,
                "rows": grid_rows,
            }
            tables.append(table)
            table_bboxes.append(table["bbox"])
            # Capture accounting must be census-identical: the ledger's
            # denominator is the census's tokenisation of the source words
            # inside the region, so the credited capture is computed the same
            # way, from the same records — never re-derived from cell text,
            # whose tokenisation can differ and falsely re-trigger vision.
            columnar_tokens.extend(
                record["token"]
                for record in source_numeric_records
                if record_inside_any_bbox(record, [table["bbox"]])
            )
            writer.write_json(f"tables/{safe_id(table_id)}.json", "table_json", table)
        lane_summaries.append({
            "lane_id": "columnar",
            "status": "pass" if columnar_grids else "empty",
            "candidate_count": len(columnar_grids),
        })

        uncovered_regions = uncovered_numeric_regions(source_numeric_records, table_bboxes)
        material_uncovered = [region for region in uncovered_regions if region["material"]]
        source_table_numeric_tokens = [
            record["token"] for record in source_numeric_records
            if record_inside_any_bbox(record, table_bboxes)
        ]
        source_table_tokens.extend(source_table_numeric_tokens)

        # Select the best bounded RULED lane for the provisional capture
        # ledger; the columnar lane owns its own disjoint regions, so its
        # tokens are additive to whichever ruled lane wins, never a competing
        # account of the same cells. Alternative ruled lanes remain
        # corroboration, never additive owners.
        bounded_lane_tokens = {
            lane_id: [token for table in lane_tables[lane_id] for token in table_tokens(table)]
            for lane_id in ("lines", "lines_strict")
        }
        best_native_tokens = min(
            bounded_lane_tokens.values(),
            key=lambda tokens: (
                sum((Counter(source_table_numeric_tokens) - Counter(tokens)).values()),
                sum((Counter(tokens) - Counter(source_table_numeric_tokens)).values()),
                -len(tokens),
            ),
            default=[],
        ) + columnar_tokens
        captured_table_tokens.extend(best_native_tokens)
        native_table_missing = bool(Counter(source_table_numeric_tokens) - Counter(best_native_tokens))

        def overlaps_bounded(box: list[float]) -> bool:
            return any(bbox_iou(box, bounded) >= 0.25 for bounded in table_bboxes)

        text_only_discovery = any(not overlaps_bounded(box) for box in text_discovery_bboxes)
        census = {
            "schema_version": "broker-surface-census/1.0",
            "document_id": descriptor["document_id"],
            "surface_id": surface_id,
            "kind": "pdf_page",
            "source_numeric_tokens": source_numeric_records,
            "whole_surface_numeric_token_count": len(source_numeric_records),
            "source_table_numeric_tokens": source_table_numeric_tokens,
            "table_discovery_lanes": lane_summaries,
            "discovered_table_bboxes": table_bboxes,
            "uncovered_numeric_regions": uncovered_regions,
            "material_uncovered_region_count": len(material_uncovered),
        }
        census_ref = writer.write_json(f"census/page-{page_index + 1:04d}.json", "surface_census", census)

        image_infos = page.get_image_info(xrefs=True) or []
        page_area = max(float(page.rect.width * page.rect.height), 1.0)
        material_image = False
        material_image_bboxes: list[list[float]] = []
        image_refs: list[str] = []
        for image_index, info in enumerate(image_infos, start=1):
            bbox = info.get("bbox")
            image_ratio = 0.0
            if bbox:
                image_area = max(0.0, float((bbox[2] - bbox[0]) * (bbox[3] - bbox[1])))
                image_ratio = image_area / page_area
                if image_ratio >= 0.08:
                    material_image = True
                    material_image_bboxes.append([float(value) for value in bbox])
            # Small logos, icons and repeated disclosure ornaments remain in
            # the immutable PDF and whole-page render; extracting each xref as
            # a second binary artifact creates no table evidence and can make a
            # long research note hundreds of megabytes larger.
            if image_ratio < 0.02:
                continue
            xref = int(info.get("xref") or 0)
            if xref > 0:
                # Discovery is retained below as compact metadata. The raw PDF
                # and page render already preserve the pixels, so a duplicate
                # embedded-image binary is not another source of evidence.
                seen_images.add(xref)

        page_numeric_count = len(source_numeric_records)
        sparse = len(text.strip()) < 40
        undetected_numeric_grid = bool(material_uncovered)
        # Two structured lanes can each look locally complete while describing
        # different segmentations of the same visible table.  Detect that
        # physical ambiguity here, while the source PDF is still open, so the
        # controller can create high-resolution crops and resolve it through
        # the bounded rendered lane.  Canonicalisation must not discover this
        # only after the source context needed for a useful remedy is gone.
        bounded_page_tables = [
            table for table in tables
            if table.get("surface_id") == surface_id
            if table.get("authority_role") != "discovery_only"
            and table.get("bbox")
        ]
        overlapping_native_lanes = any(
            bbox_iou(left.get("bbox"), right.get("bbox")) >= 0.70
            and left.get("extraction_method") != right.get("extraction_method")
            and [
                [normalise_text(cell.get("raw_text")) for cell in row]
                for row in left.get("rows", [])
            ]
            != [
                [normalise_text(cell.get("raw_text")) for cell in row]
                for row in right.get("rows", [])
            ]
            for index, left in enumerate(bounded_page_tables)
            for right in bounded_page_tables[index + 1 :]
        )
        vision_required = (
            (material_image and (page_numeric_count >= 2 or sparse))
            or undetected_numeric_grid
            or text_only_discovery
            or native_table_missing
            or overlapping_native_lanes
        )
        vision_reason = None
        surface_tables = [table for table in tables if table.get("surface_id") == surface_id]
        selected_targets = demand_targets_for_surface(
            text,
            surface_tables,
            demand_contract,
            opaque_image=bool(sparse and material_image),
        )
        # All houses receive cheap native inspection. Native-clean compatible
        # cells remain eligible regardless of which house is later chosen for
        # the single expensive recovery frontier. Only unresolved recovery work
        # is house-bounded; publication date never decides native eligibility.
        unselected_house_recovery = (
            vision_required and vision_house_id is not None and
            descriptor.get("house_id") != vision_house_id
        )
        archive_only = not selected_targets or unselected_house_recovery
        if archive_only:
            # The raw PDF, native text, geometry, census and page image remain
            # preserved.  What closes is only model-facing reconciliation for
            # a page that contains no demanded concept discoverable from its
            # readable text/tables.  Those tables are explicitly barred from
            # the candidate manifest so canonical capture cannot re-arm work.
            for table in surface_tables:
                table["authority_role"] = "archive_only"
                table["model_use"] = "prohibited"
            vision_required = False
            vision_reason = None

        if vision_required:
            unresolved = True
            reasons = []
            if material_image:
                reasons.append("material embedded image")
            if sparse:
                reasons.append("sparse native text")
            if undetected_numeric_grid:
                reasons.append(f"{len(material_uncovered)} material numeric region(s) outside discovered tables")
            if text_only_discovery:
                reasons.append("unruled table candidate discovered only by native text geometry")
            if native_table_missing:
                reasons.append("bounded native lane omits source words inside a table region")
            if overlapping_native_lanes:
                reasons.append("overlapping native table lanes require one rendered physical-grid authority")
            vision_reason = "; ".join(reasons)

        artifact_refs = [text_ref, words_ref, census_ref, *image_refs]
        # Every PDF page is preserved once as a legible workbook-facing image,
        # regardless of whether the page also needs OCR/vision adjudication.
        # Physical capture and model consumption are deliberately separate:
        # the image is the analyst's evidence tab; selected, cell-addressed
        # observations remain the only values permitted to feed the model.
        matrix = fitz.Matrix(render_dpi / 72.0, render_dpi / 72.0)
        pixmap = page.get_pixmap(matrix=matrix, alpha=False)
        page_image_ref = writer.write(
            f"pages/page-{page_index + 1:04d}.png",
            "page_image",
            pixmap.tobytes("png"),
        )
        artifact_refs.append(page_image_ref)
        if vision_required:
                # Give the two independent reads a high-resolution view of
                # every region that may contain a table.  Whole-page renders
                # remain useful for labels and continuation context, but are
                # often too small for dense broker appendices.  Region crops
                # are evidence aids only: they do not become a second source
                # or bypass the immutable whole-surface census.
                candidate_regions = [
                    *table_bboxes,
                    *text_discovery_bboxes,
                    *[region["bbox"] for region in material_uncovered],
                    *material_image_bboxes,
                ]
                unique_regions: list[list[float]] = []
                for raw_bbox in candidate_regions:
                    bbox = [float(value) for value in raw_bbox]
                    if any(bbox_iou(bbox, existing) >= 0.85 for existing in unique_regions):
                        continue
                    unique_regions.append(bbox)
                if not unique_regions:
                    unique_regions.append([
                        float(page.rect.x0), float(page.rect.y0),
                        float(page.rect.x1), float(page.rect.y1),
                    ])
                if len(unique_regions) > MAX_REGION_CROPS_PER_PAGE:
                    unique_regions = sorted(
                        unique_regions,
                        key=lambda box: max(0.0, box[2] - box[0]) * max(0.0, box[3] - box[1]),
                        reverse=True,
                    )[:MAX_REGION_CROPS_PER_PAGE]
                region_crops: list[dict[str, Any]] = []
                crop_dpi = max(300, render_dpi * 2)
                crop_matrix = fitz.Matrix(crop_dpi / 72.0, crop_dpi / 72.0)
                for region_index, bbox in enumerate(unique_regions, start=1):
                    padding = 12.0
                    clip = fitz.Rect(
                        max(float(page.rect.x0), bbox[0] - padding),
                        max(float(page.rect.y0), bbox[1] - padding),
                        min(float(page.rect.x1), bbox[2] + padding),
                        min(float(page.rect.y1), bbox[3] + padding),
                    )
                    if clip.width <= 0 or clip.height <= 0:
                        continue
                    crop_pixmap = page.get_pixmap(matrix=crop_matrix, clip=clip, alpha=False)
                    crop_ref = writer.write(
                        f"crops/page-{page_index + 1:04d}-region-{region_index:03d}.png",
                        "table_crop",
                        crop_pixmap.tobytes("png"),
                    )
                    artifact_refs.append(crop_ref)
                    region_crops.append({
                        "region_id": f"{surface_id}.crop{region_index}",
                        "image_artifact_id": crop_ref,
                        "bbox": [float(clip.x0), float(clip.y0), float(clip.x1), float(clip.y1)],
                        "dpi": crop_dpi,
                    })
                task = {
                    "schema_version": "broker-vision-task/1.0",
                    "document_id": descriptor["document_id"],
                    "surface_id": surface_id,
                    "image_artifact_id": page_image_ref,
                    "reason": vision_reason,
                    "required_passes": 2,
                    "source_census_artifact_id": census_ref,
                    "uncovered_region_ids": [region["region_id"] for region in material_uncovered],
                    "region_crops": region_crops,
                    "selected_cell_contract": {
                        "schema_version": "broker-selected-cell-task/1.0",
                        "model_demand_graph_sha256": demand_contract["model_demand_graph_sha256"],
                        "demand_contract_sha256": demand_contract["contract_sha256"],
                        "targets": selected_targets,
                    },
                    "instruction": selected_cell_instruction(selected_targets),
                }
                task_ref = writer.write_json(
                    f"vision/page-{page_index + 1:04d}.task.json",
                    "vision_task",
                    task,
                )
                artifact_refs.append(task_ref)

        surfaces.append({
            "surface_id": surface_id,
            "kind": "pdf_page",
            "ordinal": page_index + 1,
            "label": f"Page {page_index + 1}",
            "width": float(page.rect.width),
            "height": float(page.rect.height),
            "native_text_chars": len(text),
            "native_word_count": len(words),
            "numeric_token_count": page_numeric_count,
            "table_count": len([table for table in tables if table["surface_id"] == surface_id]),
            "image_count": len(image_infos),
            "artifact_refs": artifact_refs,
            "lane_status": {
                "native_text": "pass" if text.strip() else "empty",
                "geometry": "pass" if words else "empty",
                "tables": "pass" if page_tables else ("error" if all(item["status"] in {"error", "timeout"} for item in lane_summaries) else "none"),
                "images": "pass" if image_infos else "none",
                "vision": "required" if vision_required else "not_required",
            },
            "surface_census_artifact_id": census_ref,
            "whole_surface_numeric_token_count": len(source_numeric_records),
            "source_table_numeric_tokens": source_table_numeric_tokens,
            "uncovered_numeric_regions": uncovered_regions,
            "table_discovery_lanes": lane_summaries,
            "embedded_image_inventory": [
                {
                    "xref": int(info.get("xref") or 0),
                    "bbox": [float(value) for value in info.get("bbox")]
                    if info.get("bbox") else None,
                }
                for info in image_infos
            ],
            "vision_reason": vision_reason,
            "model_demand_status": (
                "selected_cell_recovery_required" if vision_required
                else "archive_only_unselected_house" if unselected_house_recovery
                else "archive_only_not_demanded" if archive_only
                else "native_candidate"
            ),
            "selected_demand_metric_ids": sorted({
                str(item.get("metric_id")) for item in selected_targets
            }),
        })

    return surfaces, tables, source_table_tokens, captured_table_tokens, "needs_vision" if unresolved else "complete"


def extract_xlsx(path: Path, descriptor: dict[str, Any], writer: ArtifactWriter) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str], list[str], str]:
    try:
        import openpyxl  # type: ignore
    except Exception as error:  # pragma: no cover - environment-specific
        raise RuntimeError(f"openpyxl is required for workbook extraction: {error}") from error

    keep_vba = descriptor["media_type"].endswith("macroEnabled.12")
    formula_book = openpyxl.load_workbook(path, data_only=False, read_only=False, keep_vba=keep_vba)
    value_book = openpyxl.load_workbook(path, data_only=True, read_only=False, keep_vba=keep_vba)
    surfaces: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    source_tokens: list[str] = []
    captured_tokens: list[str] = []

    for sheet_index, sheet in enumerate(formula_book.worksheets, start=1):
        value_sheet = value_book[sheet.title]
        nonempty = [cell for row in sheet.iter_rows() for cell in row if cell.value is not None]
        if not nonempty:
            min_row = min_col = 1
            max_row = max_col = 1
        else:
            min_row = min(cell.row for cell in nonempty)
            max_row = max(cell.row for cell in nonempty)
            min_col = min(cell.column for cell in nonempty)
            max_col = max(cell.column for cell in nonempty)
        cell_span = (max_row - min_row + 1) * (max_col - min_col + 1)
        if cell_span > 200000:
            raise RuntimeError(f"Workbook sheet {sheet.title!r} has a {cell_span:,}-cell used rectangle; trim the workbook before extraction.")
        surface_id = f"{descriptor['document_id']}.s{sheet_index}"
        source_cell_records: list[dict[str, Any]] = []
        for formula_cell in nonempty:
            value_cell = value_sheet[formula_cell.coordinate]
            formula = str(formula_cell.value) if formula_cell.data_type == "f" else None
            value = value_cell.value if formula else formula_cell.value
            token = normalise_numeric_token(value)
            source_cell_records.append({
                "cell": formula_cell.coordinate,
                "row": formula_cell.row,
                "column": formula_cell.column,
                "raw_value": serialise_value(formula_cell.value),
                "display_value": serialise_value(value),
                "formula": formula,
                "number_format": formula_cell.number_format,
                "numeric_token": token,
                "bold": bool(formula_cell.font.bold),
                "indent": float(formula_cell.alignment.indent or 0),
                "row_hidden": bool(sheet.row_dimensions[formula_cell.row].hidden),
                "column_hidden": bool(sheet.column_dimensions[formula_cell.column_letter].hidden),
            })
            if token is not None:
                source_tokens.append(token)
        rows: list[list[dict[str, Any]]] = []
        for row_number in range(min_row, max_row + 1):
            cells: list[dict[str, Any]] = []
            for column_number in range(min_col, max_col + 1):
                formula_cell = sheet.cell(row_number, column_number)
                value_cell = value_sheet.cell(row_number, column_number)
                formula = str(formula_cell.value) if formula_cell.data_type == "f" else None
                value = value_cell.value if formula else formula_cell.value
                record = cell_record(
                    row_number - min_row + 1,
                    column_number - min_col + 1,
                    value,
                    f"{path.name}#sheet={sheet.title};cell={formula_cell.coordinate}",
                    formula=formula,
                    number_format=formula_cell.number_format,
                    confidence=1.0,
                )
                record["address"] = formula_cell.coordinate
                record["bold"] = bool(formula_cell.font.bold)
                record["indent"] = float(formula_cell.alignment.indent or 0)
                cells.append(record)
                token = normalise_numeric_token(value)
                if token is not None:
                    captured_tokens.append(token)
            rows.append(cells)
        table_id = f"{surface_id}.used-range"
        table = {
            "table_id": table_id,
            "surface_id": surface_id,
            "source_location": f"sheet {sheet.title}, {sheet.cell(min_row, min_col).coordinate}:{sheet.cell(max_row, max_col).coordinate}",
            "title": sheet.title,
            "units": None,
            "bbox": None,
            "extraction_method": "xlsx_cells",
            "confidence": 1.0,
            "rows": rows,
        }
        tables.append(table)
        grid_ref = writer.write_json(f"sheets/{sheet_index:03d}-{safe_id(sheet.title)}.json", "xlsx_grid", table)
        census = {
            "schema_version": "broker-surface-census/1.0",
            "document_id": descriptor["document_id"],
            "surface_id": surface_id,
            "kind": "workbook_sheet",
            "source_cells": source_cell_records,
            "nonblank_cell_count": len(source_cell_records),
            "whole_surface_numeric_token_count": sum(1 for cell in source_cell_records if cell["numeric_token"] is not None),
            "source_table_numeric_tokens": [cell["numeric_token"] for cell in source_cell_records if cell["numeric_token"] is not None],
            "merged_ranges": sorted(str(item) for item in sheet.merged_cells.ranges),
            "hidden_rows": sorted(index for index, dimension in sheet.row_dimensions.items() if dimension.hidden),
            "hidden_columns": sorted(key for key, dimension in sheet.column_dimensions.items() if dimension.hidden),
            "table_discovery_lanes": [{"lane_id": "xlsx_used_range", "status": "pass", "candidate_count": 1}],
            "uncovered_numeric_regions": [],
            "material_uncovered_region_count": 0,
        }
        census_ref = writer.write_json(f"census/sheet-{sheet_index:04d}.json", "surface_census", census)
        surfaces.append({
            "surface_id": surface_id,
            "kind": "workbook_sheet",
            "ordinal": sheet_index,
            "label": sheet.title,
            "width": float(max_col - min_col + 1),
            "height": float(max_row - min_row + 1),
            "native_text_chars": sum(len(str(cell.value)) for cell in nonempty),
            "native_word_count": len(nonempty),
            "numeric_token_count": len(table_tokens(table)),
            "table_count": 1,
            "image_count": len(getattr(sheet, "_images", [])),
            "artifact_refs": [grid_ref, census_ref],
            "lane_status": {
                "native_text": "pass",
                "geometry": "pass",
                "tables": "pass",
                "images": "unsupported" if getattr(sheet, "_images", []) else "none",
                "vision": "not_required",
            },
            "surface_census_artifact_id": census_ref,
            "whole_surface_numeric_token_count": census["whole_surface_numeric_token_count"],
            "source_table_numeric_tokens": [cell["numeric_token"] for cell in source_cell_records if cell["numeric_token"] is not None],
            "uncovered_numeric_regions": [],
            "table_discovery_lanes": census["table_discovery_lanes"],
            "vision_reason": None,
        })
    return surfaces, tables, source_tokens, captured_tokens, "complete"


def extract_image(path: Path, descriptor: dict[str, Any], writer: ArtifactWriter,
                  demand_contract: dict[str, Any],
                  vision_house_id: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str], list[str], str]:
    """Preserve a standalone broker-table image and emit a bound vision task.

    No OCR output is accepted here.  The image remains NEEDS_VISION until
    compile_broker_vision.py receives two independently produced, hash-bound
    cell transcriptions (or an explicit reviewed resolution).
    """
    try:
        import fitz  # type: ignore
    except Exception as error:  # pragma: no cover - environment-specific
        raise RuntimeError(f"PyMuPDF is required for standalone image extraction: {error}") from error
    raw = path.read_bytes()
    image_document = fitz.open(path)
    if image_document.page_count != 1:
        image_document.close()
        raise RuntimeError("A standalone image must decode to exactly one surface.")
    image_page = image_document.load_page(0)
    width, height = image_page.rect.width, image_page.rect.height
    image_document.close()
    surface_id = f"{descriptor['document_id']}.i1"
    image_ref = writer.write("pages/page-0001.png", "page_image", raw)
    census = {
        "schema_version": "broker-surface-census/1.0",
        "document_id": descriptor["document_id"],
        "surface_id": surface_id,
        "kind": "image_page",
        "source_numeric_tokens": [],
        "whole_surface_numeric_token_count": None,
        "source_table_numeric_tokens": [],
        "table_discovery_lanes": [{"lane_id": "native", "status": "unsupported", "candidate_count": 0}],
        "uncovered_numeric_regions": [{
            "region_id": "full-image",
            "bbox": [0.0, 0.0, float(width), float(height)],
            "token_ids": [],
            "tokens": [],
            "material": True,
            "disposition": "requires_vision",
        }],
        "material_uncovered_region_count": 1,
    }
    census_ref = writer.write_json("census/image-0001.json", "surface_census", census)
    selected_targets = demand_targets_for_surface(
        "", [], demand_contract, opaque_image=True
    )
    if descriptor.get("house_id") != vision_house_id:
        selected_targets = []
    task = {
        "schema_version": "broker-vision-task/1.0",
        "document_id": descriptor["document_id"],
        "surface_id": surface_id,
        "image_artifact_id": image_ref,
        "source_census_artifact_id": census_ref,
        "uncovered_region_ids": ["full-image"],
        "selected_cell_contract": {
            "schema_version": "broker-selected-cell-task/1.0",
            "model_demand_graph_sha256": demand_contract["model_demand_graph_sha256"],
            "demand_contract_sha256": demand_contract["contract_sha256"],
            "targets": selected_targets,
        },
        "instruction": selected_cell_instruction(selected_targets),
        "required_independent_passes": 2,
    }
    task_ref = (
        writer.write_json("vision/image-0001.task.json", "vision_task", task)
        if selected_targets else None
    )
    surface = {
        "surface_id": surface_id,
        "kind": "image_page",
        "ordinal": 1,
        "label": path.name,
        "width": float(width),
        "height": float(height),
        "native_text_chars": 0,
        "native_word_count": 0,
        "numeric_token_count": 0,
        "table_count": 0,
        "image_count": 1,
        "artifact_refs": [image_ref, census_ref, *([task_ref] if task_ref else [])],
        "lane_status": {
            "native_text": "empty",
            "geometry": "unsupported",
            "tables": "none",
            "images": "pass",
            "vision": "required" if selected_targets else "not_required",
        },
        "surface_census_artifact_id": census_ref,
        "whole_surface_numeric_token_count": None,
        "source_table_numeric_tokens": [],
        "uncovered_numeric_regions": census["uncovered_numeric_regions"],
        "table_discovery_lanes": census["table_discovery_lanes"],
        "vision_reason": (
            "Standalone image requires two-pass cell-addressed table transcription."
            if selected_targets else None
        ),
        "vision_disposition": None if selected_targets else "quarantined_evidence_only",
        "model_demand_status": (
            "selected_cell_recovery_required" if selected_targets
            else "archive_only_unselected_house"
        ),
        "selected_demand_metric_ids": sorted({
            str(item.get("metric_id")) for item in selected_targets
        }),
    }
    return [surface], [], [], [], "needs_vision" if selected_targets else "complete"


def rectangular_rows(raw_rows: list[list[Any]], source_name: str, surface_id: str,
                     method: str) -> dict[str, Any]:
    width = max((len(row) for row in raw_rows), default=1)
    rows: list[list[dict[str, Any]]] = []
    for row_index, raw_row in enumerate(raw_rows, start=1):
        cells = []
        for column_index in range(1, width + 1):
            value = raw_row[column_index - 1] if column_index <= len(raw_row) else None
            cells.append(cell_record(
                row_index,
                column_index,
                value,
                f"{source_name}#r={row_index};c={column_index}",
                confidence=1.0,
            ))
        rows.append(cells)
    return {
        "table_id": f"{surface_id}.table-1",
        "surface_id": surface_id,
        "source_location": "entire document",
        "title": None,
        "units": None,
        "bbox": None,
        "extraction_method": method,
        "confidence": 1.0,
        "rows": rows,
    }


def extract_delimited_or_text(path: Path, descriptor: dict[str, Any], writer: ArtifactWriter) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str], list[str], str]:
    text = path.read_text("utf-8-sig")
    surface_id = f"{descriptor['document_id']}.text"
    text_ref = writer.write_text("document.txt", "native_text", text)
    media_type = descriptor["media_type"]
    rows: list[list[Any]] = []
    method = "csv"
    if media_type == "application/json":
        value = json.loads(text)
        if isinstance(value, list) and all(isinstance(row, list) for row in value):
            rows = value
        else:
            rows = [[json.dumps(value, sort_keys=True, ensure_ascii=False)]]
        method = "json"
    elif media_type == "text/csv":
        rows = [list(row) for row in csv.reader(text.splitlines())]
    else:
        for line in text.splitlines():
            if "\t" in line:
                rows.append(line.split("\t"))
            elif "," in line:
                rows.append(next(csv.reader([line])))
            else:
                rows.append([line])
    if not rows:
        rows = [[None]]
    table = rectangular_rows(rows, path.name, surface_id, method)
    table_ref = writer.write_json("table.json", "table_json", table)
    source_tokens = numeric_tokens(text)
    captured_tokens = table_tokens(table)
    census = {
        "schema_version": "broker-surface-census/1.0",
        "document_id": descriptor["document_id"],
        "surface_id": surface_id,
        "kind": "text_document",
        "source_numeric_tokens": [
            {"token_id": f"n{index}", "token": token, "raw_text": token, "bbox": None}
            for index, token in enumerate(source_tokens, start=1)
        ],
        "whole_surface_numeric_token_count": len(source_tokens),
        "source_table_numeric_tokens": source_tokens,
        "table_discovery_lanes": [{"lane_id": method, "status": "pass", "candidate_count": 1}],
        "uncovered_numeric_regions": [],
        "material_uncovered_region_count": 0,
    }
    census_ref = writer.write_json("census/document.json", "surface_census", census)
    surface = {
        "surface_id": surface_id,
        "kind": "text_document",
        "ordinal": 1,
        "label": path.name,
        "width": float(max(len(row) for row in rows)),
        "height": float(len(rows)),
        "native_text_chars": len(text),
        "native_word_count": len(text.split()),
        "numeric_token_count": len(numeric_tokens(text)),
        "table_count": 1,
        "image_count": 0,
        "artifact_refs": [text_ref, table_ref, census_ref],
        "lane_status": {
            "native_text": "pass",
            "geometry": "unsupported",
            "tables": "pass",
            "images": "none",
            "vision": "not_required",
        },
        "surface_census_artifact_id": census_ref,
        "whole_surface_numeric_token_count": len(source_tokens),
        "source_table_numeric_tokens": source_tokens,
        "uncovered_numeric_regions": [],
        "table_discovery_lanes": census["table_discovery_lanes"],
        "vision_reason": None,
    }
    return [surface], [table], source_tokens, captured_tokens, "complete"


def build_ledger(source_tokens: list[str], captured_tokens: list[str]) -> dict[str, Any]:
    missing = counter_difference(source_tokens, captured_tokens)
    extra = counter_difference(captured_tokens, source_tokens)
    denominator = len(source_tokens)
    recall = 1.0 if denominator == 0 else max(0.0, (denominator - len(missing)) / denominator)
    return {
        "source_tokens": source_tokens,
        "captured_tokens": captured_tokens,
        "missing_tokens": missing,
        "duplicate_tokens": extra,
        "recall": recall,
        "denominator_scope": "physical_table_regions",
        "captured_scope": "single_best_authority_lane",
    }


def extract_document(root: Path, request_dir: Path, descriptor: dict[str, Any], render_dpi: int,
                     demand_contract: dict[str, Any], vision_house_id: str) -> dict[str, Any]:
    media_type = descriptor["media_type"]
    if media_type not in SUPPORTED:
        raise RuntimeError(f"Unsupported media type {media_type!r}.")
    source_path = Path(descriptor["path"])
    if not source_path.is_absolute():
        source_path = (request_dir / source_path).resolve()
    if not source_path.is_file():
        raise RuntimeError(f"Broker source does not exist: {source_path}")
    raw_hash = sha256_file(source_path)
    if descriptor.get("expected_sha256") and descriptor["expected_sha256"] != raw_hash:
        raise RuntimeError(f"{descriptor['document_id']} expected_sha256 does not match its bytes.")
    writer = ArtifactWriter(root, descriptor["document_id"])
    if media_type == "application/pdf":
        surfaces, tables, source_tokens, captured_tokens, status = extract_pdf(
            source_path, descriptor, writer, render_dpi, demand_contract, vision_house_id
        )
    elif media_type in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroEnabled.12",
    }:
        surfaces, tables, source_tokens, captured_tokens, status = extract_xlsx(source_path, descriptor, writer)
    elif media_type in {"image/png", "image/jpeg"}:
        surfaces, tables, source_tokens, captured_tokens, status = extract_image(
            source_path, descriptor, writer, demand_contract, vision_house_id
        )
    else:
        surfaces, tables, source_tokens, captured_tokens, status = extract_delimited_or_text(source_path, descriptor, writer)
    ledger = build_ledger(source_tokens, captured_tokens)
    return {
        "document_id": descriptor["document_id"],
        "house_id": descriptor["house_id"],
        "house_name": descriptor["house_name"],
        "source_id": descriptor["source_id"],
        "file_name": source_path.name,
        "media_type": media_type,
        "published_date": descriptor["published_date"],
        "raw_sha256": raw_hash,
        "byte_length": source_path.stat().st_size,
        "surfaces": surfaces,
        "tables": tables,
        "numeric_ledger": ledger,
        "artifacts": writer.records,
        "extraction_status": status,
    }


def extract_readable_pdf_fallback(
    root: Path,
    request_dir: Path,
    descriptor: dict[str, Any],
    render_dpi: int,
    primary_error: Exception,
    demand_contract: dict[str, Any],
    vision_house_id: str,
) -> dict[str, Any]:
    """Preserve every readable PDF page when structured extraction crashes.

    A table-lane implementation failure is not evidence that the user's PDF is
    unusable. Re-open the bytes independently, render every page, and emit the
    same two-pass grid tasks the normal image lane uses. Only an unreadable,
    corrupt or unauthenticated PDF may escape this fallback as a source block.
    """
    import fitz  # type: ignore

    source_path = Path(descriptor["path"])
    if not source_path.is_absolute():
        source_path = (request_dir / source_path).resolve()
    document = fitz.open(source_path)
    if document.needs_pass:
        password = descriptor.get("password") or ""
        if not document.authenticate(password):
            raise RuntimeError(
                "PDF is encrypted and the supplied password did not authenticate."
            )
    writer = ArtifactWriter(root, descriptor["document_id"])
    surfaces = []
    for page_index in range(document.page_count):
        page = document.load_page(page_index)
        surface_id = f"{descriptor['document_id']}.p{page_index + 1}"
        dpi = max(300, int(render_dpi))
        pixmap = page.get_pixmap(matrix=fitz.Matrix(dpi / 72.0, dpi / 72.0), alpha=False)
        image_ref = writer.write(
            f"pages/page-{page_index + 1:04d}.png",
            "page_image",
            pixmap.tobytes("png"),
        )
        census = {
            "schema_version": "broker-surface-census/1.0",
            "document_id": descriptor["document_id"],
            "surface_id": surface_id,
            "kind": "pdf_page",
            "source_numeric_tokens": [],
            "whole_surface_numeric_token_count": 0,
            "source_table_numeric_tokens": [],
            "table_discovery_lanes": [{
                "lane_id": "rendered_page_fallback",
                "status": "pass",
                "candidate_count": 1,
            }],
            "uncovered_numeric_regions": [{
                "region_id": f"{surface_id}.whole-page",
                "bbox": [
                    float(page.rect.x0), float(page.rect.y0),
                    float(page.rect.x1), float(page.rect.y1),
                ],
                "material": True,
                "disposition": "requires_vision",
            }],
            "material_uncovered_region_count": 1,
            "fallback_reason": str(primary_error),
        }
        census_ref = writer.write_json(
            f"census/page-{page_index + 1:04d}.json", "surface_census", census
        )
        selected_targets = demand_targets_for_surface(
            "", [], demand_contract, opaque_image=True
        )
        if descriptor.get("house_id") != vision_house_id:
            selected_targets = []
        task = {
            "schema_version": "broker-vision-task/1.0",
            "document_id": descriptor["document_id"],
            "surface_id": surface_id,
            "image_artifact_id": image_ref,
            "reason": "structured extraction failed; readable rendered-page fallback",
            "required_passes": 2,
            "source_census_artifact_id": census_ref,
            "uncovered_region_ids": [f"{surface_id}.whole-page"],
            "region_crops": [{
                "region_id": f"{surface_id}.whole-page",
                "image_artifact_id": image_ref,
                "bbox": census["uncovered_numeric_regions"][0]["bbox"],
                "dpi": dpi,
            }],
            "selected_cell_contract": {
                "schema_version": "broker-selected-cell-task/1.0",
                "model_demand_graph_sha256": demand_contract["model_demand_graph_sha256"],
                "demand_contract_sha256": demand_contract["contract_sha256"],
                "targets": selected_targets,
            },
            "instruction": selected_cell_instruction(selected_targets),
        }
        task_ref = None
        if selected_targets:
            task_ref = writer.write_json(
                f"vision/page-{page_index + 1:04d}.task.json", "vision_task", task
            )
        surfaces.append({
            "surface_id": surface_id,
            "kind": "pdf_page",
            "ordinal": page_index + 1,
            "label": f"Page {page_index + 1}",
            "width": float(page.rect.width),
            "height": float(page.rect.height),
            "native_text_chars": 0,
            "native_word_count": 0,
            "numeric_token_count": 0,
            "table_count": 0,
            "image_count": 1,
            "artifact_refs": [image_ref, census_ref, *([task_ref] if task_ref else [])],
            "lane_status": {
                "native_text": "error",
                "geometry": "error",
                "tables": "error",
                "images": "pass",
                "vision": "required" if selected_targets else "not_required",
            },
            "surface_census_artifact_id": census_ref,
            "whole_surface_numeric_token_count": 0,
            "source_table_numeric_tokens": [],
            "uncovered_numeric_regions": census["uncovered_numeric_regions"],
            "table_discovery_lanes": census["table_discovery_lanes"],
            "vision_reason": task["reason"] if selected_targets else None,
            "vision_disposition": None if selected_targets else "quarantined_evidence_only",
            "model_demand_status": (
                "selected_cell_recovery_required" if selected_targets
                else "archive_only_unselected_house"
                if descriptor.get("house_id") != vision_house_id
                else "archive_only_not_demanded"
            ),
            "selected_demand_metric_ids": sorted({
                str(item.get("metric_id")) for item in selected_targets
            }),
        })
    raw_hash = sha256_file(source_path)
    return {
        "document_id": descriptor["document_id"],
        "house_id": descriptor["house_id"],
        "house_name": descriptor["house_name"],
        "source_id": descriptor["source_id"],
        "file_name": source_path.name,
        "media_type": descriptor["media_type"],
        "published_date": descriptor["published_date"],
        "raw_sha256": raw_hash,
        "byte_length": source_path.stat().st_size,
        "surfaces": surfaces,
        "tables": [],
        "numeric_ledger": build_ledger([], []),
        "artifacts": writer.records,
        "extraction_status": (
            "needs_vision"
            if any(surface.get("lane_status", {}).get("vision") == "required" for surface in surfaces)
            else "complete"
        ),
        "internal_recovery": {
            "strategy": "rendered_page_fallback",
            "primary_error": str(primary_error),
        },
    }


def validate_request(request: dict[str, Any]) -> None:
    if request.get("schema_version") != "broker-extraction-request/1.0":
        raise ValueError("Unsupported broker extraction request schema_version.")
    if not re.fullmatch(r"[a-z0-9][a-z0-9_.-]*", str(request.get("run_id", ""))):
        raise ValueError("run_id must be a stable lower-case identifier.")
    documents = request.get("documents")
    if not isinstance(documents, list) or not 1 <= len(documents) <= 10:
        raise ValueError("documents must contain 1-10 broker sources.")
    required = {"document_id", "house_id", "house_name", "source_id", "path", "media_type", "published_date"}
    seen_documents: set[str] = set()
    for index, descriptor in enumerate(documents):
        missing = sorted(required - set(descriptor or {}))
        if missing:
            raise ValueError(f"documents[{index}] is missing {', '.join(missing)}.")
        if descriptor["document_id"] in seen_documents:
            raise ValueError(f"Duplicate document_id {descriptor['document_id']!r}.")
        seen_documents.add(descriptor["document_id"])


def tool_versions() -> dict[str, Any]:
    versions: dict[str, Any] = {"python": sys.version.split()[0]}
    try:
        import fitz  # type: ignore
        versions["pymupdf"] = getattr(fitz, "VersionBind", getattr(fitz, "__version__", "unknown"))
    except Exception:
        versions["pymupdf"] = None
    try:
        import openpyxl  # type: ignore
        versions["openpyxl"] = getattr(openpyxl, "__version__", "unknown")
    except Exception:
        versions["openpyxl"] = None
    return versions


CORE_BROKER_DEMANDS = (
    ("revenue", "Revenue", ("revenue", "sales", "turnover")),
    ("ebit", "EBIT / Operating Profit", ("ebit", "operating profit", "operating income")),
    ("adjusted_ebitda", "Adjusted EBITDA", ("adjusted ebitda", "ebitda")),
    ("depreciation_and_amortisation", "Depreciation and amortisation", ("d&a", "depreciation", "amortisation")),
    ("effective_tax_rate", "Effective tax rate", ("effective tax rate", "tax rate")),
    ("capex", "Capital expenditure", ("capex", "capital expenditure")),
    ("change_in_working_capital", "Change in working capital", ("working capital", "change in working capital")),
    ("dividends", "Dividends", ("dividend", "dividends")),
)


def _canonical_contract_hash(value: object) -> str:
    return hashlib.sha256(
        (json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False) + "\n").encode("utf-8")
    ).hexdigest()


def augment_core_broker_demand_contract(contract: dict[str, Any]) -> dict[str, Any]:
    """Ensure the debt-overlay Tier-1 surface exists before broker recovery."""
    output = json.loads(json.dumps(contract))
    targets = output.setdefault("targets", [])
    existing = {str(item.get("metric_id") or item.get("concept_id") or "") for item in targets}
    periods = list(output.get("forecast_periods") or [])
    for metric_id, label, aliases in CORE_BROKER_DEMANDS:
        if metric_id in existing:
            continue
        targets.append({
            "target_id": f"tier1.{metric_id}",
            "metric_id": metric_id,
            "concept_id": metric_id,
            "label": label,
            "search_terms": list(aliases),
            "aliases": list(aliases),
            "forecast_periods": periods,
            "periods": periods,
            "material": True,
            "tier": 1,
        })
    graph_text = json.dumps(output.get("model_demand_graph") or output, sort_keys=True).lower()
    if any(term in graph_text for term in ("buyback", "share repurchase", "share_repurchase")) and "share_buybacks" not in existing:
        targets.append({
            "target_id": "tier1.share_buybacks", "metric_id": "share_buybacks",
            "concept_id": "share_buybacks", "label": "Share buybacks",
            "search_terms": ["share buybacks", "share repurchases", "buybacks"],
            "aliases": ["share buybacks", "share repurchases", "buybacks"],
            "forecast_periods": periods, "periods": periods, "material": True, "tier": 1,
        })
    output.pop("contract_sha256", None)
    output["contract_sha256"] = _canonical_contract_hash(output)
    return output



def ensure_core_broker_demand_contract(contract: dict[str, Any]) -> dict[str, Any]:
    """Compatibility seam; v2 model demand is already complete by construction.

    Legacy v1 carriers must retain their original demand semantics.  Widening an
    old graph by cloning targets would silently change a resumed run.
    """
    return json.loads(json.dumps(contract))


def _recovery_search_terms(demand_contract: dict[str, Any]) -> set[str]:
    terms: set[str] = set()
    for target in demand_contract.get("targets", []):
        for key in ("metric_id", "concept_id", "label", "source_label"):
            value = target.get(key)
            if isinstance(value, str) and value.strip():
                terms.add(normalise_text(value))
        for value in target.get("aliases", []) or target.get("search_terms", []) or []:
            if isinstance(value, str) and value.strip():
                terms.add(normalise_text(value))
    return {term for term in terms if term}


def _native_house_metrics(document: dict[str, Any]) -> set[str]:
    metrics: set[str] = set()
    for surface in document.get("surfaces", []):
        if (surface.get("lane_status") or {}).get("vision") == "required":
            continue
        metrics.update(
            str(value)
            for value in surface.get("selected_demand_metric_ids", [])
            if value
        )
    return metrics


def _select_recovery_house_from_request(
    request: dict[str, Any], request_dir: Path, demand_contract: dict[str, Any],
) -> str:
    """Choose one recovery frontier from cheap native evidence for every house."""
    terms = _recovery_search_terms(demand_contract)
    ranked: list[tuple[tuple[int, int, int, str, str, str], str]] = []
    for descriptor in request.get("documents", []):
        source_path = Path(str(descriptor.get("path") or ""))
        if not source_path.is_absolute():
            source_path = (request_dir / source_path).resolve()
        native_text = ""
        try:
            if descriptor.get("media_type") == "application/pdf" and source_path.is_file():
                import fitz  # type: ignore
                with fitz.open(source_path) as document:
                    native_text = "\n".join(page.get_text("text") or "" for page in document)
        except Exception:
            native_text = ""
        normalized = normalise_text(native_text)
        demand_hits = sum(1 for term in terms if term and term in normalized)
        numeric_tokens = len(re.findall(r"(?<![A-Za-z])[-+]?\(?\d[\d.,]*%?\)?", native_text))
        score = (
            demand_hits,
            min(numeric_tokens, 10000),
            min(len(native_text), 1_000_000),
            str(descriptor.get("published_date") or ""),
            str(descriptor.get("house_id") or ""),
            str(descriptor.get("document_id") or ""),
        )
        ranked.append((score, str(descriptor.get("house_id") or "")))
    if not ranked:
        raise ValueError("Broker extraction request contains no documents.")
    ranked.sort(reverse=True)
    if not ranked[0][1]:
        raise ValueError("No broker house is available for bounded recovery selection.")
    return ranked[0][1]


def _select_recovery_house_from_documents(
    documents: list[dict[str, Any]],
    descriptors: list[dict[str, Any]],
    demand_contract: dict[str, Any],
) -> tuple[str, list[dict[str, Any]]]:
    """Rank verified native usefulness; publication date is only a late tie-breaker."""
    descriptor_by_id = {str(item.get("document_id")): item for item in descriptors}
    demanded = {
        str(item.get("metric_id") or item.get("concept_id"))
        for item in demand_contract.get("targets", [])
    }
    rows: list[dict[str, Any]] = []
    for document in documents:
        descriptor = descriptor_by_id.get(str(document.get("document_id")), {})
        native = _native_house_metrics(document)
        unresolved = sum(
            1
            for surface in document.get("surfaces", [])
            if (surface.get("lane_status") or {}).get("vision") == "required"
        )
        native_tables = sum(
            1
            for table in document.get("tables", [])
            if table.get("model_use") != "prohibited"
            and table.get("authority_role") != "archive_only"
        )
        material_coverage = len(native & demanded)
        rows.append({
            "house_id": str(document.get("house_id") or descriptor.get("house_id") or ""),
            "document_id": str(document.get("document_id") or ""),
            "published_date": str(descriptor.get("published_date") or ""),
            "native_demand_coverage": material_coverage,
            "material_demand_coverage": material_coverage,
            "native_metric_ids": sorted(native & demanded),
            "unresolved_surface_count": unresolved,
            "native_table_count": native_tables,
            "material_contradiction_count": 0,
        })
    rows.sort(
        key=lambda item: (
            -item["material_contradiction_count"],
            item["material_demand_coverage"],
            item["native_demand_coverage"],
            -item["unresolved_surface_count"],
            item["native_table_count"],
            item["published_date"],
            item["house_id"],
            item["document_id"],
        ),
        reverse=True,
    )
    if not rows or not rows[0]["house_id"]:
        raise ValueError("No broker house is available for bounded recovery selection.")
    for rank, row in enumerate(rows, 1):
        row["rank"] = rank
    return rows[0]["house_id"], rows


def select_recovery_house_id(
    source: Any, context: Any, demand_contract: dict[str, Any],
) -> Any:
    """Compatibility dispatcher for preflight and post-extraction quality ranking."""
    if isinstance(source, dict):
        return _select_recovery_house_from_request(source, Path(context), demand_contract)
    return _select_recovery_house_from_documents(list(source), list(context), demand_contract)


def enforce_one_recovery_frontier(documents: list[dict[str, Any]], selected_house_id: str) -> None:
    for document in documents:
        if str(document.get("house_id") or "") == selected_house_id:
            continue
        for surface in document.get("surfaces", []):
            if (surface.get("lane_status") or {}).get("vision") != "required":
                continue
            surface.setdefault("lane_status", {})["vision"] = "not_required"
            surface["model_demand_status"] = "archive_only_unselected_house_recovery"
            surface["selected_demand_metric_ids"] = []
            surface["recovery_prohibited"] = True
            surface_id = surface.get("surface_id")
            for table in document.get("tables", []):
                if table.get("surface_id") != surface_id:
                    continue
                if table.get("authority_role") in {"native_structured_authority", "bounded_native"}:
                    continue
                table["authority_role"] = "archive_only"
                table["model_use"] = "prohibited"
        document["extraction_status"] = (
            "needs_vision"
            if any(
                (surface.get("lane_status") or {}).get("vision") == "required"
                for surface in document.get("surfaces", [])
            )
            else "complete"
        )

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("request", help="broker-extraction-request/1.0 JSON")
    parser.add_argument("--out", required=True, help="output directory outside the skill tree")
    parser.add_argument("--render-dpi", type=int, default=150)
    parser.add_argument("--native-timeout-ms", type=int, default=120_000)
    args = parser.parse_args()
    request_path = Path(args.request).resolve()
    output_root = Path(args.out).resolve()
    request = json.loads(request_path.read_text("utf-8"))
    validate_request(request)
    if not 0 < args.native_timeout_ms <= 120_000:
        raise ValueError("Per-document broker-native timeout must be within 1..120000 ms.")
    demand_contract = ensure_core_broker_demand_contract(
        compile_broker_demand_contract(request.get("model_context") or {})
    )
    vision_house_id = select_recovery_house_id(
        request, request_path.parent, demand_contract
    )
    output_root.mkdir(parents=True, exist_ok=True)

    documents: list[dict[str, Any]] = []
    findings: list[dict[str, Any]] = []
    native_budget_executions: list[dict[str, Any]] = []
    def native_timeout(_signum, _frame):
        raise TimeoutError(
            f"broker native extraction exceeded {args.native_timeout_ms} ms for one document"
        )
    prior_alarm_handler = signal.signal(signal.SIGALRM, native_timeout)
    for descriptor in request["documents"]:
        native_started = time.monotonic()
        outcome = "PASS"
        signal.setitimer(signal.ITIMER_REAL, args.native_timeout_ms / 1000)
        try:
            documents.append(extract_document(
                output_root,
                request_path.parent,
                descriptor,
                args.render_dpi,
                demand_contract,
                vision_house_id,
            ))
        except Exception as error:
            if isinstance(error, TimeoutError):
                outcome = "TIMEOUT"
                findings.append({
                    "id": "broker_extraction.document_native_timeout",
                    "severity": "blocker",
                    "document_id": descriptor.get("document_id"),
                    "surface_id": None,
                    "message": str(error),
                })
                continue
            try:
                if descriptor.get("media_type") != "application/pdf":
                    raise
                # The declared-hash gate is NOT recoverable: the rendered
                # fallback exists for structured-extraction failures on the
                # pinned bytes, never as a route around a source-integrity
                # mismatch. Re-verify before any recovery.
                source_path = (request_path.parent / str(descriptor.get("path") or "")).resolve() \
                    if not Path(str(descriptor.get("path") or "")).is_absolute() \
                    else Path(str(descriptor.get("path") or ""))
                expected = descriptor.get("expected_sha256")
                if expected and (not source_path.is_file() or sha256_file(source_path) != expected):
                    raise
                recovered = extract_readable_pdf_fallback(
                    output_root,
                    request_path.parent,
                    descriptor,
                    args.render_dpi,
                    error,
                    demand_contract,
                    vision_house_id,
                )
                documents.append(recovered)
                findings.append({
                    "id": "broker_extraction.rendered_fallback_activated",
                    "severity": "warning",
                    "document_id": descriptor.get("document_id"),
                    "surface_id": None,
                    "message": (
                        "Structured extraction failed, but the readable PDF was preserved page-for-page and "
                        "routed to bounded rendered-grid recovery. No replacement source is required."
                    ),
                })
            except Exception as fallback_error:
                outcome = "FAIL"
                findings.append({
                    "id": "broker_extraction.document_failed",
                    "severity": "blocker",
                    "document_id": descriptor.get("document_id"),
                    "surface_id": None,
                    "message": f"primary: {error}; rendered fallback: {fallback_error}",
                })
        finally:
            signal.setitimer(signal.ITIMER_REAL, 0)
            native_budget_executions.append({
                "document_id": descriptor.get("document_id"),
                "budget_ms": args.native_timeout_ms,
                "duration_ms": round((time.monotonic() - native_started) * 1000),
                "outcome": outcome,
            })
    signal.signal(signal.SIGALRM, prior_alarm_handler)

    native_budget_body = {
        "schema_version": "broker-native-budget-receipt/1.0",
        "per_document_budget_ms": args.native_timeout_ms,
        "executions": native_budget_executions,
        "status": "PASS" if all(
            item["outcome"] == "PASS" and item["duration_ms"] <= args.native_timeout_ms + 1000
            for item in native_budget_executions
        ) else "DEGRADED",
    }
    native_budget_receipt = {
        **native_budget_body,
        "receipt_sha256": hashlib.sha256(canonical_json(native_budget_body)).hexdigest(),
    }
    native_budget_path = output_root / "broker-native-budget-receipt.json"
    native_budget_path.write_bytes(canonical_json(native_budget_receipt))

    house_ranking: list[dict[str, Any]] = []
    if documents:
        _, house_ranking = _select_recovery_house_from_documents(
            documents, request["documents"], demand_contract
        )
        enforce_one_recovery_frontier(documents, vision_house_id)

    unresolved = sum(
        1
        for document in documents
        for surface in document["surfaces"]
        if surface["lane_status"]["vision"] == "required"
    )
    for document in documents:
        if document["numeric_ledger"]["missing_tokens"]:
            findings.append({
                "id": "broker_extraction.numeric_tokens_outside_tables",
                "severity": "warning",
                "document_id": document["document_id"],
                "surface_id": None,
                "message": f"{len(document['numeric_ledger']['missing_tokens'])} numeric tokens inside bounded physical-table regions are absent from the best native lane; visual verification is required.",
            })
        if document["extraction_status"] == "needs_vision":
            findings.append({
                "id": "broker_extraction.vision_required",
                "severity": "warning",
                "document_id": document["document_id"],
                "surface_id": None,
                "message": "One or more pages require two-pass vision transcription before normalization.",
            })

    if any(item["severity"] == "blocker" for item in findings):
        gate_status = "BLOCKED"
    elif unresolved:
        gate_status = "NEEDS_VISION"
    else:
        gate_status = "PASS"
    surface_count = sum(len(document["surfaces"]) for document in documents)
    table_count = sum(len(document["tables"]) for document in documents)
    cell_count = sum(len(row) for document in documents for table in document["tables"] for row in table["rows"])
    source_token_count = sum(len(document["numeric_ledger"]["source_tokens"]) for document in documents)
    missing_token_count = sum(len(document["numeric_ledger"]["missing_tokens"]) for document in documents)
    recall = 1.0 if source_token_count == 0 else (source_token_count - missing_token_count) / source_token_count
    duplicate_count = sum(len(document["numeric_ledger"]["duplicate_tokens"]) for document in documents)
    uncovered_region_count = sum(
        len(surface.get("uncovered_numeric_regions", []))
        for document in documents for surface in document["surfaces"]
    )
    material_uncovered_region_count = sum(
        sum(1 for region in surface.get("uncovered_numeric_regions", []) if region.get("material"))
        for document in documents for surface in document["surfaces"]
    )
    bundle = {
        "schema_version": "broker-extraction-bundle/1.0",
        "run_id": request["run_id"],
        "created_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "extractor_version": VERSION,
        "tool_versions": tool_versions(),
        "selected_cell_demand_contract": demand_contract,
        "selected_cell_recovery_policy": {
            "schema_version": "broker-selected-house-recovery/1.0",
            "policy": "quality_ranked_native_then_one_recovery_frontier",
            "selected_house_id": vision_house_id,
            "maximum_recovery_house_count": 1,
            "house_ranking": house_ranking,
        },
        "artifact_root": str(output_root),
        "native_budget_receipt": str(native_budget_path),
        "native_budget_receipt_sha256": sha256_file(native_budget_path),
        "documents": documents,
        "summary": {
            "document_count": len(documents),
            "surface_count": surface_count,
            "table_count": table_count,
            "cell_count": cell_count,
            "numeric_token_count": source_token_count,
            "native_numeric_recall": recall,
            "unresolved_surface_count": unresolved,
            "duplicate_cell_count": duplicate_count,
            "uncovered_numeric_region_count": uncovered_region_count,
            "material_uncovered_region_count": material_uncovered_region_count,
        },
        "gate_status": gate_status,
        "findings": findings,
    }
    bundle_path = output_root / "broker-extraction-bundle.json"
    bundle_path.write_bytes(canonical_json(bundle))
    print(json.dumps({
        "status": gate_status,
        "bundle": str(bundle_path),
        "documents": len(documents),
        "surfaces": surface_count,
        "tables": table_count,
        "cells": cell_count,
        "native_numeric_recall": recall,
        "unresolved_surfaces": unresolved,
    }, sort_keys=True))
    return 0 if gate_status != "BLOCKED" else 2


if __name__ == "__main__":
    raise SystemExit(main())
