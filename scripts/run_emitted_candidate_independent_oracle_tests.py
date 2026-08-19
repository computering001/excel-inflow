#!/usr/bin/env python3
"""Build real candidate artifacts and challenge them with independent oracles."""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

from verify.emitted_candidate_artifact_oracle import (
    FORBIDDEN_PRODUCTION_IMPORTS,
    inspect_acquisition_and_lease,
    inspect_consensus,
    inspect_debt_and_interest,
    inspect_emitted_candidate,
    inspect_forecast_ownership,
    inspect_protected_cash_and_rcf,
    inspect_source_arithmetic,
)
from verify.oracle_independence import production_imports as scan_production_imports
from verify.oracle_independence import scan_directory

# Same-language (JavaScript) checkers in scripts/verify share the production
# language and module graph; they are honestly NOT_INDEPENDENT and are never
# scanned as if the Python independence scan could clear them.
NOT_INDEPENDENT_SAME_LANGUAGE_CHECKERS = (
    "verify/recalc_second_opinion.mjs",
    "verify/run_linked_debt_addback_proof_test.mjs",
)


HERE = Path(__file__).resolve().parent
ROOT = HERE.parent


def run(command: list[str], *, env: dict[str, str] | None = None, timeout: int = 600) -> str:
    completed = subprocess.run(
        command, cwd=ROOT, text=True, capture_output=True, timeout=timeout,
        env=env, check=False,
    )
    if completed.returncode != 0:
        raise AssertionError(
            f"command failed ({completed.returncode}): {' '.join(command)}\n"
            f"{completed.stdout[-6000:]}\n{completed.stderr[-6000:]}"
        )
    return completed.stdout


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def build_candidate(representative: Path, output: Path) -> tuple[Path, Path, Path, Path]:
    source = json.loads(representative.read_text(encoding="utf-8"))
    source["execution_profile"] = "reference_parity"
    source.setdefault("controls", {})["broker_case"] = "Model Consensus"
    case_path = output / "candidate-case.json"
    case_path.write_text(json.dumps(source, indent=2) + "\n", encoding="utf-8")
    workbook = output / "candidate.xlsx"
    run(["node", str(HERE / "build_dynamic_model.mjs"), str(case_path), "--out", str(workbook), "--plan-only"])
    run(
        [sys.executable, "-m", "emit", "build", f"{workbook}.plan.json", "--out", str(workbook)],
        env={**os.environ, "PYTHONDONTWRITEBYTECODE": "1", "PYTHONPATH": str(HERE)},
    )
    source_arithmetic = output / "source-arithmetic-artifact.json"
    run([
        sys.executable, str(HERE / "emit_source_arithmetic_candidate_artifact.py"),
        "--model-case", str(case_path), "--out", str(source_arithmetic),
    ])
    return (
        workbook,
        case_path,
        Path(f"{workbook}.forecast-receipt.json"),
        Path(f"{workbook}.workbook-proof-contract.json"),
    )


def row_for(sheet, label: str, occurrence: int = 0) -> int:
    rows = [row for row in range(1, sheet.max_row + 1) if sheet[f"B{row}"].value == label]
    if len(rows) <= occurrence:
        raise AssertionError(f"missing emitted label {label!r}")
    return rows[occurrence]


def workbook_mutation(source: Path, target: Path, mutate) -> None:
    workbook = load_workbook(source, data_only=False)
    mutate(workbook)
    workbook.save(target)


def recalculate(source: Path, target_directory: Path, soffice: str) -> Path:
    target_directory.mkdir(parents=True, exist_ok=True)
    run([soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(target_directory), str(source)])
    target = target_directory / source.name
    if not target.is_file():
        raise AssertionError("portable recalculation did not emit an xlsx")
    return target


def caught(findings: list[dict], code: str) -> bool:
    return any(item.get("code") == code for item in findings)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("representative")
    parser.add_argument("--soffice", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()
    output = Path(args.out).resolve()
    output.mkdir(parents=True, exist_ok=True)
    workbook, case_path, forecast_receipt, proof_contract = build_candidate(
        Path(args.representative).resolve(), output
    )
    source_arithmetic = output / "source-arithmetic-artifact.json"

    baseline = inspect_emitted_candidate(
        workbook_path=workbook,
        model_case_path=case_path,
        forecast_receipt_path=forecast_receipt,
        proof_contract_path=proof_contract,
        source_arithmetic_path=source_arithmetic,
    )
    assert baseline["status"] == "PASS", json.dumps(baseline, indent=2)

    oracle_path = HERE / "verify" / "emitted_candidate_artifact_oracle.py"
    assert FORBIDDEN_PRODUCTION_IMPORTS, "forbidden production-import list must be non-empty"
    independence_scan = scan_directory(HERE / "verify", FORBIDDEN_PRODUCTION_IMPORTS)
    assert oracle_path.name in independence_scan
    contaminated = {name: hits for name, hits in independence_scan.items() if hits}
    assert contaminated == {}, json.dumps(contaminated, indent=2)
    production_imports = independence_scan[oracle_path.name]
    assert production_imports == []
    # Negative self-test: an injected forbidden import MUST be caught, or the
    # scan is decorative.
    injected_copy = output / "injected-forbidden-import-oracle.py"
    injected_copy.write_text(
        oracle_path.read_text(encoding="utf-8") + "\nimport emit  # injected forbidden import\n",
        encoding="utf-8",
    )
    assert scan_production_imports(injected_copy, FORBIDDEN_PRODUCTION_IMPORTS) == ["emit"]
    injected_copy.unlink()

    mutations: list[dict] = []

    def record(domain: str, mutation_id: str, expected_code: str, findings: list[dict]) -> None:
        result = {
            "domain": domain,
            "mutation_id": mutation_id,
            "expected_code": expected_code,
            "caught": caught(findings, expected_code),
        }
        mutations.append(result)
        assert result["caught"], json.dumps({"mutation": result, "findings": findings}, indent=2)

    consensus_mutations = [
        ("named-source-black", "CONSENSUS_NAMED_SOURCE_NOT_BLUE", lambda wb: setattr(wb["Brokers"]["D5"], "font", copy.copy(wb["Brokers"]["D10"].font))),
        ("provider-black", "CONSENSUS_PROVIDER_NOT_BLUE", lambda wb: setattr(wb["Brokers"]["D11"], "font", copy.copy(wb["Brokers"]["D10"].font))),
        ("provider-lineage-removed", "CONSENSUS_PROVIDER_LINEAGE", lambda wb: setattr(wb["Brokers"]["D11"], "comment", None)),
        ("model-blue", "CONSENSUS_MODEL_NOT_BLACK", lambda wb: setattr(wb["Brokers"]["D10"], "font", copy.copy(wb["Brokers"]["D5"].font))),
        ("selected-cross-sheet-black", "CONSENSUS_CROSS_SHEET_LINK_NOT_GREEN", lambda wb: setattr(wb["Operating Model"]["J18"], "font", copy.copy(wb["Brokers"]["D10"].font))),
        ("provider-enters-model-membership", "CONSENSUS_MEMBERSHIP_INVALID", lambda wb: setattr(wb["Brokers"]["D10"], "value", '=IFERROR(AVERAGE(D5,D6,D7,D8,D9,D11),"")')),
        ("named-house-dropped-from-membership", "CONSENSUS_MEMBERSHIP_MISMATCH", lambda wb: setattr(wb["Brokers"]["D10"], "value", '=IFERROR(AVERAGE(D5,D6,D7,D8),"")')),
        ("selected-drops-model-consensus", "CONSENSUS_SELECTED_MISSING_MODEL", lambda wb: setattr(wb["Brokers"]["D18"], "value", "=D11")),
    ]
    for mutation_id, code, mutate in consensus_mutations:
        target = output / f"mutation-{mutation_id}.xlsx"
        workbook_mutation(workbook, target, mutate)
        findings: list[dict] = []
        inspect_consensus(target, findings)
        record("native_workbook_style_provenance" if "black" in mutation_id or "blue" in mutation_id else "broker_consensus_membership", mutation_id, code, findings)

    # A real portable recalculation proves that a changed named-house hardcode
    # flows through Model Consensus, Selected Forecast and the green model link.
    dynamic_source = output / "dynamic-source.xlsx"
    original_value = load_workbook(workbook, data_only=True)["Brokers"]["D5"].value
    workbook_mutation(workbook, dynamic_source, lambda wb: setattr(wb["Brokers"]["D5"], "value", float(original_value) + 100.0))
    recalculated = recalculate(dynamic_source, output / "dynamic-recalculated", args.soffice)
    baseline_values = load_workbook(workbook, data_only=True)
    changed_values = load_workbook(recalculated, data_only=True)
    expected_delta = 20.0
    for sheet, address in (("Brokers", "D10"), ("Brokers", "D18"), ("Operating Model", "J18")):
        delta = float(changed_values[sheet][address].value) - float(baseline_values[sheet][address].value)
        assert abs(delta - expected_delta) < 1e-6, (sheet, address, delta)

    formula_book = load_workbook(workbook, data_only=False)
    operating = formula_book["Operating Model"]
    case = json.loads(case_path.read_text(encoding="utf-8"))
    years = [int(str(period["date"])[:4]) for period in case["periods"] if period.get("status") == "forecast"]
    close_column = ("N", "O", "P")[years.index(int(case["acquisition"]["close_year"]))]
    economic_workbook_mutations = [
        ("acquisition_cash_zero", "acquisition_cash_debt_interest", "ACQUISITION_CASH_DEBT", f"{close_column}{row_for(operating, 'Acquisitions, net of cash acquired')}", "=0"),
        ("acquisition_debt_zero", "acquisition_cash_debt_interest", "ACQUISITION_CASH_DEBT", f"{close_column}{row_for(operating, 'Acquisition term debt')}", "=0"),
        ("acquisition_interest_zero", "acquisition_cash_debt_interest", "ACQUISITION_INTEREST", f"{close_column}{row_for(operating, 'Acquisition debt interest')}", "=0"),
        ("lease_roll_forward_zero", "lease_roll_forward", "LEASE_ROLL_FORWARD", f"J{row_for(operating, 'Lease liabilities — after assumed new lease additions')}", "=0"),
    ]
    for mutation_id, domain, code, address, formula in economic_workbook_mutations:
        source = output / f"mutation-source-{mutation_id}.xlsx"
        workbook_mutation(workbook, source, lambda wb, a=address, f=formula: setattr(wb["Operating Model"][a], "value", f))
        target = recalculate(source, output / f"recalculated-{mutation_id}", args.soffice)
        findings = []
        inspect_acquisition_and_lease(target, case_path, findings)
        record(domain, mutation_id, code, findings)

    # Artifact-only mutations for the remaining independent domains.
    protected_mutation = output / "mutation-protected-cash.xlsx"
    financing_row = row_for(operating, "Net cash from financing activities")
    workbook_mutation(workbook, protected_mutation, lambda wb: setattr(wb["Operating Model"][f"J{financing_row}"], "value", "=J53+J54"))
    findings = []
    inspect_protected_cash_and_rcf(protected_mutation, proof_contract, findings)
    record("cash_flow_protected_identities", "drop-change-in-debt-from-financing", "PROTECTED_CASH_IDENTITY", findings)

    contract_mutation = output / "mutation-rcf-contract.json"
    contract = json.loads(proof_contract.read_text(encoding="utf-8"))
    contract["semantic_scope_captures"] = [entry for entry in contract["semantic_scope_captures"] if entry.get("child_label") != "RCF draw"]
    contract_mutation.write_text(json.dumps(contract, indent=2) + "\n", encoding="utf-8")
    findings = []
    inspect_protected_cash_and_rcf(workbook, contract_mutation, findings)
    record("rcf_financing_cash", "drop-rcf-draw-capture", "RCF_CAPTURE_CONTRACT", findings)

    interest_mutation = output / "mutation-interest-lineage.xlsx"
    workbook_mutation(workbook, interest_mutation, lambda wb: setattr(wb["Operating Model"]["J152"], "value", str(wb["Operating Model"]["J152"].value).replace("$D73", "$D74")))
    findings = []
    inspect_debt_and_interest(interest_mutation, case_path, findings)
    record("instrument_interest_lineage", "wrong-instrument-balance-edge", "INTEREST_LINEAGE", findings)

    case_mutation = output / "mutation-debt-classification.json"
    changed_case = json.loads(case_path.read_text(encoding="utf-8"))
    changed_case["instruments"][0]["class"] = "manual_all_in"
    case_mutation.write_text(json.dumps(changed_case, indent=2) + "\n", encoding="utf-8")
    findings = []
    inspect_debt_and_interest(workbook, case_mutation, findings)
    record("debt_classification", "bond-reclassified-as-other-debt", "DEBT_CLASSIFICATION", findings)

    receipt_mutation = output / "mutation-forecast-ownership.json"
    receipt = json.loads(forecast_receipt.read_text(encoding="utf-8"))
    receipt.append(copy.deepcopy(receipt[0]))
    receipt_mutation.write_text(json.dumps(receipt, indent=2) + "\n", encoding="utf-8")
    findings = []
    inspect_forecast_ownership(receipt_mutation, findings)
    record("forecast_ownership", "duplicate-forecast-writer", "FORECAST_OWNERSHIP_DUPLICATE_WRITER", findings)

    arithmetic = json.loads(source_arithmetic.read_text(encoding="utf-8"))
    compatible_child = next(row for row in arithmetic["rows"] if row.get("parent_source_line_id"))
    compatible_parent = next(row for row in arithmetic["rows"] if row["source_line_id"] == compatible_child["parent_source_line_id"])
    dimensional = copy.deepcopy(arithmetic)
    next(row for row in dimensional["rows"] if row["source_line_id"] == compatible_child["source_line_id"])["reporting_currency"] = "USD"
    dimensional_path = output / "mutation-source-dimension.json"
    dimensional_path.write_text(json.dumps(dimensional, indent=2) + "\n", encoding="utf-8")
    findings = []
    inspect_source_arithmetic(dimensional_path, findings)
    record("source_arithmetic_dimensional_compatibility", "linked-child-wrong-currency", "SOURCE_ARITHMETIC_DIMENSION", findings)

    arithmetic_value = copy.deepcopy(arithmetic)
    next(row for row in arithmetic_value["rows"] if row["source_line_id"] == compatible_parent["source_line_id"])["values"][1] += 1
    arithmetic_value_path = output / "mutation-source-value.json"
    arithmetic_value_path.write_text(json.dumps(arithmetic_value, indent=2) + "\n", encoding="utf-8")
    findings = []
    inspect_source_arithmetic(arithmetic_value_path, findings)
    record("source_arithmetic_dimensional_compatibility", "linked-total-does-not-foot", "SOURCE_ARITHMETIC_VALUE", findings)

    required_domains = {
        "forecast_ownership", "cash_flow_protected_identities", "broker_consensus_membership",
        "debt_classification", "instrument_interest_lineage", "acquisition_cash_debt_interest",
        "lease_roll_forward", "native_workbook_style_provenance", "rcf_financing_cash",
        "source_arithmetic_dimensional_compatibility",
    }
    covered_domains = {item["domain"] for item in mutations}
    assert required_domains == covered_domains, sorted(required_domains - covered_domains)
    assert all(item["caught"] for item in mutations)

    report = {
        "schema_version": "emitted-candidate-independent-oracle-report/1.0",
        "status": "PASS",
        "candidate_workbook_sha256": sha256(workbook),
        "candidate_case_sha256": sha256(case_path),
        "candidate_forecast_receipt_sha256": sha256(forecast_receipt),
        "candidate_proof_contract_sha256": sha256(proof_contract),
        "candidate_source_arithmetic_sha256": sha256(source_arithmetic),
        "baseline": baseline,
        "domains": sorted(required_domains),
        "mutations": mutations,
        "mutations_caught": len(mutations),
        "dynamic_recalculation": {
            "status": "PASS",
            "named_house_delta": 100.0,
            "model_consensus_delta": expected_delta,
            "selected_forecast_delta": expected_delta,
            "operating_model_delta": expected_delta,
        },
        "production_imports": production_imports,
        "independence_scan": {
            "forbidden_production_imports": list(FORBIDDEN_PRODUCTION_IMPORTS),
            "python_files_scanned": sorted(independence_scan),
            "not_independent_same_language_checkers": list(NOT_INDEPENDENT_SAME_LANGUAGE_CHECKERS),
        },
        "artifact_origin": "real_candidate_emission_with_sealed_source_arithmetic",
        "native_excel": "NOT_RUN_PORTABLE_OOXML_AND_LIBREOFFICE",
        "total_violations": 0,
    }
    report_path = output / "emitted-candidate-independent-oracle-report.json"
    report_path.write_text(json.dumps(report, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(report, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
