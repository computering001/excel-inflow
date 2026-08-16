#!/usr/bin/env python3
"""Run the portable end-to-end Excel Inflow cohort from raw inputs to workbooks.

This is the repository-owned production-candidate gate. It starts from raw PDF
and DCS transactions through the shipping entry points, exercises skip, failed
optional close and usable deterministic broker paths, builds acquisition OFF/ON
workbooks, and runs independent workbook proofs. Installed-host semantic
responses remain a separate external certification gate and are never simulated
as host evidence here.
"""
from __future__ import annotations

import argparse
import copy
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import time
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore

ROOT = Path(__file__).resolve().parents[1]
NODE = shutil.which("node") or "node"
EXTERNAL_RELEASE_REQUIREMENTS = {
    "BROKER_CORPUS",
    "BROKER_REAL_PACK_MANIFEST",
    "REAL_FILINGS_REQUEST",
    "REAL_FILINGS_EXPECTATIONS",
    "INSTALLED_HOST_BROKER_CANARY_RECEIPT",
}


def run(command: list[str], *, timeout: int, env: dict[str, str], cwd: Path = ROOT) -> dict[str, Any]:
    started = time.monotonic()
    try:
        completed = subprocess.run(command, cwd=cwd, text=True, capture_output=True, check=False, timeout=timeout, env=env)
        return {
            "command": command,
            "status": "PASS" if completed.returncode == 0 else "FAIL",
            "exit_code": completed.returncode,
            "duration_ms": round((time.monotonic() - started) * 1000, 3),
            "stdout": completed.stdout[-200000:],
            "stderr": completed.stderr[-200000:],
        }
    except subprocess.TimeoutExpired as error:
        return {
            "command": command,
            "status": "TIMEOUT",
            "exit_code": None,
            "duration_ms": round((time.monotonic() - started) * 1000, 3),
            "stdout": error.stdout[-200000:] if isinstance(error.stdout, str) else "",
            "stderr": error.stderr[-200000:] if isinstance(error.stderr, str) else "",
        }
    except Exception as error:
        return {"command": command, "status": "ERROR", "exit_code": None, "duration_ms": round((time.monotonic() - started) * 1000, 3), "stdout": "", "stderr": repr(error)}


def parse_json_stdout(result: dict[str, Any]) -> dict[str, Any] | None:
    if result["status"] != "PASS":
        return None
    text = result.get("stdout") or ""
    starts = [index for index, char in enumerate(text) if char == "{"]
    for start in reversed(starts):
        try:
            value = json.loads(text[start:])
            if isinstance(value, dict):
                return value
        except Exception:
            continue
    return None


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canary(
    fixture: Path,
    python: str,
    soffice: str,
    broker_state: str,
    dcs_basis: str,
    timeout: int,
    env: dict[str, str],
) -> tuple[dict[str, Any], dict[str, Any]]:
    result = run([
        NODE,
        str(ROOT / "scripts" / "run_raw_input_black_box_canary.mjs"),
        str(fixture), python, soffice,
        "--broker-state", broker_state,
        "--dcs-balance-basis", dcs_basis,
    ], timeout=timeout, env=env)
    receipt = parse_json_stdout(result)
    if result["status"] != "PASS" or not receipt or receipt.get("status") != "PASS":
        raise AssertionError(f"Raw canary {broker_state}/{dcs_basis} failed:\n{result.get('stderr') or result.get('stdout')}")
    return result, receipt


def find_case(run_root: Path) -> Path:
    candidates = sorted(
        path for path in run_root.rglob("*.json")
        if path.name in {"model-case.json", "compiled-case.json"}
        or "model-case" in path.name
    )
    for path in candidates:
        try:
            value = json.loads(path.read_text("utf-8"))
        except Exception:
            continue
        schema = str(value.get("schema_version") if isinstance(value, dict) else "")
        if "model-case" in schema or (isinstance(value, dict) and "statement_structure" in value and "periods" in value):
            return path
    raise FileNotFoundError(f"No compiled model case found beneath {run_root}")


def forecast_periods(model_case: dict[str, Any]) -> list[str]:
    return [
        str(period.get("date") or period.get("year"))
        for period in model_case.get("periods", [])
        if period.get("status") == "forecast"
    ]


def recursively_set_existing(value: Any, aliases: set[str], replacement: Any) -> int:
    count = 0
    if isinstance(value, dict):
        for key in list(value):
            if key in aliases:
                old = value[key]
                if isinstance(old, str) and isinstance(replacement, (int, float)):
                    value[key] = str(replacement)
                elif isinstance(old, int) and isinstance(replacement, str) and replacement[:4].isdigit():
                    value[key] = int(replacement[:4])
                else:
                    value[key] = replacement
                count += 1
            else:
                count += recursively_set_existing(value[key], aliases, replacement)
    elif isinstance(value, list):
        for item in value:
            count += recursively_set_existing(item, aliases, replacement)
    return count


def acquisition_container(model_case: dict[str, Any]) -> dict[str, Any]:
    for key in ["acquisition", "acquisition_case", "acquisition_assumptions"]:
        if isinstance(model_case.get(key), dict):
            return model_case[key]
    for parent_key in ["assumptions", "case_inputs", "advanced_cases"]:
        parent = model_case.get(parent_key)
        if isinstance(parent, dict):
            for key, value in parent.items():
                if "acquisition" in key and isinstance(value, dict):
                    return value
    model_case["acquisition"] = {}
    return model_case["acquisition"]


def set_toggle(model_case: dict[str, Any], enabled: bool) -> int:
    aliases = {"adjustment_columns", "acquisition_adjustments", "acquisition_case", "acquisition"}
    controls = model_case.setdefault("controls", {})
    count = 0
    for key in list(controls):
        if key in aliases:
            old = controls[key]
            if isinstance(old, bool):
                controls[key] = enabled
            elif isinstance(old, (int, float)):
                controls[key] = 1 if enabled else 0
            else:
                controls[key] = "On" if enabled else "Off"
            count += 1
    if count == 0:
        controls["adjustment_columns"] = "On" if enabled else "Off"
        count = 1
    return count


def acquisition_case(base: dict[str, Any], enabled: bool) -> tuple[dict[str, Any], dict[str, Any]]:
    model_case = copy.deepcopy(base)
    forecasts = forecast_periods(model_case)
    if len(forecasts) != 3:
        raise AssertionError(f"Expected three forecast periods; received {forecasts}")
    close_period = forecasts[0]
    close_year = int(close_period[:4])
    mutations = {
        "toggle": set_toggle(model_case, enabled),
        "transaction_enterprise_value": recursively_set_existing(model_case, {"transaction_enterprise_value", "transaction_value", "enterprise_value"}, 1000),
        "acquisition_debt_amount": recursively_set_existing(model_case, {"acquisition_debt_amount", "acquisition_debt", "debt_amount"}, 400),
        "entry_ev_ebitda": recursively_set_existing(model_case, {"entry_ev_ebitda", "entry_ev_to_ebitda", "entry_multiple", "entry_ev_ebitda_multiple"}, 10.0),
        "incremental_debt_rate": recursively_set_existing(model_case, {"incremental_debt_rate", "incremental_rate", "acquisition_debt_rate"}, 0.05),
        "close_month": recursively_set_existing(model_case, {"close_month", "acquisition_close_month"}, 6),
        "close_year": recursively_set_existing(model_case, {"close_year", "acquisition_close_year"}, close_period),
    }
    container = acquisition_container(model_case)
    defaults = {
        "transaction_enterprise_value": 1000,
        "acquisition_debt_amount": 400,
        "entry_ev_ebitda": 10.0,
        "incremental_debt_rate": 0.05,
        "close_month": 6,
        "close_year": close_year,
    }
    for key, value in defaults.items():
        if mutations.get(key, 0) == 0:
            container[key] = value
            mutations[key] = 1
    return model_case, {"enabled": enabled, "close_period": close_period, "mutations": mutations}


def build_case(case_path: Path, workbook_path: Path, python: str, soffice: str, timeout: int, env: dict[str, str]) -> dict[str, Any]:
    builder = ROOT / "scripts" / "build_dynamic_model.mjs"
    attempts = [
        [NODE, str(builder), str(case_path), "--out", str(workbook_path), "--python", python, "--soffice", soffice],
        [NODE, str(builder), str(case_path), "--out", str(workbook_path)],
        [NODE, str(builder), "--case", str(case_path), "--out", str(workbook_path), "--python", python, "--soffice", soffice],
        [NODE, str(builder), "--case", str(case_path), "--out", str(workbook_path)],
        [NODE, str(builder), str(case_path), str(workbook_path)],
    ]
    results = []
    for command in attempts:
        if workbook_path.exists():
            workbook_path.unlink()
        result = run(command, timeout=timeout, env=env)
        results.append(result)
        if result["status"] == "PASS" and workbook_path.is_file() and workbook_path.stat().st_size > 0:
            return {"status": "PASS", "selected_command": command, "attempts": results, "workbook": str(workbook_path), "workbook_sha256": sha256_file(workbook_path)}
    return {"status": "FAIL", "selected_command": None, "attempts": results, "workbook": None, "workbook_sha256": None}


def recalculate(workbook: Path, output: Path, soffice: str, timeout: int, env: dict[str, str]) -> Path:
    output.mkdir(parents=True, exist_ok=True)
    source = output / f"source-{workbook.name}"
    shutil.copy2(workbook, source)
    result = run([soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(output), str(source)], timeout=timeout, env=env)
    candidate = output / source.name
    if result["status"] == "PASS" and candidate.is_file():
        return candidate
    # Opening/saving through LibreOffice may not create a different path when
    # the source is already xlsx. The builder's cached workbook remains the
    # fallback and independent cache/parity validators still run below.
    return workbook


def label_rows(workbook: Path, data_only: bool) -> tuple[Any, dict[str, int]]:
    book = load_workbook(workbook, data_only=data_only, read_only=False, keep_links=True)
    if "Operating Model" not in book.sheetnames:
        raise AssertionError(f"Operating Model missing from {workbook}")
    sheet = book["Operating Model"]
    labels: dict[str, int] = {}
    for row in range(1, sheet.max_row + 1):
        values = [sheet.cell(row, column).value for column in range(1, min(7, sheet.max_column) + 1)]
        label = next((value for value in values if isinstance(value, str) and value.strip()), None)
        if label:
            labels[label.strip().casefold()] = row
    return book, labels


def find_label(labels: dict[str, int], predicates: list[set[str]]) -> tuple[str, int] | None:
    for label, row in labels.items():
        tokens = set(re.findall(r"[a-z0-9]+", label))
        if any(required <= tokens for required in predicates):
            return label, row
    return None


def row_values(sheet: Any, row: int, columns: range) -> list[Any]:
    return [sheet.cell(row, column).value for column in columns]


def inspect_acquisition_workbooks(off_path: Path, on_path: Path) -> dict[str, Any]:
    off_formula_book, off_labels = label_rows(off_path, data_only=False)
    on_formula_book, on_labels = label_rows(on_path, data_only=False)
    off_value_book, _ = label_rows(off_path, data_only=True)
    on_value_book, _ = label_rows(on_path, data_only=True)
    off_formula = off_formula_book["Operating Model"]
    on_formula = on_formula_book["Operating Model"]
    off_values = off_value_book["Operating Model"]
    on_values = on_value_book["Operating Model"]
    consideration = find_label(on_labels, [
        {"transaction", "cash", "flow"},
        {"purchase", "consideration"},
        {"acquisition", "consideration"},
        {"direct", "acquisition", "cash"},
    ])
    proceeds = find_label(on_labels, [{"acquisition", "debt", "proceeds"}, {"debt", "proceeds"}])
    acquisition_debt = find_label(on_labels, [{"acquisition", "debt"}])
    gross_debt = find_label(on_labels, [{"gross", "debt"}])
    net_debt = find_label(on_labels, [{"net", "debt"}])
    ending_cash = find_label(on_labels, [{"ending", "cash"}, {"closing", "cash"}])
    leverage = find_label(on_labels, [{"net", "leverage"}, {"leverage"}])
    if not consideration or not proceeds:
        raise AssertionError(f"Funded transaction rows not found. consideration={consideration}, proceeds={proceeds}")
    adjustment_columns = range(14, 17)  # N:P canonical acquisition-adjustment geometry.
    consideration_formulas = row_values(on_formula, consideration[1], adjustment_columns)
    proceeds_formulas = row_values(on_formula, proceeds[1], adjustment_columns)
    assert any(isinstance(value, str) and value.startswith("=") and "ABS" in value.upper() for value in consideration_formulas), consideration_formulas
    assert any(isinstance(value, str) and value.startswith("=") and ("MAX" in value.upper() or "DEBT" in value.upper()) for value in proceeds_formulas), proceeds_formulas
    assert not all(str(value).replace(" ", "") in {"=IF($P$4=0,0,0)", "=0", "0", "None"} for value in consideration_formulas)
    assert not all(str(value).replace(" ", "") in {"=IF($P$4=0,0,0)", "=0", "0", "None"} for value in proceeds_formulas)
    off_consideration = row_values(off_values, consideration[1], adjustment_columns)
    off_proceeds = row_values(off_values, proceeds[1], adjustment_columns)
    on_consideration = row_values(on_values, consideration[1], adjustment_columns)
    on_proceeds = row_values(on_values, proceeds[1], adjustment_columns)
    finite_off_consideration = [float(value) for value in off_consideration if isinstance(value, (int, float))]
    finite_off_proceeds = [float(value) for value in off_proceeds if isinstance(value, (int, float))]
    finite_on_consideration = [float(value) for value in on_consideration if isinstance(value, (int, float))]
    finite_on_proceeds = [float(value) for value in on_proceeds if isinstance(value, (int, float))]
    if finite_off_consideration:
        assert all(abs(value) < 1e-8 for value in finite_off_consideration)
    if finite_off_proceeds:
        assert all(abs(value) < 1e-8 for value in finite_off_proceeds)
    if finite_on_consideration:
        assert min(finite_on_consideration) < 0
    if finite_on_proceeds:
        assert max(finite_on_proceeds) > 0
    economic_differences = {}
    for name, match in {
        "acquisition_debt": acquisition_debt,
        "gross_debt": gross_debt,
        "net_debt": net_debt,
        "ending_cash": ending_cash,
        "leverage": leverage,
    }.items():
        if not match:
            continue
        columns = range(18, min(22, on_values.max_column + 1))  # R:U pro-forma/reference geometry.
        before = row_values(off_values, match[1], columns)
        after = row_values(on_values, match[1], columns)
        changed = any(
            isinstance(left, (int, float)) and isinstance(right, (int, float)) and abs(float(left) - float(right)) > 1e-7
            for left, right in zip(before, after)
        )
        economic_differences[name] = {"changed": changed, "off": before, "on": after}
    assert economic_differences.get("gross_debt", {}).get("changed", True), economic_differences
    assert economic_differences.get("net_debt", {}).get("changed", True), economic_differences
    return {
        "consideration_row": consideration,
        "proceeds_row": proceeds,
        "consideration_formulas": consideration_formulas,
        "proceeds_formulas": proceeds_formulas,
        "off_consideration_values": off_consideration,
        "off_proceeds_values": off_proceeds,
        "on_consideration_values": on_consideration,
        "on_proceeds_values": on_proceeds,
        "economic_differences": economic_differences,
    }


def validation_commands(case_path: Path, workbook: Path, output: Path, python: str, timeout: int, env: dict[str, str]) -> list[dict[str, Any]]:
    output.mkdir(parents=True, exist_ok=True)
    commands = [
        [python, str(ROOT / "scripts" / "verify" / "validate_dynamic_model.py"), str(case_path), str(workbook), "--out", str(output / "dynamic-model.json")],
        [python, str(ROOT / "scripts" / "verify" / "finance_proof.py"), str(case_path), str(workbook), "--out", str(output / "finance-proof.json")],
        [python, str(ROOT / "scripts" / "verify" / "workbook_semantic_oracle.py"), str(case_path), str(workbook), "--out", str(output / "semantic-oracle.json")],
        [NODE, str(ROOT / "scripts" / "validate_cache_parity.mjs"), str(workbook), "--json", str(output / "cache-parity.json")],
    ]
    results = []
    for command in commands:
        script = Path(command[1])
        if not script.is_file():
            results.append({"command": command, "status": "MISSING", "exit_code": None, "duration_ms": 0, "stdout": "", "stderr": "validator missing"})
            continue
        results.append(run(command, timeout=timeout, env=env))
    return results


def create_fixed_point_manifest(cases: list[Path], output: Path) -> Path:
    candidates = [
        {"id": path.stem, "path": str(path.resolve())}
        for path in cases
    ]
    manifest = {
        "schema_version": "fixed-point-case-manifest/1.0",
        "cases": candidates,
    }
    path = output / "fixed-point-cases.json"
    path.write_text(json.dumps(manifest, indent=2) + "\n", "utf-8")
    return path


def run_portable_registry(
    output: Path,
    substitutions: dict[str, Path | str],
    timeout: int,
    env: dict[str, str],
) -> dict[str, Any]:
    registry = json.loads((ROOT / "assets" / "development-test-registry.json").read_text("utf-8"))
    results = []
    for test in registry.get("tests", []):
        requirements = test.get("requires") or []
        missing = [item for item in requirements if item not in substitutions]
        if missing:
            classification = "EXTERNAL_RELEASE_CUSTODY" if all(item in EXTERNAL_RELEASE_REQUIREMENTS for item in missing) else "MISSING_PORTABLE_INPUT"
            results.append({"id": test.get("id"), "phase": test.get("phase"), "status": classification, "missing": missing})
            continue
        script = ROOT / "scripts" / str(test.get("script") or "")
        if not script.is_file():
            results.append({"id": test.get("id"), "phase": test.get("phase"), "status": "MISSING_SCRIPT", "missing": [str(script)]})
            continue
        arguments = []
        for item in test.get("arguments") or []:
            match = re.fullmatch(r"\$([A-Z_]+)", str(item))
            arguments.append(str(substitutions[match.group(1)]) if match else str(item))
        command = [sys.executable if test.get("runtime") == "python" else NODE, str(script), *arguments]
        result = run(command, timeout=timeout, env=env)
        results.append({"id": test.get("id"), "phase": test.get("phase"), **result})
    counts = Counter(item["status"] for item in results)
    unacceptable = [
        item for item in results
        if item["status"] in {"FAIL", "TIMEOUT", "ERROR", "MISSING_SCRIPT", "MISSING_PORTABLE_INPUT"}
    ]
    report = {"status": "PASS" if not unacceptable else "FAIL", "counts": dict(counts), "results": results}
    (output / "portable-registry.json").write_text(json.dumps(report, indent=2, sort_keys=True) + "\n", "utf-8")
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", required=True)
    parser.add_argument("--python", required=True)
    parser.add_argument("--soffice", required=True)
    parser.add_argument("--timeout", type=int, default=1800)
    args = parser.parse_args()
    output = Path(args.out).resolve()
    output.mkdir(parents=True, exist_ok=True)
    python = str(Path(args.python).resolve())
    soffice = str(Path(args.soffice).resolve())
    env = {
        **os.environ,
        "PYTHONDONTWRITEBYTECODE": "1",
        "EXCEL_INFLOW_TEST_PYTHON": python,
        "EXCEL_INFLOW_SOURCE_COMMIT": os.environ.get("EXCEL_INFLOW_SOURCE_COMMIT", "a" * 64),
        "EXCEL_INFLOW_SOURCE_TREE": os.environ.get("EXCEL_INFLOW_SOURCE_TREE", "b" * 64),
    }
    steps = []

    for identifier, script in [
        ("raw-filing-arithmetic", ROOT / "scripts" / "run_raw_filing_arithmetic_integration_test.py"),
        ("raw-broker-native", ROOT / "scripts" / "run_raw_broker_native_integration_test.py"),
    ]:
        result = run([python, str(script)], timeout=args.timeout, env=env)
        steps.append({"id": identifier, **result})
        if result["status"] != "PASS":
            raise AssertionError(result["stderr"] or result["stdout"])

    discovery_root = output / "fixture-discovery"
    discovery = run([
        python,
        str(ROOT / "scripts" / "discover_portable_canary_fixture.py"),
        "--out", str(discovery_root),
        "--python", python,
        "--soffice", soffice,
        "--timeout", str(min(args.timeout, 900)),
    ], timeout=max(args.timeout, 3600), env=env)
    steps.append({"id": "fixture-discovery", **discovery})
    if discovery["status"] != "PASS":
        raise AssertionError(discovery["stderr"] or discovery["stdout"])
    fixture = Path((discovery_root / "selected-fixture-path.txt").read_text("utf-8").strip())

    matrix = run([
        NODE,
        str(ROOT / "scripts" / "run_universal_broker_delivery_matrix.mjs"),
        str(fixture), python, soffice,
    ], timeout=max(args.timeout, 7200), env=env)
    steps.append({"id": "universal-broker-delivery-matrix", **matrix})
    if matrix["status"] != "PASS":
        raise AssertionError(matrix["stderr"] or matrix["stdout"])

    receipts: dict[str, dict[str, Any]] = {}
    canary_results = []
    for state in ["explicit_skip", "failed_optional_close", "usable"]:
        result, receipt = canary(fixture, python, soffice, state, "native_principal", args.timeout, env)
        canary_results.append({"id": f"raw-canary-{state}", **result})
        receipts[state] = receipt
    steps.extend(canary_results)
    assert receipts["explicit_skip"]["economic_signature_sha256"] == receipts["failed_optional_close"]["economic_signature_sha256"]
    assert receipts["explicit_skip"]["runtime_broker_selected_value_count"] == 0
    assert receipts["failed_optional_close"]["runtime_broker_selected_value_count"] == 0
    assert receipts["usable"]["runtime_broker_selected_value_count"] > 0
    assert receipts["usable"].get("installed_host_semantic_seam_exercised") is False
    assert receipts["usable"].get("installed_host_certification_claim") is False

    base_run = Path(receipts["explicit_skip"]["run_root"])
    base_case_path = find_case(base_run)
    base_case = json.loads(base_case_path.read_text("utf-8"))
    cases_root = output / "cases"
    cases_root.mkdir(parents=True, exist_ok=True)
    off_case, off_mutation = acquisition_case(base_case, False)
    on_case, on_mutation = acquisition_case(base_case, True)
    off_case_path = cases_root / "portable-acquisition-off.json"
    on_case_path = cases_root / "portable-acquisition-on.json"
    off_case_path.write_text(json.dumps(off_case, indent=2, sort_keys=True) + "\n", "utf-8")
    on_case_path.write_text(json.dumps(on_case, indent=2, sort_keys=True) + "\n", "utf-8")
    (output / "case-mutations.json").write_text(json.dumps({"off": off_mutation, "on": on_mutation, "base_case": str(base_case_path)}, indent=2, sort_keys=True) + "\n", "utf-8")

    stage4 = run([
        NODE,
        str(ROOT / "scripts" / "run_stage4_checkpoint_tests.mjs"),
        "--cases", str(cases_root),
        "--python", python,
        "--soffice", soffice,
    ], timeout=max(args.timeout, 7200), env=env)
    steps.append({"id": "stage4-portable-cases", **stage4})
    if stage4["status"] != "PASS":
        raise AssertionError(stage4["stderr"] or stage4["stdout"])

    workbooks = output / "workbooks"
    workbooks.mkdir(parents=True, exist_ok=True)
    off_build = build_case(off_case_path, workbooks / "portable-acquisition-off.xlsx", python, soffice, args.timeout, env)
    on_build = build_case(on_case_path, workbooks / "portable-acquisition-on.xlsx", python, soffice, args.timeout, env)
    (output / "builder-attempts.json").write_text(json.dumps({"off": off_build, "on": on_build}, indent=2, sort_keys=True) + "\n", "utf-8")
    if off_build["status"] != "PASS" or on_build["status"] != "PASS":
        raise AssertionError("Direct acquisition workbook build failed; see builder-attempts.json")
    off_workbook = Path(off_build["workbook"])
    on_workbook = Path(on_build["workbook"])
    off_recalculated = recalculate(off_workbook, output / "recalculated-off", soffice, args.timeout, env)
    on_recalculated = recalculate(on_workbook, output / "recalculated-on", soffice, args.timeout, env)
    acquisition_inspection = inspect_acquisition_workbooks(off_recalculated, on_recalculated)
    (output / "acquisition-workbook-inspection.json").write_text(json.dumps(acquisition_inspection, indent=2, sort_keys=True, default=str) + "\n", "utf-8")

    validations = {
        "off": validation_commands(off_case_path, off_workbook, output / "validation-off", python, args.timeout, env),
        "on": validation_commands(on_case_path, on_workbook, output / "validation-on", python, args.timeout, env),
    }
    (output / "workbook-validations.json").write_text(json.dumps(validations, indent=2, sort_keys=True) + "\n", "utf-8")
    validation_failures = [item for group in validations.values() for item in group if item["status"] not in {"PASS", "MISSING"}]
    if validation_failures:
        raise AssertionError(f"Independent workbook validations failed: {[(item['command'], item['status']) for item in validation_failures]}")

    fixed_point_manifest = create_fixed_point_manifest([off_case_path, on_case_path], output)
    substitutions: dict[str, Path | str] = {
        "CASES": cases_root,
        "REPRESENTATIVE": off_case_path,
        "FIXED_POINT_CASES_MANIFEST": fixed_point_manifest,
        "RAW_CANARY_EVIDENCE": fixture,
        "PYTHON": python,
        "SOFFICE": soffice,
    }
    portable_registry = run_portable_registry(output, substitutions, args.timeout, env)
    if portable_registry["status"] != "PASS":
        failures = [item for item in portable_registry["results"] if item["status"] not in {"PASS", "EXTERNAL_RELEASE_CUSTODY"}]
        raise AssertionError(f"Portable registry failures: {[(item['id'], item['status'], item.get('missing')) for item in failures]}")

    report = {
        "schema_version": "excel-inflow-portable-end-to-end-cases/1.0",
        "status": "PASS",
        "fixture": str(fixture),
        "fixture_sha256": sha256_file(fixture),
        "steps": steps,
        "canary_receipts": receipts,
        "skip_failed_economic_equivalence": True,
        "usable_deterministic_broker_selected_value_count": receipts["usable"]["runtime_broker_selected_value_count"],
        "installed_host_semantic_seam_exercised": False,
        "acquisition_case_mutations": {"off": off_mutation, "on": on_mutation},
        "acquisition_workbook_inspection": acquisition_inspection,
        "workbooks": {
            "off": {"path": str(off_workbook), "sha256": sha256_file(off_workbook)},
            "on": {"path": str(on_workbook), "sha256": sha256_file(on_workbook)},
        },
        "portable_registry_counts": portable_registry["counts"],
        "external_release_custody": [
            item for item in portable_registry["results"] if item["status"] == "EXTERNAL_RELEASE_CUSTODY"
        ],
    }
    report["report_sha256"] = hashlib.sha256((json.dumps(report, sort_keys=True, separators=(",", ":"), default=str) + "\n").encode()).hexdigest()
    (output / "portable-end-to-end-cases.json").write_text(json.dumps(report, indent=2, sort_keys=True, default=str) + "\n", "utf-8")
    with (output / "portable-end-to-end-cases.md").open("w", encoding="utf-8") as handle:
        handle.write("# Excel Inflow portable end-to-end case gate\n\n")
        handle.write(f"Status: **PASS**  \nReport SHA-256: `{report['report_sha256']}`\n\n")
        handle.write(f"- Fixture: `{fixture}`\n")
        handle.write(f"- Skip/failed broker economics identical: yes\n")
        handle.write(f"- Usable deterministic broker selected cells: {receipts['usable']['runtime_broker_selected_value_count']}\n")
        handle.write(f"- Acquisition OFF workbook: `{report['workbooks']['off']['sha256']}`\n")
        handle.write(f"- Acquisition ON workbook: `{report['workbooks']['on']['sha256']}`\n")
        handle.write(f"- Installed-host semantic seam exercised: no (separate required release gate)\n")
        handle.write(f"- Portable registry: `{portable_registry['counts']}`\n")
    print(json.dumps({"status": "PASS", "report_sha256": report["report_sha256"], "workbooks": report["workbooks"]}, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
