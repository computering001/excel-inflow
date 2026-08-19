#!/usr/bin/env python3
"""Reconstruct a render plan FROM an emitted .xlsx and compare it to the recorded plan.

WHY THIS EXISTS
---------------
Every other reverse check in this repository is handed its expectation by the
build.  `scripts/verify/workbook_semantic_oracle.py` takes `--contract`
(line 1334) and every check it performs is `contract.get(...)`-keyed
(lines 430-1238); the only contract any caller ever passes is
`<workbook>.workbook-proof-contract.json`, written by the builder itself at
`scripts/build_dynamic_model.mjs:12184` from `workbookSemanticProofContract()`
in `scripts/lib/model_ir_v3.mjs:815`.  A workbook aspect the builder does not
enumerate is therefore an aspect no reverse check can see: merges, data
validations, conditional formats, column widths, freeze panes, comment TRUTH
and `calcPr` were all unread for exactly that reason.  `extract_plan.mjs` --
named in `scripts/emit/verify.py:5`, `scripts/lib/plan_builder.mjs:4` and the
shipped design contract -- does not exist.

This module is the missing direction.  It reads the .xlsx package bytes and
nothing else, rebuilds a plan-shaped description of the workbook, and compares
that description against the RECORDED plan (`<workbook>.plan.json`).  The
recorded plan is one of the two things being compared, never the source of the
reading; no proof contract, semantic manifest, model IR, row map, forecast
receipt or solution file is opened, and none of their names appear in this file.

INDEPENDENCE
------------
Standard library plus openpyxl.  Nothing from `scripts/lib`, the builder, the
case/forecast compilers, the row planner, the solver, the emitter or the render
pipeline is imported -- see FORBIDDEN_PRODUCTION_IMPORTS, which the shared AST
scan in `oracle_independence.py` applies to every file in this directory.
openpyxl is used twice: for its published builtin-number-format table (a
specification constant, not repository code) and as a SECOND, third-party
reader whose disagreement with this file's own stdlib reader is itself a typed
finding -- a parser bug here cannot quietly become a clean bill of health.

WHAT "RECONSTRUCTED" MEANS
--------------------------
Style and differential-style INDICES are never compared.  The emitted package
deduplicates: 203 recorded styles become 202 `cellXfs`, and 120 recorded
differential styles become 9 `dxfs`, so equal indices would prove nothing and
unequal indices would prove nothing either.  Each side dereferences its own
table and the resulting records are compared field by field, which is the only
comparison that says anything about what a reader sees.

Effective values are compared, not spellings.  An absent `<outlinePr/>` and
`summaryBelow="1" summaryRight="1"` are the same sheet; `C4:C4` and `C4` are
the same range.  A comparator that cries wolf about orthography gets its real
findings ignored.

WHAT IS NOT SILENCE
-------------------
Two kinds of gap are reported rather than skipped:

  * `unrecordable_surfaces` -- content the .xlsx carries for which the plan
    schema has no field at all (threaded-comment person ids, comment box
    geometry, sheet visibility, workbook views, theme, docProps).  These are
    enumerated in the report with the part they were found in.  They are not
    blocking, because no plan could ever record them; they are also not hidden.
  * `declared_gaps` -- aspects this reconstructor cannot recover from the file
    even in principle, each with the reason.

Anything else unmapped is a BLOCKING finding: `RECON_UNMAPPED_XML_ELEMENT` and
`RECON_UNMAPPED_XML_ATTRIBUTE` fire on any element or attribute in the workbook
and worksheet parts that this reader neither maps to a plan channel nor has
classified as unrecordable.  A new emitter capability therefore cannot slip
past this oracle by being invisible to it.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.styles.numbers import BUILTIN_FORMATS

# Production-side module tokens no oracle in this directory may import; the
# classification rule lives in oracle_independence.py.
FORBIDDEN_PRODUCTION_IMPORTS: list[str] = [
    "build_dynamic_model",
    "case_compiler",
    "emit",
    "forecast_candidate_compiler",
    "generated",
    "render",
    "row_plan",
    "scripts.lib",
    "solver",
]

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
TC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"

REL_OFFICE_DOCUMENT = DOC_REL_NS + "/officeDocument"
REL_WORKSHEET = DOC_REL_NS + "/worksheet"
REL_STYLES = DOC_REL_NS + "/styles"
REL_SHARED_STRINGS = DOC_REL_NS + "/sharedStrings"
REL_COMMENTS = DOC_REL_NS + "/comments"
REL_DRAWING = DOC_REL_NS + "/drawing"
REL_VML = DOC_REL_NS + "/vmlDrawing"

BORDER_EDGES = ("left", "right", "top", "bottom", "diagonal")
CELL_RE = re.compile(r"^\$?([A-Z]{1,3})\$?([0-9]+)$")

# Cached numeric values are decimal text on both sides.  A round-tripped IEEE
# double is exact, so the only tolerance permitted is the one that absorbs the
# last-place difference between two decimal spellings of the same double.  It is
# named and reported rather than left implicit.
VALUE_RELATIVE_TOLERANCE = 1e-12
VALUE_ABSOLUTE_TOLERANCE = 1e-12

# Findings are capped per code so one corrupted style table cannot bury the
# report; the COUNT is never capped.
FINDINGS_PER_CODE = 20


def _q(tag: str, ns: str = MAIN_NS) -> str:
    return "{%s}%s" % (ns, tag)


def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _flag(value: str | None, default: bool = False) -> bool:
    """An OOXML boolean attribute.  Absent means the caller's default."""
    if value is None:
        return default
    return str(value).strip().lower() in ("1", "true", "on")


def _child_flag(node, tag: str) -> bool:
    """A `<b/>`-style boolean CHILD: absent is false, present without @val is true."""
    child = node.find(_q(tag))
    if child is None:
        return False
    return _flag(child.get("val"), True)


def column_number(letters: str) -> int:
    value = 0
    for char in letters:
        value = value * 26 + ord(char) - 64
    return value


def column_letters(number: int) -> str:
    out = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        out = chr(65 + remainder) + out
    return out


def normalise_address(address: str) -> str:
    match = CELL_RE.match(str(address).strip().upper())
    if not match:
        return str(address).strip().upper()
    return "%s%d" % (match.group(1), int(match.group(2)))


def normalise_sqref(sqref: object) -> str:
    """`C4:C4` and `C4` are the same range, and a sqref is a SET of ranges."""
    tokens = []
    for token in str(sqref or "").replace(",", " ").split():
        token = token.replace("$", "").upper()
        if ":" in token:
            first, last = token.split(":", 1)
            if first == last:
                token = first
        tokens.append(token)
    return " ".join(sorted(tokens))


def canonical_colour(colour: object) -> object:
    if colour is None:
        return None
    if isinstance(colour, dict):
        return {key: colour[key] for key in sorted(colour)}
    text = str(colour).strip().upper()
    if not text:
        return None
    if len(text) == 6:
        text = "FF" + text
    return text


def _sha256(path: Path) -> str:
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()


# ---------------------------------------------------------------------------
# Canonical style records: the two sides dereference their own tables and the
# resolved records are compared, never the indices.
# ---------------------------------------------------------------------------


def canonical_style(record: object) -> dict:
    """A cell style resolved to its EFFECTIVE value, with every field present.

    An `xf` always resolves to a complete effective style, so absence on either
    side means the OOXML default and must compare equal to it.
    """
    record = dict(record or {})
    font = dict(record.get("font") or {})
    border = dict(record.get("border") or {})
    align = dict(record.get("alignment") or {})
    size = font.get("size")
    return {
        "number_format": record.get("number_format") or "General",
        "font": {
            "name": font.get("name") or None,
            "size": float(size) if size is not None else None,
            "color": canonical_colour(font.get("color")),
            "bold": bool(font.get("bold")),
            "italic": bool(font.get("italic")),
            "underline": font.get("underline") or None,
            "strike": bool(font.get("strike")),
        },
        "fill": canonical_fill(record.get("fill")),
        "border": {
            edge: {
                "style": (dict(spec or {})).get("style"),
                "color": canonical_colour((dict(spec or {})).get("color")),
            }
            for edge, spec in sorted(border.items())
            if spec and (dict(spec)).get("style")
        },
        "alignment": {
            "horizontal": align.get("horizontal") or None,
            "vertical": align.get("vertical") or None,
            "indent": int(align.get("indent") or 0),
            "wrap_text": bool(align.get("wrap_text")),
            "shrink_to_fit": bool(align.get("shrink_to_fit")),
        },
    }


def canonical_fill(fill: object) -> object:
    if not fill:
        return None
    fill = dict(fill)
    record = {"pattern": fill.get("pattern") or None}
    for key in ("fg_color", "bg_color"):
        colour = canonical_colour(fill.get(key))
        if colour is not None:
            record[key] = colour
    if record["pattern"] == "gradient" and fill.get("gradient_type"):
        record["gradient_type"] = fill["gradient_type"]
    return record


def canonical_differential_style(record: object) -> dict:
    """A dxf is PRESENCE-shaped, not effective-shaped.

    A differential style that says nothing about bold is not a differential
    style that says bold=false: it leaves the cell's own bold alone.  So absent
    keys stay absent here, unlike `canonical_style`, and a `<b val="0"/>` the
    plan never recorded is a real difference rather than a shared default.
    """
    record = dict(record or {})
    out: dict = {}
    if record.get("number_format") is not None:
        out["number_format"] = record["number_format"]
    font = dict(record.get("font") or {})
    if font:
        resolved = {}
        for key in ("name", "underline"):
            if font.get(key) is not None:
                resolved[key] = font[key]
        if font.get("size") is not None:
            resolved["size"] = float(font["size"])
        if font.get("color") is not None:
            resolved["color"] = canonical_colour(font["color"])
        for key in ("bold", "italic", "strike"):
            if font.get(key) is not None:
                resolved[key] = bool(font[key])
        if resolved:
            out["font"] = {key: resolved[key] for key in sorted(resolved)}
    fill = canonical_fill(record.get("fill"))
    if fill is not None:
        out["fill"] = fill
    border = dict(record.get("border") or {})
    resolved_border = {
        edge: {
            "style": (dict(spec or {})).get("style"),
            "color": canonical_colour((dict(spec or {})).get("color")),
        }
        for edge, spec in sorted(border.items())
        if spec and (dict(spec)).get("style")
    }
    if resolved_border:
        out["border"] = resolved_border
    return out


# ---------------------------------------------------------------------------
# The unmapped sweep: every element and attribute is either a plan channel, a
# named unrecordable surface, or a blocking finding.
# ---------------------------------------------------------------------------

# The sweep tables are PART-SCOPED.  A global table would accept `<left/>`
# (a border edge in styles.xml) inside a worksheet, which is exactly the kind of
# hole that lets a new emitter capability pass unnoticed.
#
#   mapped                  -> read into the reconstruction
#   unrecordable            -> present in the file, no plan field, reported
#   attributes              -> attributes of a mapped element that are read
#   unrecordable_attributes -> attributes with no plan field, reported
#
# Anything outside those four is RECON_UNMAPPED_XML_ELEMENT / _ATTRIBUTE.
SWEEP_TABLES = {
    "worksheet": {
        "mapped": {
            "worksheet", "sheetPr", "outlinePr", "sheetView", "pane",
            "sheetFormatPr", "cols", "col", "sheetData", "row", "c", "f", "v",
            "is", "t", "r", "rPr", "mergeCells", "mergeCell",
            "conditionalFormatting", "cfRule", "formula", "dataValidations",
            "dataValidation", "formula1", "formula2", "pageMargins",
            "pageSetup", "drawing", "hyperlinks", "hyperlink",
        },
        "unrecordable": {
            "dimension": "cached used range; derivable from the cells themselves",
            "sheetViews": "container for the view whose pane and gridlines ARE planned",
            "selection": "cursor position; the plan has no selection field",
            "pageSetUpPr": "print-setup container; empty in the emitted package",
            "legacyDrawing": "comment box geometry lives in VML; the plan records only cell and text",
        },
        "attributes": {
            "worksheet": set(),
            "sheetPr": set(),
            "outlinePr": {"summaryBelow", "summaryRight"},
            "sheetView": {"showGridLines", "workbookViewId", "tabSelected",
                          "zoomScale", "zoomScaleNormal", "showRowColHeaders",
                          "showZeros", "rightToLeft", "view"},
            "pane": {"xSplit", "ySplit", "topLeftCell", "activePane", "state"},
            "sheetFormatPr": {"defaultRowHeight", "outlineLevelRow",
                              "outlineLevelCol", "defaultColWidth",
                              "customHeight", "zeroHeight", "thickTop",
                              "thickBottom"},
            "cols": set(),
            "col": {"min", "max", "width", "customWidth", "hidden",
                    "outlineLevel", "style", "bestFit", "collapsed"},
            "sheetData": set(),
            "row": {"r", "ht", "customHeight", "hidden", "outlineLevel",
                    "collapsed", "s", "customFormat", "spans", "thickBot",
                    "thickTop"},
            "c": {"r", "s", "t", "cm", "vm", "ph"},
            "f": {"t", "ref", "si", "ca", "shared", "aca", "dt2D", "dtr",
                  "del1", "del2", "r1", "r2", "bx"},
            "v": set(),
            "is": set(),
            "t": {"space"},
            "mergeCells": {"count"},
            "mergeCell": {"ref"},
            "conditionalFormatting": {"sqref", "pivot"},
            "cfRule": {"type", "priority", "operator", "dxfId", "stopIfTrue",
                       "text", "percent", "bottom", "rank", "stdDev",
                       "aboveAverage", "equalAverage", "timePeriod"},
            "formula": set(),
            "dataValidations": {"count", "disablePrompts", "xWindow", "yWindow"},
            "dataValidation": {"sqref", "type", "operator", "allowBlank",
                               "showErrorMessage", "showInputMessage",
                               "promptTitle", "prompt", "errorTitle", "error",
                               "errorStyle", "imeMode"},
            "formula1": set(),
            "formula2": set(),
            "pageMargins": {"left", "right", "top", "bottom", "header", "footer"},
            "pageSetup": {"orientation", "scale", "fitToWidth", "fitToHeight",
                          "paperSize", "horizontalDpi", "verticalDpi", "r:id"},
            "drawing": {"r:id"},
            "hyperlinks": set(),
            "hyperlink": {"ref", "r:id", "location", "display", "tooltip"},
        },
        "unrecordable_attributes": {
            ("sheetFormatPr", "baseColWidth"): "default column width unit; no plan field",
            ("dataValidation", "showDropDown"): "in-cell dropdown suppression; no plan field",
        },
    },
    "workbook": {
        "mapped": {"workbook", "sheets", "sheet", "definedNames", "definedName", "calcPr"},
        "unrecordable": {
            "workbookPr": "workbook-level date and calc spelling flags; no plan field",
            "workbookProtection": "protection element; no plan field",
            "bookViews": "window geometry; no plan field",
            "workbookView": "window geometry; no plan field",
            "fileVersion": "producer identity; no plan field",
        },
        "attributes": {
            "workbook": set(),
            "sheets": set(),
            "sheet": {"name", "r:id"},
            "definedNames": set(),
            "definedName": {"name", "localSheetId", "hidden", "comment"},
            "calcPr": {"calcId", "calcMode", "fullCalcOnLoad", "forceFullCalc",
                       "iterate", "iterateCount", "iterateDelta", "refMode",
                       "fullPrecision", "calcCompleted", "calcOnSave",
                       "concurrentCalc", "concurrentManualCount"},
        },
        "unrecordable_attributes": {
            ("sheet", "state"): "sheet visibility; no plan field",
            ("sheet", "sheetId"): "sheet id; the plan records name and order",
        },
    },
    "styles": {
        "mapped": {
            "styleSheet", "numFmts", "numFmt", "fonts", "font", "name", "sz",
            "color", "b", "i", "u", "strike", "family", "scheme", "charset",
            "fills", "fill", "patternFill", "fgColor", "bgColor",
            "gradientFill", "stop", "borders", "border", "left", "right",
            "top", "bottom", "diagonal", "cellXfs", "xf", "alignment",
            "dxfs", "dxf",
        },
        "unrecordable": {
            "cellStyleXfs": "style parents; the plan records the resolved style of each cell",
            "cellStyles": "named-style table; the plan records resolved styles",
            "cellStyle": "named-style entry; no plan field",
            "colors": "indexed-colour palette; the plan records explicit ARGB",
            "indexedColors": "indexed-colour palette; the plan records explicit ARGB",
            "rgbColor": "indexed-colour palette entry; the plan records explicit ARGB",
            "tableStyles": "table style defaults; no plan field",
        },
        "attributes": {
            "styleSheet": set(),
            "numFmts": {"count"},
            "numFmt": {"numFmtId", "formatCode"},
            "fonts": {"count", "knownFonts"},
            "font": set(),
            "name": {"val"},
            "sz": {"val"},
            "color": {"rgb", "theme", "tint", "indexed", "auto"},
            "b": {"val"}, "i": {"val"}, "u": {"val"}, "strike": {"val"},
            "family": {"val"}, "scheme": {"val"}, "charset": {"val"},
            "fills": {"count"},
            "fill": set(),
            "patternFill": {"patternType"},
            "fgColor": {"rgb", "theme", "tint", "indexed", "auto"},
            "bgColor": {"rgb", "theme", "tint", "indexed", "auto"},
            "gradientFill": {"type", "degree", "left", "right", "top", "bottom"},
            "stop": {"position"},
            "borders": {"count"},
            "border": {"diagonalUp", "diagonalDown", "outline"},
            "left": {"style"}, "right": {"style"}, "top": {"style"},
            "bottom": {"style"}, "diagonal": {"style"},
            "cellXfs": {"count"},
            "xf": {"numFmtId", "fontId", "fillId", "borderId"},
            "alignment": {"horizontal", "vertical", "indent", "wrapText",
                          "shrinkToFit", "textRotation", "readingOrder",
                          "justifyLastLine", "relativeIndent"},
            "dxfs": {"count"},
            "dxf": set(),
        },
        "unrecordable_attributes": {
            ("xf", "pivotButton"): "pivot affordance; no plan field",
            ("xf", "quotePrefix"): "quote prefix; no plan field",
            ("xf", "xfId"): "style parent index; the plan records resolved styles",
            ("xf", "applyNumberFormat"): "inheritance gate; see declared_gaps 'apply* style flags'",
            ("xf", "applyFont"): "inheritance gate; see declared_gaps 'apply* style flags'",
            ("xf", "applyFill"): "inheritance gate; see declared_gaps 'apply* style flags'",
            ("xf", "applyBorder"): "inheritance gate; see declared_gaps 'apply* style flags'",
            ("xf", "applyAlignment"): "inheritance gate; see declared_gaps 'apply* style flags'",
            ("xf", "applyProtection"): "inheritance gate; see declared_gaps 'apply* style flags'",
        },
    },
    "comments": {
        "mapped": {"comments", "authors", "author", "commentList", "comment",
                   "text", "t", "r", "rPr", "sz", "color", "rFont", "family",
                   "scheme", "b", "i", "u", "charset"},
        "unrecordable": {},
        "attributes": {
            "comments": set(), "authors": set(), "author": set(),
            "commentList": set(), "comment": {"ref"}, "text": set(),
            "t": {"space"}, "r": set(), "rPr": set(),
            "sz": {"val"}, "color": {"rgb", "theme", "tint", "indexed"},
            "rFont": {"val"}, "family": {"val"}, "scheme": {"val"},
            "b": {"val"}, "i": {"val"}, "u": {"val"}, "charset": {"val"},
        },
        "unrecordable_attributes": {
            ("comment", "shapeId"): "comment box shape; the plan records cell and text",
            ("comment", "authorId"): "index into the comment author table",
        },
    },
}


# ---------------------------------------------------------------------------
# The stdlib package reader
# ---------------------------------------------------------------------------


class Package:
    """The .xlsx as a zip of XML parts.  Nothing outside the file is opened."""

    def __init__(self, path: Path) -> None:
        self.path = Path(path)
        self.zip = zipfile.ZipFile(self.path)
        self.names = set(self.zip.namelist())

    def close(self) -> None:
        self.zip.close()

    def xml(self, name: str):
        return ET.fromstring(self.zip.read(name))

    def rels(self, part: str) -> dict:
        directory, _, base = part.rpartition("/")
        rels_name = "%s/_rels/%s.rels" % (directory, base) if directory else "_rels/%s.rels" % base
        if rels_name not in self.names:
            return {}
        out = {}
        for node in self.xml(rels_name):
            out[node.get("Id")] = (node.get("Type"), node.get("Target"), node.get("TargetMode"))
        return out

    def resolve(self, base_part: str, target: str) -> str:
        if target.startswith("/"):
            return target[1:]
        directory = base_part.rpartition("/")[0]
        parts = [segment for segment in (directory.split("/") if directory else [])]
        for segment in target.split("/"):
            if segment == "..":
                if parts:
                    parts.pop()
            elif segment not in ("", "."):
                parts.append(segment)
        return "/".join(parts)


def _font_effective(node) -> dict:
    if node is None:
        return {}
    name = node.find(_q("name"))
    size = node.find(_q("sz"))
    colour = node.find(_q("color"))
    underline = node.find(_q("u"))
    return {
        "name": name.get("val") if name is not None else None,
        "size": float(size.get("val")) if size is not None and size.get("val") else None,
        "color": _colour_of(colour),
        "bold": _child_flag(node, "b"),
        "italic": _child_flag(node, "i"),
        "underline": (underline.get("val") or "single") if underline is not None else None,
        "strike": _child_flag(node, "strike"),
    }


def _font_present(node) -> dict:
    """Only the font facts the element PHYSICALLY carries (for dxf records)."""
    if node is None:
        return {}
    out: dict = {}
    name = node.find(_q("name"))
    if name is not None and name.get("val"):
        out["name"] = name.get("val")
    size = node.find(_q("sz"))
    if size is not None and size.get("val"):
        out["size"] = float(size.get("val"))
    colour = node.find(_q("color"))
    if colour is not None:
        out["color"] = _colour_of(colour)
    underline = node.find(_q("u"))
    if underline is not None:
        out["underline"] = underline.get("val") or "single"
    for tag, key in (("b", "bold"), ("i", "italic"), ("strike", "strike")):
        child = node.find(_q(tag))
        if child is not None:
            out[key] = _flag(child.get("val"), True)
    return out


def _colour_of(node) -> object:
    if node is None:
        return None
    if node.get("rgb"):
        return node.get("rgb")
    if node.get("theme") is not None:
        record = {"theme": int(node.get("theme"))}
        if node.get("tint"):
            record["tint"] = float(node.get("tint"))
        return record
    if node.get("indexed") is not None:
        return {"indexed": int(node.get("indexed"))}
    if _flag(node.get("auto")):
        return {"auto": True}
    return None


def _fill_of(node) -> object:
    if node is None:
        return None
    pattern = node.find(_q("patternFill"))
    if pattern is not None:
        kind = pattern.get("patternType")
        if not kind:
            return None
        record = {"pattern": kind}
        fg = _colour_of(pattern.find(_q("fgColor")))
        bg = _colour_of(pattern.find(_q("bgColor")))
        if fg is not None:
            record["fg_color"] = fg
        if bg is not None:
            record["bg_color"] = bg
        return record
    gradient = node.find(_q("gradientFill"))
    if gradient is not None:
        return {"pattern": "gradient", "gradient_type": gradient.get("type") or "linear"}
    return None


def _border_of(node) -> dict:
    if node is None:
        return {}
    out = {}
    for edge in BORDER_EDGES:
        child = node.find(_q(edge))
        if child is None:
            continue
        style = child.get("style")
        if not style:
            continue
        out[edge] = {"style": style, "color": _colour_of(child.find(_q("color")))}
    return out


def _alignment_of(node) -> dict:
    if node is None:
        return {}
    out: dict = {}
    for attr, key in (("horizontal", "horizontal"), ("vertical", "vertical")):
        if node.get(attr):
            out[key] = node.get(attr)
    if node.get("indent") and int(node.get("indent")) != 0:
        out["indent"] = int(node.get("indent"))
    for attr, key in (("wrapText", "wrap_text"), ("shrinkToFit", "shrink_to_fit")):
        if _flag(node.get(attr)):
            out[key] = True
    return out


def read_style_tables(package: Package, styles_part: str) -> tuple[list, list]:
    """`cellXfs` -> resolved style records, `dxfs` -> resolved differential records."""
    if styles_part is None or styles_part not in package.names:
        return [canonical_style({})], []
    root = package.xml(styles_part)
    number_formats = dict(BUILTIN_FORMATS)
    numfmts = root.find(_q("numFmts"))
    if numfmts is not None:
        for node in numfmts.findall(_q("numFmt")):
            number_formats[int(node.get("numFmtId"))] = node.get("formatCode")
    fonts = [_font_effective(node) for node in (root.find(_q("fonts")) or [])]
    fills = [_fill_of(node) for node in (root.find(_q("fills")) or [])]
    borders = [_border_of(node) for node in (root.find(_q("borders")) or [])]

    styles = []
    cell_xfs = root.find(_q("cellXfs"))
    for node in (cell_xfs if cell_xfs is not None else []):
        numfmt_id = int(node.get("numFmtId") or 0)
        font_id = int(node.get("fontId") or 0)
        fill_id = int(node.get("fillId") or 0)
        border_id = int(node.get("borderId") or 0)
        styles.append(canonical_style({
            "number_format": number_formats.get(numfmt_id, "General"),
            "font": fonts[font_id] if font_id < len(fonts) else {},
            "fill": fills[fill_id] if fill_id < len(fills) else None,
            "border": borders[border_id] if border_id < len(borders) else {},
            "alignment": _alignment_of(node.find(_q("alignment"))),
        }))
    if not styles:
        styles = [canonical_style({})]

    differentials = []
    dxfs = root.find(_q("dxfs"))
    for node in (dxfs if dxfs is not None else []):
        numfmt = node.find(_q("numFmt"))
        record = {
            "font": _font_present(node.find(_q("font"))),
            "fill": _fill_of(node.find(_q("fill"))),
            "border": _border_of(node.find(_q("border"))),
        }
        if numfmt is not None and numfmt.get("formatCode"):
            record["number_format"] = numfmt.get("formatCode")
        differentials.append(canonical_differential_style(record))
    return styles, differentials


def read_shared_strings(package: Package, part: str | None) -> list:
    if not part or part not in package.names:
        return []
    root = package.xml(part)
    return [_rich_text(node) for node in root.findall(_q("si"))]


def _rich_text(node) -> str:
    if node is None:
        return ""
    pieces = []
    direct = node.find(_q("t"))
    if direct is not None:
        pieces.append(direct.text or "")
    for run in node.findall(_q("r")):
        text = run.find(_q("t"))
        pieces.append((text.text or "") if text is not None else "")
    return "".join(pieces)


def _cells_of(sheet_root, styles: list, shared: list) -> dict:
    cells: dict = {}
    data = sheet_root.find(_q("sheetData"))
    if data is None:
        return cells
    for row in data.findall(_q("row")):
        for cell in row.findall(_q("c")):
            address = normalise_address(cell.get("r") or "")
            index = int(cell.get("s") or 0)
            record: dict = {"style": styles[index] if index < len(styles) else canonical_style({})}
            formula = cell.find(_q("f"))
            if formula is not None:
                record["formula"] = formula.text or ""
                if formula.get("t"):
                    record["formula_kind"] = formula.get("t")
                if formula.get("ref"):
                    record["formula_ref"] = formula.get("ref")
                if formula.get("si") is not None:
                    record["formula_si"] = formula.get("si")
            kind = cell.get("t") or "n"
            value_node = cell.find(_q("v"))
            # The plan has ONE string marker, "s"; OOXML has three spellings.
            # The spelling is not lost, it is DERIVABLE -- a formula whose result
            # is text must be `str`, a literal must be `inlineStr` or `s` -- so
            # it is recorded here and checked rather than declared unreachable.
            if kind == "inlineStr":
                record["value"] = _rich_text(cell.find(_q("is")))
                record["type"] = "s"
                record["type_spelling"] = kind
            elif kind == "s":
                position = int(value_node.text) if value_node is not None and value_node.text else 0
                record["value"] = shared[position] if position < len(shared) else ""
                record["type"] = "s"
                record["type_spelling"] = kind
            elif kind == "str":
                record["value"] = (value_node.text or "") if value_node is not None else ""
                record["type"] = "s"
                record["type_spelling"] = kind
            elif kind == "b":
                record["value"] = _flag(value_node.text if value_node is not None else None)
                record["type"] = "b"
            elif kind == "e":
                record["value"] = value_node.text if value_node is not None else ""
                record["type"] = "e"
            elif value_node is not None and (value_node.text or "").strip() != "":
                record["value"] = float(value_node.text)
                record["type"] = "n"
            cells[address] = record
    return cells


def _columns_of(sheet_root) -> list:
    out = []
    cols = sheet_root.find(_q("cols"))
    for node in (cols if cols is not None else []):
        record = {
            "min": int(node.get("min")),
            "max": int(node.get("max")),
            "width": float(node.get("width")) if node.get("width") else None,
            "hidden": _flag(node.get("hidden")),
            "custom_width": _flag(node.get("customWidth")),
        }
        if node.get("outlineLevel") and int(node.get("outlineLevel")) != 0:
            record["outline_level"] = int(node.get("outlineLevel"))
        if node.get("style") and int(node.get("style")) != 0:
            record["style"] = int(node.get("style"))
        out.append(record)
    return out


def _rows_of(sheet_root) -> list:
    """A row record exists only where the row carries a non-default property."""
    out = []
    data = sheet_root.find(_q("sheetData"))
    for node in (data.findall(_q("row")) if data is not None else []):
        record = {"row": int(node.get("r"))}
        interesting = False
        if node.get("ht"):
            record["height"] = float(node.get("ht"))
            interesting = True
        if _flag(node.get("customHeight")):
            record["custom_height"] = True
            interesting = True
        if _flag(node.get("hidden")):
            record["hidden"] = True
            interesting = True
        if node.get("outlineLevel") and int(node.get("outlineLevel")) != 0:
            record["outline_level"] = int(node.get("outlineLevel"))
            interesting = True
        if _flag(node.get("collapsed")):
            record["collapsed"] = True
            interesting = True
        if node.get("s") and _flag(node.get("customFormat")):
            record["style"] = int(node.get("s"))
            interesting = True
        if interesting:
            out.append(record)
    return out


def _freeze_of(sheet_root) -> object:
    view = sheet_root.find("%s/%s" % (_q("sheetViews"), _q("sheetView")))
    if view is None:
        return None
    pane = view.find(_q("pane"))
    if pane is None:
        return None
    record = {
        "top_left_cell": normalise_address(pane.get("topLeftCell") or ""),
        "x_split": int(float(pane.get("xSplit") or 0)),
        "y_split": int(float(pane.get("ySplit") or 0)),
        "active_pane": pane.get("activePane") or None,
        "state": pane.get("state") or None,
    }
    return record


def _conditional_formats_of(sheet_root, differentials: list) -> list:
    out = []
    for node in sheet_root.findall(_q("conditionalFormatting")):
        rules = []
        for rule in node.findall(_q("cfRule")):
            index = rule.get("dxfId")
            record = {
                "type": rule.get("type"),
                "operator": rule.get("operator") or None,
                "priority": int(rule.get("priority")) if rule.get("priority") else None,
                "stop_if_true": _flag(rule.get("stopIfTrue")),
                "formulas": [(child.text or "") for child in rule.findall(_q("formula"))],
                "dxf": (
                    differentials[int(index)]
                    if index is not None and int(index) < len(differentials)
                    else None
                ),
            }
            if rule.get("text"):
                record["text"] = rule.get("text")
            rules.append(record)
        out.append({"sqref": normalise_sqref(node.get("sqref")), "rules": rules})
    return out


def _data_validations_of(sheet_root) -> list:
    out = []
    container = sheet_root.find(_q("dataValidations"))
    for node in (container if container is not None else []):
        record = {
            "sqref": normalise_sqref(node.get("sqref")),
            "type": node.get("type") or "none",
            "operator": node.get("operator") or None,
            "allow_blank": _flag(node.get("allowBlank")),
            "show_error_message": _flag(node.get("showErrorMessage")),
            "show_input_message": _flag(node.get("showInputMessage")),
            "formula1": None,
            "formula2": None,
        }
        for tag, key in (("formula1", "formula1"), ("formula2", "formula2")):
            child = node.find(_q(tag))
            if child is not None:
                record[key] = child.text or ""
        for attr, key in (
            ("promptTitle", "prompt_title"), ("prompt", "prompt"),
            ("errorTitle", "error_title"), ("error", "error"),
        ):
            if node.get(attr):
                record[key] = node.get(attr)
        out.append(record)
    return out


def _merges_of(sheet_root) -> list:
    container = sheet_root.find(_q("mergeCells"))
    if container is None:
        return []
    return sorted(normalise_sqref(node.get("ref")) for node in container.findall(_q("mergeCell")))


def _outline_of(sheet_root) -> dict:
    outline = sheet_root.find("%s/%s" % (_q("sheetPr"), _q("outlinePr")))
    fmt = sheet_root.find(_q("sheetFormatPr"))
    record = {
        "summary_below": _flag(outline.get("summaryBelow"), True) if outline is not None else True,
        "summary_right": _flag(outline.get("summaryRight"), True) if outline is not None else True,
        "outline_level_row": int(fmt.get("outlineLevelRow")) if fmt is not None and fmt.get("outlineLevelRow") else 0,
        "outline_level_col": int(fmt.get("outlineLevelCol")) if fmt is not None and fmt.get("outlineLevelCol") else 0,
    }
    return record


def _page_setup_of(sheet_root) -> object:
    node = sheet_root.find(_q("pageSetup"))
    if node is None:
        return None
    record = {}
    for attr, key, cast in (
        ("orientation", "orientation", str), ("scale", "scale", int),
        ("fitToWidth", "fit_to_width", int), ("fitToHeight", "fit_to_height", int),
        ("paperSize", "paper_size", int),
    ):
        if node.get(attr):
            record[key] = cast(node.get(attr))
    return record or None


def read_comments(package: Package, sheet_part: str) -> tuple[list, object, list]:
    """Legacy comment text keyed by cell, plus the threaded-comment author.

    The emitted package writes the visible text into `xl/comments/*.xml` and the
    same text again as a threaded comment whose author resolves through
    `xl/persons/person.xml`.  Both are read: the legacy part is what a
    spreadsheet reader shows, the persons part is where the author name lives.
    """
    comments: list = []
    author = None
    surfaces: list = []
    unmapped: list = []
    relations = package.rels(sheet_part)
    for _, (kind, target, mode) in sorted(relations.items()):
        if kind != REL_COMMENTS or mode == "External":
            continue
        part = package.resolve(sheet_part, target)
        if part not in package.names:
            continue
        root = package.xml(part)
        part_unmapped, part_surfaces = _unmapped_sweep(root, part, "comments")
        unmapped.extend(part_unmapped)
        surfaces.extend(part_surfaces)
        comment_list = root.find(_q("commentList"))
        for node in (comment_list if comment_list is not None else []):
            text_node = node.find(_q("text"))
            comments.append({
                "cell": normalise_address(node.get("ref") or ""),
                "text": _rich_text(text_node),
            })
        authors = root.find(_q("authors"))
        for node in (authors if authors is not None else []):
            value = (node.text or "").strip()
            if value.startswith("tc="):
                surfaces.append({
                    "surface": "threaded_comment_author_marker",
                    "part": part,
                    "why": "author slot holds a threaded-comment id; the display name lives in xl/persons",
                })
                break
    person_part = "xl/persons/person.xml"
    if person_part in package.names:
        root = package.xml(person_part)
        people = [node.get("displayName") for node in root.findall(_q("person", TC_NS))]
        if people:
            author = people[0]
        surfaces.append({
            "surface": "threaded_comment_person_ids",
            "part": person_part,
            "why": "person GUIDs and comment timestamps have no plan field",
        })
    return comments, author, surfaces, unmapped


def read_images(package: Package, sheet_part: str) -> tuple[list, list]:
    """Drawing anchors and media hashes; the plan's source `path` is not in the file."""
    images: list = []
    surfaces: list = []
    for _, (kind, target, mode) in sorted(package.rels(sheet_part).items()):
        if kind != REL_DRAWING or mode == "External":
            continue
        part = package.resolve(sheet_part, target)
        if part not in package.names:
            continue
        drawing = package.xml(part)
        media_rels = package.rels(part)
        for anchor in list(drawing):
            frm = anchor.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}from")
            column = row = None
            if frm is not None:
                column_node = frm.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}col")
                row_node = frm.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}row")
                if column_node is not None and row_node is not None:
                    column = int(column_node.text) + 1
                    row = int(row_node.text) + 1
            embed = None
            for node in anchor.iter():
                rid = node.get("{%s}embed" % DOC_REL_NS)
                if rid:
                    embed = rid
                    break
            digest = None
            if embed and embed in media_rels:
                media = package.resolve(part, media_rels[embed][1])
                if media in package.names:
                    digest = hashlib.sha256(package.zip.read(media)).hexdigest()
            images.append({
                "anchor": ("%s%d" % (column_letters(column), row)) if column and row else None,
                "sha256": digest,
            })
        surfaces.append({
            "surface": "drawing_geometry",
            "part": part,
            "why": "EMU extents and anchor mode are reconstructible but the plan records only anchor cell and pixel size",
        })
    return images, surfaces


def _unmapped_sweep(root, part: str, kind: str) -> tuple[list, list]:
    """Every element and attribute is mapped, named-unrecordable, or a finding."""
    table = SWEEP_TABLES[kind]
    mapped = table["mapped"]
    unrecordable = table["unrecordable"]
    attributes = table["attributes"]
    unrecordable_attributes = table["unrecordable_attributes"]
    findings: list = []
    surfaces: list = []
    seen_finding: set = set()
    seen_surface: set = set()
    for node in root.iter():
        tag = _local(node.tag)
        if tag in unrecordable:
            if tag not in seen_surface:
                seen_surface.add(tag)
                surfaces.append({"surface": "element:%s" % tag, "part": part,
                                 "why": unrecordable[tag]})
            continue
        if tag not in mapped:
            if ("element", tag) not in seen_finding:
                seen_finding.add(("element", tag))
                findings.append({
                    "code": "RECON_UNMAPPED_XML_ELEMENT",
                    "part": part, "locus": tag,
                    "detail": "element present in the emitted package that this reconstructor neither maps to a plan channel nor classifies as unrecordable",
                })
            continue
        allowed = attributes.get(tag)
        if allowed is None:
            continue
        for attribute in node.keys():
            name = attribute if not attribute.startswith("{") else "r:" + _local(attribute)
            if name in allowed:
                continue
            key = (tag, name)
            if key in unrecordable_attributes:
                marker = "attribute:%s@%s" % key
                if marker not in seen_surface:
                    seen_surface.add(marker)
                    surfaces.append({"surface": marker, "part": part,
                                     "why": unrecordable_attributes[key]})
                continue
            if ("attribute", tag, name) in seen_finding:
                continue
            seen_finding.add(("attribute", tag, name))
            findings.append({
                "code": "RECON_UNMAPPED_XML_ATTRIBUTE",
                "part": part, "locus": "%s@%s" % (tag, name),
                "detail": "attribute present in the emitted package that this reconstructor neither reads nor classifies as unrecordable",
            })
    return findings, surfaces


def reconstruct_plan(xlsx_path: Path) -> dict:
    """A plan-shaped description of the workbook, built from the .xlsx alone."""
    package = Package(xlsx_path)
    try:
        root_rels = package.rels("")
        workbook_part = "xl/workbook.xml"
        for _, (kind, target, _mode) in sorted(root_rels.items()):
            if kind == REL_OFFICE_DOCUMENT:
                workbook_part = target.lstrip("/")
        workbook_root = package.xml(workbook_part)
        workbook_rels = package.rels(workbook_part)

        styles_part = shared_part = None
        sheet_parts: dict = {}
        for rid, (kind, target, mode) in workbook_rels.items():
            if mode == "External":
                continue
            resolved = package.resolve(workbook_part, target)
            if kind == REL_STYLES:
                styles_part = resolved
            elif kind == REL_SHARED_STRINGS:
                shared_part = resolved
            elif kind == REL_WORKSHEET:
                sheet_parts[rid] = resolved

        styles, differentials = read_style_tables(package, styles_part)
        shared = read_shared_strings(package, shared_part)

        unmapped, surfaces = _unmapped_sweep(workbook_root, workbook_part, "workbook")
        if styles_part and styles_part in package.names:
            style_unmapped, style_surfaces = _unmapped_sweep(
                package.xml(styles_part), styles_part, "styles")
            unmapped.extend(style_unmapped)
            surfaces.extend(style_surfaces)

        calc = workbook_root.find(_q("calcPr"))
        calc_properties = None
        if calc is not None:
            calc_properties = {
                "calc_id": calc.get("calcId"),
                "calc_mode": calc.get("calcMode") or "auto",
                "full_calc_on_load": _flag(calc.get("fullCalcOnLoad")),
                "force_full_calc": _flag(calc.get("forceFullCalc")),
                "iterate": _flag(calc.get("iterate")),
                "iterate_count": int(calc.get("iterateCount")) if calc.get("iterateCount") else None,
                "iterate_delta": float(calc.get("iterateDelta")) if calc.get("iterateDelta") else None,
            }

        defined_names = []
        container = workbook_root.find(_q("definedNames"))
        for node in (container if container is not None else []):
            defined_names.append({"name": node.get("name"), "refers_to": (node.text or "").strip()})

        sheets = []
        sheet_list = workbook_root.find(_q("sheets"))
        for entry in (sheet_list if sheet_list is not None else []):
            rid = entry.get("{%s}id" % DOC_REL_NS)
            part = sheet_parts.get(rid)
            if part is None or part not in package.names:
                continue
            sheet_root = package.xml(part)
            sheet_unmapped, sheet_surfaces = _unmapped_sweep(sheet_root, part, "worksheet")
            unmapped.extend(sheet_unmapped)
            surfaces.extend(sheet_surfaces)
            view = sheet_root.find("%s/%s" % (_q("sheetViews"), _q("sheetView")))
            fmt = sheet_root.find(_q("sheetFormatPr"))
            margins = sheet_root.find(_q("pageMargins"))
            comments, author, comment_surfaces, comment_unmapped = read_comments(package, part)
            images, image_surfaces = read_images(package, part)
            surfaces.extend(comment_surfaces)
            surfaces.extend(image_surfaces)
            unmapped.extend(comment_unmapped)
            sheets.append({
                "name": entry.get("name"),
                "show_gridlines": _flag(view.get("showGridLines"), True) if view is not None else True,
                "default_row_height": (
                    float(fmt.get("defaultRowHeight"))
                    if fmt is not None and fmt.get("defaultRowHeight") else None
                ),
                "freeze_pane": _freeze_of(sheet_root),
                "outline": _outline_of(sheet_root),
                "page_margins": (
                    {key: float(margins.get(key)) for key in
                     ("left", "right", "top", "bottom", "header", "footer")
                     if margins.get(key) is not None}
                    if margins is not None else None
                ),
                "page_setup": _page_setup_of(sheet_root),
                "columns": _columns_of(sheet_root),
                "rows": _rows_of(sheet_root),
                "cells": _cells_of(sheet_root, styles, shared),
                "merges": _merges_of(sheet_root),
                "conditional_formats": _conditional_formats_of(sheet_root, differentials),
                "data_validations": _data_validations_of(sheet_root),
                "comments": comments,
                "comment_author": author,
                "images": images,
            })

        default_font = None
        if styles_part and styles_part in package.names:
            fonts = package.xml(styles_part).find(_q("fonts"))
            if fonts is not None and len(list(fonts)):
                first = _font_effective(list(fonts)[0])
                default_font = {"name": first.get("name"), "size": first.get("size")}

        authors = [sheet["comment_author"] for sheet in sheets if sheet["comment_author"]]
        return {
            "reconstructed_from": {
                "xlsx": str(xlsx_path),
                "xlsx_sha256": _sha256(xlsx_path),
                "parts": sorted(package.names),
            },
            "workbook": {
                "default_font": default_font,
                "calc_properties": calc_properties,
                "comment_author": authors[0] if authors else None,
                "defined_names": defined_names,
                "sheets": sheets,
            },
            "unmapped": unmapped,
            "unrecordable_surfaces": surfaces,
        }
    finally:
        package.close()


# ---------------------------------------------------------------------------
# The recorded plan, normalised into the SAME shape (indices dereferenced)
# ---------------------------------------------------------------------------


def normalise_recorded_plan(plan: dict) -> dict:
    workbook = dict(plan.get("workbook") or {})
    styles = [canonical_style(record) for record in (workbook.get("styles") or [])] or [canonical_style({})]
    differentials = [canonical_differential_style(record) for record in (workbook.get("differential_styles") or [])]

    sheets = []
    for sheet in workbook.get("sheets") or []:
        cells = {}
        for address, record in (sheet.get("cells") or {}).items():
            index = record.get("s")
            out = {"style": styles[index] if isinstance(index, int) and index < len(styles) else canonical_style({})}
            if record.get("f"):
                out["formula"] = record["f"]
            if "t" in record:
                out["type"] = record["t"]
            if "v" in record and record["v"] is not None:
                out["value"] = float(record["v"]) if record.get("t") == "n" else record["v"]
            cells[normalise_address(address)] = out

        conditional_formats = []
        for entry in sheet.get("conditional_formats") or []:
            rules = []
            for rule in entry.get("rules") or []:
                index = rule.get("dxf")
                rules.append({
                    "type": rule.get("type"),
                    "operator": rule.get("operator") or None,
                    "priority": rule.get("priority"),
                    "stop_if_true": bool(rule.get("stop_if_true")),
                    "formulas": list(rule.get("formulas") or []),
                    "dxf": (
                        differentials[index]
                        if isinstance(index, int) and index < len(differentials) else None
                    ),
                })
                if rule.get("text"):
                    rules[-1]["text"] = rule["text"]
            conditional_formats.append({"sqref": normalise_sqref(entry.get("sqref")), "rules": rules})

        validations = []
        for entry in sheet.get("data_validations") or []:
            record = {
                "sqref": normalise_sqref(entry.get("sqref")),
                "type": entry.get("type") or "none",
                "operator": entry.get("operator") or None,
                "allow_blank": bool(entry.get("allow_blank")),
                "show_error_message": bool(entry.get("show_error_message")),
                "show_input_message": bool(entry.get("show_input_message")),
                "formula1": entry.get("formula1"),
                "formula2": entry.get("formula2"),
            }
            for key in ("prompt_title", "prompt", "error_title", "error"):
                if entry.get(key):
                    record[key] = entry[key]
            validations.append(record)

        columns = []
        for entry in sheet.get("columns") or []:
            record = {
                "min": entry.get("min"), "max": entry.get("max"),
                "width": float(entry["width"]) if entry.get("width") is not None else None,
                "hidden": bool(entry.get("hidden")),
                "custom_width": bool(entry.get("custom_width")),
            }
            if entry.get("outline_level"):
                record["outline_level"] = entry["outline_level"]
            if entry.get("style"):
                record["style"] = entry["style"]
            columns.append(record)

        rows = []
        for entry in sheet.get("rows") or []:
            record = {"row": entry.get("row")}
            if entry.get("height") is not None:
                record["height"] = float(entry["height"])
            for key in ("custom_height", "hidden", "collapsed"):
                if entry.get(key):
                    record[key] = True
            for key in ("outline_level", "style"):
                if entry.get(key):
                    record[key] = entry[key]
            rows.append(record)

        freeze = sheet.get("freeze_pane")
        if freeze:
            freeze = {
                "top_left_cell": normalise_address(freeze.get("top_left_cell") or ""),
                "x_split": int(freeze.get("x_split") or 0),
                "y_split": int(freeze.get("y_split") or 0),
                "active_pane": freeze.get("active_pane") or None,
                "state": freeze.get("state") or None,
            }

        outline = dict(sheet.get("outline") or {})
        sheets.append({
            "name": sheet.get("name"),
            "show_gridlines": bool(sheet.get("show_gridlines", True)),
            "default_row_height": (
                float(sheet["default_row_height"])
                if sheet.get("default_row_height") is not None else None
            ),
            "freeze_pane": freeze or None,
            "outline": {
                "summary_below": bool(outline.get("summary_below", True)),
                "summary_right": bool(outline.get("summary_right", True)),
                "outline_level_row": int(outline.get("outline_level_row") or 0),
                "outline_level_col": int(outline.get("outline_level_col") or 0),
            },
            "page_margins": (
                {key: float(value) for key, value in (sheet.get("page_margins") or {}).items()}
                if sheet.get("page_margins") else None
            ),
            "page_setup": sheet.get("page_setup") or None,
            "columns": columns,
            "rows": rows,
            "cells": cells,
            "merges": sorted(normalise_sqref(ref) for ref in (sheet.get("merges") or [])),
            "conditional_formats": conditional_formats,
            "data_validations": validations,
            "comments": [
                {"cell": normalise_address(entry.get("cell") or ""), "text": entry.get("text") or ""}
                for entry in (sheet.get("comments") or [])
            ],
            "images": [
                {"anchor": normalise_address(entry.get("anchor") or ""), "sha256": entry.get("sha256")}
                for entry in (sheet.get("images") or [])
            ],
        })

    calc = dict(workbook.get("calc_properties") or {})
    return {
        "workbook": {
            "default_font": (
                {"name": (workbook.get("default_font") or {}).get("name"),
                 "size": float((workbook.get("default_font") or {}).get("size"))
                 if (workbook.get("default_font") or {}).get("size") is not None else None}
                if workbook.get("default_font") else None
            ),
            "calc_properties": {
                "calc_id": str(calc.get("calc_id")) if calc.get("calc_id") is not None else None,
                "calc_mode": calc.get("calc_mode") or "auto",
                "full_calc_on_load": bool(calc.get("full_calc_on_load")),
                "force_full_calc": bool(calc.get("force_full_calc")),
                "iterate": bool(calc.get("iterate")),
                "iterate_count": calc.get("iterate_count"),
                "iterate_delta": float(calc["iterate_delta"]) if calc.get("iterate_delta") is not None else None,
            } if calc else None,
            "comment_author": workbook.get("comment_author") or None,
            "defined_names": [
                {"name": entry.get("name"), "refers_to": entry.get("refers_to")}
                for sheet in (workbook.get("sheets") or [])
                for entry in (sheet.get("defined_names") or [])
            ],
            "sheets": sheets,
        },
    }


# ---------------------------------------------------------------------------
# openpyxl as the second, third-party reader
# ---------------------------------------------------------------------------


def cross_read_with_openpyxl(xlsx_path: Path, reconstruction: dict) -> list:
    """A disagreement between this file's reader and openpyxl is a finding.

    openpyxl is a third-party OOXML reader with no relationship to this
    repository.  Where the two readers disagree about a formula, a cached value
    or a resolved number format, the reconstruction cannot be trusted and says
    so instead of quietly reporting a clean workbook.
    """
    findings: list = []
    formulas = load_workbook(xlsx_path, data_only=False)
    values = load_workbook(xlsx_path, data_only=True)
    try:
        if formulas.sheetnames != [sheet["name"] for sheet in reconstruction["workbook"]["sheets"]]:
            findings.append({
                "code": "RECON_READER_DISAGREEMENT",
                "locus": "sheet_names",
                "stdlib": [sheet["name"] for sheet in reconstruction["workbook"]["sheets"]],
                "openpyxl": formulas.sheetnames,
            })
            return findings
        for sheet in reconstruction["workbook"]["sheets"]:
            formula_sheet = formulas[sheet["name"]]
            value_sheet = values[sheet["name"]]
            for address, record in sheet["cells"].items():
                other_formula = formula_sheet[address].value
                mine = record.get("formula")
                if mine is not None:
                    expected = "=" + mine
                    if not isinstance(other_formula, str) or other_formula.replace(" ", "") != expected.replace(" ", ""):
                        findings.append({
                            "code": "RECON_READER_DISAGREEMENT",
                            "sheet": sheet["name"], "locus": address, "channel": "formula",
                            "stdlib": expected, "openpyxl": other_formula,
                        })
                        continue
                elif isinstance(other_formula, str) and other_formula.startswith("="):
                    findings.append({
                        "code": "RECON_READER_DISAGREEMENT",
                        "sheet": sheet["name"], "locus": address, "channel": "formula",
                        "stdlib": None, "openpyxl": other_formula,
                    })
                    continue
                other_value = value_sheet[address].value
                mine_value = record.get("value")
                if record.get("type") == "n" and isinstance(other_value, (int, float)):
                    if not _numbers_agree(float(mine_value), float(other_value)):
                        findings.append({
                            "code": "RECON_READER_DISAGREEMENT",
                            "sheet": sheet["name"], "locus": address, "channel": "cached_value",
                            "stdlib": mine_value, "openpyxl": other_value,
                        })
                elif record.get("type") == "s" and other_value is not None:
                    if str(mine_value) != str(other_value):
                        findings.append({
                            "code": "RECON_READER_DISAGREEMENT",
                            "sheet": sheet["name"], "locus": address, "channel": "cached_value",
                            "stdlib": mine_value, "openpyxl": other_value,
                        })
                other_format = formula_sheet[address].number_format
                if other_format != record["style"]["number_format"]:
                    findings.append({
                        "code": "RECON_READER_DISAGREEMENT",
                        "sheet": sheet["name"], "locus": address, "channel": "number_format",
                        "stdlib": record["style"]["number_format"], "openpyxl": other_format,
                    })
    finally:
        formulas.close()
        values.close()
    return findings


def _numbers_agree(left: float, right: float) -> bool:
    if left == right:
        return True
    difference = abs(left - right)
    return difference <= max(VALUE_ABSOLUTE_TOLERANCE, VALUE_RELATIVE_TOLERANCE * max(abs(left), abs(right)))


# ---------------------------------------------------------------------------
# The comparison
# ---------------------------------------------------------------------------

DECLARED_GAPS = [
    {
        "aspect": "images[].path",
        "reason": "the plan records the filesystem path of a source image; the .xlsx carries only the media bytes, so the path is not reconstructible. The anchor cell and the media sha256 ARE reconstructed and compared.",
    },
    {
        "aspect": "plan_version / case_id / generator",
        "reason": "provenance of the plan document, not a property of the workbook; nothing in the .xlsx records it.",
    },
    {
        "aspect": "cached value TRUTH",
        "reason": "this oracle proves the file's cached values agree with the recorded plan's; whether either is the correct arithmetic answer is a recalculation question and belongs to a recalculation oracle.",
    },
    {
        "aspect": "apply* style flags",
        "reason": "applyFont/applyFill/applyBorder/applyNumberFormat gate inheritance from cellStyleXfs. The emitted package carries a single default parent, so the xf's own values are the effective values; a package with real named-style parents would need the flags honoured.",
    },
    {
        "aspect": "shared and array formula expansion",
        "reason": "a `<f t=\"shared\">` master and its `si` followers are reconstructed as written (kind, ref and si are recorded and compared) but not expanded into per-cell formula text; the emitted package writes every formula in full, so this is unexercised rather than wrong.",
    },
]


class _Report:
    def __init__(self) -> None:
        self.findings: list = []
        self.counts: dict = {}

    def add(self, code: str, **detail) -> None:
        self.counts[code] = self.counts.get(code, 0) + 1
        if self.counts[code] <= FINDINGS_PER_CODE:
            self.findings.append(dict(code=code, **detail))


def compare(reconstruction: dict, recorded: dict) -> dict:
    """Typed findings for every disagreement between the file and the plan."""
    report = _Report()
    file_workbook = reconstruction["workbook"]
    plan_workbook = recorded["workbook"]

    for entry in reconstruction.get("unmapped") or []:
        report.add(entry["code"], part=entry.get("part"), locus=entry.get("locus"),
                   detail=entry.get("detail"))

    if file_workbook["default_font"] != plan_workbook["default_font"]:
        report.add("RECON_DEFAULT_FONT_MISMATCH",
                   file=file_workbook["default_font"], plan=plan_workbook["default_font"])

    file_calc = file_workbook["calc_properties"] or {}
    plan_calc = plan_workbook["calc_properties"] or {}
    for key in sorted(set(file_calc) | set(plan_calc)):
        if file_calc.get(key) != plan_calc.get(key):
            report.add("RECON_CALC_PROPERTY_MISMATCH", locus=key,
                       file=file_calc.get(key), plan=plan_calc.get(key))

    if file_workbook["comment_author"] != plan_workbook["comment_author"]:
        report.add("RECON_COMMENT_AUTHOR_MISMATCH",
                   file=file_workbook["comment_author"], plan=plan_workbook["comment_author"])

    file_names = {entry["name"]: entry["refers_to"] for entry in file_workbook["defined_names"]}
    plan_names = {entry["name"]: entry["refers_to"] for entry in plan_workbook["defined_names"]}
    for name in sorted(set(file_names) - set(plan_names)):
        report.add("RECON_DEFINED_NAME_ABSENT_FROM_PLAN", locus=name, file=file_names[name])
    for name in sorted(set(plan_names) - set(file_names)):
        report.add("RECON_DEFINED_NAME_ABSENT_FROM_FILE", locus=name, plan=plan_names[name])
    for name in sorted(set(plan_names) & set(file_names)):
        if file_names[name] != plan_names[name]:
            report.add("RECON_DEFINED_NAME_MISMATCH", locus=name,
                       file=file_names[name], plan=plan_names[name])

    file_sheets = {sheet["name"]: sheet for sheet in file_workbook["sheets"]}
    plan_sheets = {sheet["name"]: sheet for sheet in plan_workbook["sheets"]}
    for name in sorted(set(file_sheets) - set(plan_sheets)):
        report.add("RECON_SHEET_ABSENT_FROM_PLAN", sheet=name)
    for name in sorted(set(plan_sheets) - set(file_sheets)):
        report.add("RECON_SHEET_ABSENT_FROM_FILE", sheet=name)
    file_order = [sheet["name"] for sheet in file_workbook["sheets"]]
    plan_order = [sheet["name"] for sheet in plan_workbook["sheets"]]
    if [name for name in file_order if name in plan_sheets] != [name for name in plan_order if name in file_sheets]:
        report.add("RECON_SHEET_ORDER_MISMATCH", file=file_order, plan=plan_order)

    cells_compared = 0
    for name in sorted(set(file_sheets) & set(plan_sheets)):
        cells_compared += _compare_sheet(report, name, file_sheets[name], plan_sheets[name])

    for entry in cross_read_with_openpyxl(
        Path(reconstruction["reconstructed_from"]["xlsx"]), reconstruction
    ):
        report.add(**entry)

    metrics = {
        "sheets_compared": len(set(file_sheets) & set(plan_sheets)),
        "cells_compared": cells_compared,
        "formula_cells": sum(
            1 for sheet in file_workbook["sheets"] for record in sheet["cells"].values()
            if record.get("formula")
        ),
        "merges": sum(len(sheet["merges"]) for sheet in file_workbook["sheets"]),
        "conditional_format_ranges": sum(len(sheet["conditional_formats"]) for sheet in file_workbook["sheets"]),
        "data_validations": sum(len(sheet["data_validations"]) for sheet in file_workbook["sheets"]),
        "comments": sum(len(sheet["comments"]) for sheet in file_workbook["sheets"]),
        "columns": sum(len(sheet["columns"]) for sheet in file_workbook["sheets"]),
        "rows": sum(len(sheet["rows"]) for sheet in file_workbook["sheets"]),
        "images": sum(len(sheet["images"]) for sheet in file_workbook["sheets"]),
        "findings_total": sum(report.counts.values()),
    }
    return {
        "status": "PASS" if not report.counts else "FAIL",
        "xlsx_sha256": reconstruction["reconstructed_from"]["xlsx_sha256"],
        "value_tolerance": {
            "relative": VALUE_RELATIVE_TOLERANCE,
            "absolute": VALUE_ABSOLUTE_TOLERANCE,
        },
        "finding_counts": dict(sorted(report.counts.items())),
        "findings": report.findings,
        "findings_truncated_per_code_at": FINDINGS_PER_CODE,
        "metrics": metrics,
        "reconstructed_aspects": sorted(RECONSTRUCTED_ASPECTS),
        "declared_gaps": DECLARED_GAPS,
        "unrecordable_surfaces": reconstruction.get("unrecordable_surfaces") or [],
    }


RECONSTRUCTED_ASPECTS = {
    "sheet names and order",
    "cell formulas",
    "cell cached values and value types",
    "cell number formats",
    "cell fonts (name, size, colour, bold, italic, underline, strike)",
    "cell fills",
    "cell borders (per edge, style and colour)",
    "cell alignment (horizontal, vertical, indent, wrap, shrink)",
    "merged ranges",
    "column widths, hidden flags, custom-width flags, outline levels",
    "row heights, hidden flags, custom-height flags, outline levels",
    "freeze panes (split, top-left cell, active pane, state)",
    "sheet gridline visibility",
    "default row height",
    "outline properties (summary below/right, outline levels)",
    "page margins",
    "page setup (orientation, scale, fit-to, paper size)",
    "data validations (type, operator, formulas, messages)",
    "conditional formats (sqref, rule type, operator, priority, formulas, resolved dxf)",
    "comments (cell and text) and the comment author",
    "workbook calculation properties (iterative calculation settings)",
    "defined names",
    "drawing/image anchors and media hashes",
    "default font",
}


def _compare_sheet(report: _Report, name: str, file_sheet: dict, plan_sheet: dict) -> int:
    for key, code in (
        ("show_gridlines", "RECON_SHOW_GRIDLINES_MISMATCH"),
        ("default_row_height", "RECON_DEFAULT_ROW_HEIGHT_MISMATCH"),
        ("freeze_pane", "RECON_FREEZE_PANE_MISMATCH"),
        ("outline", "RECON_OUTLINE_MISMATCH"),
        ("page_margins", "RECON_PAGE_MARGINS_MISMATCH"),
        ("page_setup", "RECON_PAGE_SETUP_MISMATCH"),
    ):
        if file_sheet.get(key) != plan_sheet.get(key):
            report.add(code, sheet=name, file=file_sheet.get(key), plan=plan_sheet.get(key))

    file_merges = set(file_sheet["merges"])
    plan_merges = set(plan_sheet["merges"])
    for ref in sorted(file_merges - plan_merges):
        report.add("RECON_MERGE_ABSENT_FROM_PLAN", sheet=name, locus=ref)
    for ref in sorted(plan_merges - file_merges):
        report.add("RECON_MERGE_ABSENT_FROM_FILE", sheet=name, locus=ref)

    file_columns = {(entry["min"], entry["max"]): entry for entry in file_sheet["columns"]}
    plan_columns = {(entry["min"], entry["max"]): entry for entry in plan_sheet["columns"]}
    for key in sorted(set(file_columns) - set(plan_columns)):
        report.add("RECON_COLUMN_ABSENT_FROM_PLAN", sheet=name, locus=str(key), file=file_columns[key])
    for key in sorted(set(plan_columns) - set(file_columns)):
        report.add("RECON_COLUMN_ABSENT_FROM_FILE", sheet=name, locus=str(key), plan=plan_columns[key])
    for key in sorted(set(file_columns) & set(plan_columns)):
        if file_columns[key] != plan_columns[key]:
            report.add("RECON_COLUMN_MISMATCH", sheet=name, locus=str(key),
                       file=file_columns[key], plan=plan_columns[key])

    file_rows = {entry["row"]: entry for entry in file_sheet["rows"]}
    plan_rows = {entry["row"]: entry for entry in plan_sheet["rows"]}
    for key in sorted(set(file_rows) - set(plan_rows)):
        report.add("RECON_ROW_ABSENT_FROM_PLAN", sheet=name, locus=key, file=file_rows[key])
    for key in sorted(set(plan_rows) - set(file_rows)):
        report.add("RECON_ROW_ABSENT_FROM_FILE", sheet=name, locus=key, plan=plan_rows[key])
    for key in sorted(set(file_rows) & set(plan_rows)):
        if file_rows[key] != plan_rows[key]:
            report.add("RECON_ROW_MISMATCH", sheet=name, locus=key,
                       file=file_rows[key], plan=plan_rows[key])

    _compare_validations(report, name, file_sheet, plan_sheet)
    _compare_conditional_formats(report, name, file_sheet, plan_sheet)
    _compare_comments(report, name, file_sheet, plan_sheet)
    _compare_images(report, name, file_sheet, plan_sheet)
    return _compare_cells(report, name, file_sheet, plan_sheet)


def _compare_cells(report: _Report, name: str, file_sheet: dict, plan_sheet: dict) -> int:
    file_cells = file_sheet["cells"]
    plan_cells = plan_sheet["cells"]
    for address in sorted(set(file_cells) - set(plan_cells)):
        report.add("RECON_CELL_ABSENT_FROM_PLAN", sheet=name, locus=address, file=file_cells[address])
    for address in sorted(set(plan_cells) - set(file_cells)):
        report.add("RECON_CELL_ABSENT_FROM_FILE", sheet=name, locus=address, plan=plan_cells[address])
    shared = sorted(set(file_cells) & set(plan_cells))
    for address in shared:
        file_cell = file_cells[address]
        plan_cell = plan_cells[address]
        if file_cell.get("formula") != plan_cell.get("formula"):
            report.add("RECON_FORMULA_MISMATCH", sheet=name, locus=address,
                       file=file_cell.get("formula"), plan=plan_cell.get("formula"))
        for key, code in (("formula_kind", "RECON_FORMULA_KIND_UNPLANNED"),
                          ("formula_ref", "RECON_FORMULA_REF_UNPLANNED"),
                          ("formula_si", "RECON_FORMULA_SHARED_INDEX_UNPLANNED")):
            if file_cell.get(key) is not None:
                report.add(code, sheet=name, locus=address, file=file_cell[key],
                           detail="the plan has no field for this formula attribute")
        spelling = file_cell.get("type_spelling")
        if spelling is not None:
            wanted = "str" if file_cell.get("formula") is not None else ("inlineStr", "s")
            if (spelling != wanted) if isinstance(wanted, str) else (spelling not in wanted):
                report.add("RECON_STRING_TYPE_SPELLING", sheet=name, locus=address,
                           file=spelling, plan=plan_cell.get("type"),
                           detail="a text result of a formula must be spelled t=\"str\"; a literal must be t=\"inlineStr\" or t=\"s\"")
        if file_cell.get("type") != plan_cell.get("type"):
            report.add("RECON_CACHED_TYPE_MISMATCH", sheet=name, locus=address,
                       file=file_cell.get("type"), plan=plan_cell.get("type"))
        elif "value" in file_cell or "value" in plan_cell:
            if ("value" in file_cell) != ("value" in plan_cell):
                report.add("RECON_CACHED_VALUE_PRESENCE_MISMATCH", sheet=name, locus=address,
                           file=file_cell.get("value"), plan=plan_cell.get("value"))
            elif file_cell.get("type") == "n":
                if not _numbers_agree(float(file_cell["value"]), float(plan_cell["value"])):
                    report.add("RECON_VALUE_MISMATCH", sheet=name, locus=address,
                               file=file_cell["value"], plan=plan_cell["value"])
            elif file_cell.get("value") != plan_cell.get("value"):
                report.add("RECON_VALUE_MISMATCH", sheet=name, locus=address,
                           file=file_cell.get("value"), plan=plan_cell.get("value"))
        file_style = file_cell["style"]
        plan_style = plan_cell["style"]
        if file_style != plan_style:
            for channel, code in (
                ("number_format", "RECON_NUMBER_FORMAT_MISMATCH"),
                ("font", "RECON_FONT_MISMATCH"),
                ("fill", "RECON_FILL_MISMATCH"),
                ("border", "RECON_BORDER_MISMATCH"),
                ("alignment", "RECON_ALIGNMENT_MISMATCH"),
            ):
                if file_style.get(channel) != plan_style.get(channel):
                    report.add(code, sheet=name, locus=address,
                               file=file_style.get(channel), plan=plan_style.get(channel))
    return len(shared)


def _compare_validations(report: _Report, name: str, file_sheet: dict, plan_sheet: dict) -> None:
    file_map = {entry["sqref"]: entry for entry in file_sheet["data_validations"]}
    plan_map = {entry["sqref"]: entry for entry in plan_sheet["data_validations"]}
    for key in sorted(set(file_map) - set(plan_map)):
        report.add("RECON_DATA_VALIDATION_ABSENT_FROM_PLAN", sheet=name, locus=key, file=file_map[key])
    for key in sorted(set(plan_map) - set(file_map)):
        report.add("RECON_DATA_VALIDATION_ABSENT_FROM_FILE", sheet=name, locus=key, plan=plan_map[key])
    for key in sorted(set(file_map) & set(plan_map)):
        if file_map[key] != plan_map[key]:
            report.add("RECON_DATA_VALIDATION_MISMATCH", sheet=name, locus=key,
                       file=file_map[key], plan=plan_map[key])


def _compare_conditional_formats(report: _Report, name: str, file_sheet: dict, plan_sheet: dict) -> None:
    file_map: dict = {}
    for entry in file_sheet["conditional_formats"]:
        file_map.setdefault(entry["sqref"], []).extend(entry["rules"])
    plan_map: dict = {}
    for entry in plan_sheet["conditional_formats"]:
        plan_map.setdefault(entry["sqref"], []).extend(entry["rules"])
    for key in sorted(set(file_map) - set(plan_map)):
        report.add("RECON_CONDITIONAL_FORMAT_ABSENT_FROM_PLAN", sheet=name, locus=key, file=file_map[key])
    for key in sorted(set(plan_map) - set(file_map)):
        report.add("RECON_CONDITIONAL_FORMAT_ABSENT_FROM_FILE", sheet=name, locus=key, plan=plan_map[key])
    for key in sorted(set(file_map) & set(plan_map)):
        file_rules = sorted(file_map[key], key=lambda rule: (rule.get("priority") or 0))
        plan_rules = sorted(plan_map[key], key=lambda rule: (rule.get("priority") or 0))
        if len(file_rules) != len(plan_rules):
            report.add("RECON_CONDITIONAL_FORMAT_RULE_COUNT_MISMATCH", sheet=name, locus=key,
                       file=len(file_rules), plan=len(plan_rules))
            continue
        for file_rule, plan_rule in zip(file_rules, plan_rules):
            if file_rule != plan_rule:
                report.add("RECON_CONDITIONAL_FORMAT_RULE_MISMATCH", sheet=name, locus=key,
                           file=file_rule, plan=plan_rule)


def _compare_comments(report: _Report, name: str, file_sheet: dict, plan_sheet: dict) -> None:
    file_map: dict = {}
    for entry in file_sheet["comments"]:
        file_map.setdefault(entry["cell"], []).append(entry["text"])
    plan_map: dict = {}
    for entry in plan_sheet["comments"]:
        plan_map.setdefault(entry["cell"], []).append(entry["text"])
    for cell in sorted(set(file_map) - set(plan_map)):
        report.add("RECON_COMMENT_ABSENT_FROM_PLAN", sheet=name, locus=cell, file=file_map[cell])
    for cell in sorted(set(plan_map) - set(file_map)):
        report.add("RECON_COMMENT_ABSENT_FROM_FILE", sheet=name, locus=cell, plan=plan_map[cell])
    for cell in sorted(set(file_map) & set(plan_map)):
        if sorted(file_map[cell]) != sorted(plan_map[cell]):
            report.add("RECON_COMMENT_TEXT_MISMATCH", sheet=name, locus=cell,
                       file=file_map[cell], plan=plan_map[cell])


def _compare_images(report: _Report, name: str, file_sheet: dict, plan_sheet: dict) -> None:
    file_map = {(entry.get("anchor"), entry.get("sha256")) for entry in file_sheet["images"]}
    plan_map = {(entry.get("anchor"), entry.get("sha256")) for entry in plan_sheet["images"]}
    for key in sorted(file_map - plan_map, key=repr):
        report.add("RECON_IMAGE_ABSENT_FROM_PLAN", sheet=name, locus=repr(key))
    for key in sorted(plan_map - file_map, key=repr):
        report.add("RECON_IMAGE_ABSENT_FROM_FILE", sheet=name, locus=repr(key))


def verify_workbook(xlsx_path: Path, plan_path: Path) -> dict:
    """Reconstruct from the .xlsx, normalise the recorded plan, compare."""
    reconstruction = reconstruct_plan(Path(xlsx_path))
    raw_plan = json.loads(Path(plan_path).read_text(encoding="utf-8"))
    if not isinstance(raw_plan, dict) or not isinstance(raw_plan.get("workbook"), dict):
        return {
            "status": "FAIL",
            "xlsx_sha256": reconstruction["reconstructed_from"]["xlsx_sha256"],
            "finding_counts": {"RECON_PLAN_NOT_A_PLAN": 1},
            "findings": [{
                "code": "RECON_PLAN_NOT_A_PLAN",
                "detail": "the document supplied as the recorded plan has no workbook; a reconstruction cannot be compared against it",
                "plan": str(plan_path),
            }],
            "metrics": {"cells_compared": 0},
            "declared_gaps": DECLARED_GAPS,
            "unrecordable_surfaces": reconstruction.get("unrecordable_surfaces") or [],
        }
    recorded = normalise_recorded_plan(raw_plan)
    report = compare(reconstruction, recorded)
    report["plan_sha256"] = _sha256(plan_path)
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--xlsx", required=True, type=Path)
    parser.add_argument("--plan", required=True, type=Path)
    parser.add_argument("--out", type=Path)
    parser.add_argument("--reconstruction-out", type=Path)
    args = parser.parse_args()
    report = verify_workbook(args.xlsx, args.plan)
    if args.reconstruction_out:
        args.reconstruction_out.parent.mkdir(parents=True, exist_ok=True)
        args.reconstruction_out.write_text(
            json.dumps(reconstruct_plan(args.xlsx), indent=2, default=str) + "\n",
            encoding="utf-8",
        )
    if args.out:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(json.dumps(report, indent=2, default=str) + "\n", encoding="utf-8")
    sys.stdout.write(json.dumps({
        "status": report["status"],
        "finding_counts": report.get("finding_counts") or {},
        **(report.get("metrics") or {}),
    }, default=str) + "\n")
    return 0 if report["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
