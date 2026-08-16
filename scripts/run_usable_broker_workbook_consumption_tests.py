#!/usr/bin/env python3
"""Prove a designated usable raw broker input is consumed by the delivered XLSX.

This is deliberately a local simulated-semantic component test, not an
installed-host canary.  Its purpose is to make broker usefulness non-vacuous:
the raw canary must deliver a workbook whose Operating Model actually links to
the compact Brokers sheet with one headline authority and no Bxx dependency.
"""
from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent


def json_objects(text: str) -> list[dict[str, Any]]:
    decoder = json.JSONDecoder()
    values: list[dict[str, Any]] = []
    for index, character in enumerate(text):
        if character != "{":
            continue
        try:
            value, _end = decoder.raw_decode(text[index:])
        except Exception:
            continue
        if isinstance(value, dict):
            values.append(value)
    return values


def formula(value: Any) -> str | None:
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def normalise(value: Any) -> str:
    return " ".join(str(value or "").lower().replace("&", " and ").split())


def row_label(sheet, row: int) -> str:
    pieces = [
        str(sheet.cell(row=row, column=column).value or "").strip()
        for column in range(1, min(sheet.max_column, 6) + 1)
    ]
    return " ".join(piece for piece in pieces if piece)


def broker_links(sheet, row: int) -> list[str]:
    """Return broker formulas only from standalone forecast columns J:L."""
    links = []
    for column in range(10, 13):
        value = formula(sheet.cell(row=row, column=column).value)
        if value and "brokers" in value.lower():
            links.append(value)
    return links


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--fixture", required=True)
    parser.add_argument("--python", required=True)
    parser.add_argument("--soffice", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    output = Path(args.out).resolve()
    output.mkdir(parents=True, exist_ok=True)
    command = [
        sys.executable if Path(sys.executable).resolve() == Path(args.python).resolve() else "node",
    ]
    # The shipping canary is Node.  The explicit Python and soffice arguments
    # are forwarded so every descendant uses the same selected toolchain.
    command = [
        "node",
        str(HERE / "run_raw_input_local_semantic_canary.mjs"),
        str(Path(args.fixture).resolve()),
        str(Path(args.python).resolve()),
        str(Path(args.soffice).resolve()),
        "--broker-state",
        "usable",
        "--dcs-balance-basis",
        "native_principal",
    ]
    environment = {
        **os.environ,
        "EXCEL_INFLOW_TEST_PYTHON": str(Path(args.python).resolve()),
    }
    completed = subprocess.run(
        command,
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=False,
        timeout=3600,
        env=environment,
    )
    (output / "canary.stdout.txt").write_text(completed.stdout, "utf-8")
    (output / "canary.stderr.txt").write_text(completed.stderr, "utf-8")
    if completed.returncode != 0:
        raise AssertionError(
            f"Usable local semantic canary failed rc={completed.returncode}: "
            f"{completed.stderr[-4000:]}"
        )
    receipts = [item for item in json_objects(completed.stdout) if item.get("status") == "PASS"]
    if not receipts:
        raise AssertionError("Usable canary emitted no PASS receipt.")
    receipt = receipts[-1]
    assert receipt.get("broker_state") == "usable", receipt
    assert receipt.get("installed_host_semantics_exercised") is False, receipt
    assert receipt.get("semantic_response_boundary") == "simulated_local_fixture", receipt
    assert receipt.get("preauthored_broker_crosswalk") is True, receipt
    assert int(receipt.get("runtime_broker_selected_value_count") or 0) > 0, receipt
    assert receipt.get("selected_broker_case") != "Forecast Waterfall", receipt

    workbook = Path(str(receipt.get("workbook") or "")).resolve()
    assert workbook.is_file() and workbook.stat().st_size > 0, workbook
    model = load_workbook(workbook, data_only=False, read_only=False)
    assert "Operating Model" in model.sheetnames
    assert "Brokers" in model.sheetnames
    operating = model["Operating Model"]
    brokers = model["Brokers"]

    operating_links: list[tuple[str, str]] = []
    direct_bxx_links: list[tuple[str, str]] = []
    for row in operating.iter_rows():
        for cell in row:
            value = formula(cell.value)
            if not value:
                continue
            lowered = value.lower()
            if "brokers" in lowered:
                operating_links.append((cell.coordinate, value))
            if any(f"b{index:02d}" in lowered for index in range(1, 11)):
                direct_bxx_links.append((cell.coordinate, value))
    assert operating_links, "Operating Model contains no links to Brokers."
    assert not direct_bxx_links, direct_bxx_links

    screenshot_formula_cells: list[str] = []
    screenshot_sheets = [name for name in model.sheetnames if name.startswith("B") and name[1:].isdigit()]
    for name in screenshot_sheets:
        sheet = model[name]
        for row in sheet.iter_rows():
            for cell in row:
                if formula(cell.value):
                    screenshot_formula_cells.append(f"{name}!{cell.coordinate}")
    assert not screenshot_formula_cells, screenshot_formula_cells

    headline_rows: list[int] = []
    da_rows: list[int] = []
    for row in range(1, operating.max_row + 1):
        label = normalise(row_label(operating, row))
        if label in {"ebit", "adjusted ebitda", "ebitda", "ebitda proxy"}:
            headline_rows.append(row)
        if (
            "depreciation and amortisation" in label
            or label in {"d and a", "d&a", "depreciation amortisation"}
        ):
            da_rows.append(row)
    assert len(headline_rows) >= 2, headline_rows
    headline_links = [
        (row, value)
        for row in headline_rows
        for value in broker_links(operating, row)
    ]
    # One selected headline over three forecast periods.  The other headline is
    # formula-derived through D&A and may retain broker values only as memo data
    # on the Brokers sheet.
    assert len(headline_links) == 3, headline_links
    assert da_rows, "No visible D&A bridge row found."
    da_links = [(row, value) for row in da_rows for value in broker_links(operating, row)]
    assert len(da_links) >= 3, da_links

    broker_text = {
        normalise(cell.value)
        for row in brokers.iter_rows()
        for cell in row
        if cell.value not in {None, ""}
    }
    required_groups = {
        "revenue": ("revenue",),
        "d_and_a": ("depreciation and amortisation", "d and a", "d&a"),
        "tax": ("effective tax rate",),
        "capex": ("capital expenditure", "capex"),
        "working_capital": ("change in working capital", "working capital"),
        "dividends": ("dividends paid", "dividends"),
    }
    missing = []
    for group, aliases in required_groups.items():
        if not any(any(alias in value for alias in aliases) for value in broker_text):
            missing.append(group)
    if missing:
        raise AssertionError(f"Brokers sheet lacks required consumed metric groups: {missing}")
    assert any(
        any(alias in value for alias in ("ebit", "adjusted ebitda"))
        for value in broker_text
    ), "Brokers sheet lacks a headline anchor."

    report = {
        "schema_version": "usable-broker-workbook-consumption/1.0",
        "status": "PASS",
        "semantic_boundary": "simulated_local_fixture",
        "installed_host_claim": False,
        "workbook": str(workbook),
        "selected_value_count": receipt.get("runtime_broker_selected_value_count"),
        "operating_model_broker_link_count": len(operating_links),
        "headline_link_count": len(headline_links),
        "d_and_a_link_count": len(da_links),
        "screenshot_sheet_count": len(screenshot_sheets),
        "direct_bxx_link_count": len(direct_bxx_links),
        "screenshot_formula_count": len(screenshot_formula_cells),
        "checks": 18,
    }
    (output / "usable-broker-workbook-consumption.json").write_text(
        json.dumps(report, indent=2, sort_keys=True) + "\n",
        "utf-8",
    )
    print(json.dumps(report, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
