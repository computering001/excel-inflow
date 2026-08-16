#!/usr/bin/env python3
"""Build and inspect a funded-acquisition workbook from portable custody."""
from __future__ import annotations
import argparse, json, os, shutil, subprocess, sys, tempfile, time
from pathlib import Path
from typing import Any
from openpyxl import load_workbook

HERE=Path(__file__).resolve().parent
ROOT=HERE.parent

def run(command:list[str],cwd:Path=ROOT,timeout:int=1800)->subprocess.CompletedProcess[str]:
    return subprocess.run(command,cwd=cwd,text=True,capture_output=True,timeout=timeout,check=False,env={**os.environ,"PYTHONDONTWRITEBYTECODE":"1","EXCEL_INFLOW_TEST_PYTHON":sys.executable})

def build(case:Path,out:Path,soffice:str)->dict[str,Any]:
    attempts=[
      ["node",str(HERE/"build_dynamic_model.mjs"),str(case),"--out",str(out),"--python",sys.executable,"--soffice",soffice],
      ["node",str(HERE/"build_dynamic_model.mjs"),str(case),"--out",str(out)],
      ["node",str(HERE/"build_dynamic_model.mjs"),"--case",str(case),"--out",str(out)],
      ["node",str(HERE/"build_dynamic_model.mjs"),str(case),str(out)],
    ]
    receipts=[]
    for index,command in enumerate(attempts,1):
        if out.exists():out.unlink()
        completed=run(command)
        receipts.append({"attempt":index,"command":command,"returncode":completed.returncode,"stdout":completed.stdout[-4000:],"stderr":completed.stderr[-8000:]})
        if completed.returncode==0 and out.exists() and out.stat().st_size>0:
            return {"attempt":index,"receipts":receipts}
    raise AssertionError("No supported build_dynamic_model CLI produced a workbook: "+json.dumps(receipts[-1],indent=2))

def norm(value:Any)->str:
    return " ".join(str(value or "").lower().replace("_"," ").split())

def find_rows(ws,terms:set[str])->list[int]:
    rows=[]
    for row in ws.iter_rows():
        for cell in row:
            if norm(cell.value) in terms:
                rows.append(cell.row);break
    return sorted(set(rows))

def formula_cells(ws,row:int)->list[Any]:
    return [cell for cell in ws[row] if isinstance(cell.value,str) and cell.value.startswith("=")]

def numeric_cells(ws,row:int)->list[float]:
    result=[]
    for cell in ws[row]:
        if isinstance(cell.value,(int,float)) and not isinstance(cell.value,bool):result.append(float(cell.value))
    return result

def main()->int:
    parser=argparse.ArgumentParser()
    parser.add_argument("case")
    parser.add_argument("--soffice",required=True)
    parser.add_argument("--out",required=True)
    args=parser.parse_args()
    case=Path(args.case).resolve();outdir=Path(args.out).resolve();outdir.mkdir(parents=True,exist_ok=True)
    workbook=outdir/"acquisition-funded.xlsx"
    build_receipt=build(case,workbook,args.soffice)

    formulas=load_workbook(workbook,data_only=False,read_only=False)
    values=load_workbook(workbook,data_only=True,read_only=False)
    assert "Operating Model" in formulas.sheetnames
    ws=formulas["Operating Model"]; vws=values["Operating Model"]
    consideration_rows=find_rows(ws,{"direct acquisition cash flow","acquisition consideration","purchase consideration"})
    debt_rows=find_rows(ws,{"acquisition debt proceeds","acquisition financing proceeds"})
    assert consideration_rows,"No acquisition-consideration row was emitted"
    assert debt_rows,"No acquisition-debt-proceeds row was emitted"
    consideration_formulas=[cell.value for row in consideration_rows for cell in formula_cells(ws,row)]
    debt_formulas=[cell.value for row in debt_rows for cell in formula_cells(ws,row)]
    assert consideration_formulas,"Acquisition consideration has no formula cells"
    assert debt_formulas,"Acquisition debt proceeds has no formula cells"
    assert all("IF($P$4=0,0,0)" not in formula.replace(" ","") for formula in consideration_formulas+debt_formulas)
    assert any("$P$5" in formula and "-$P$5" in formula.replace(" ","") for formula in consideration_formulas),consideration_formulas
    assert any("$P$8" in formula for formula in debt_formulas),debt_formulas
    assert any("$P$10" in formula for formula in consideration_formulas+debt_formulas)

    consideration_values=[value for row in consideration_rows for value in numeric_cells(vws,row)]
    debt_values=[value for row in debt_rows for value in numeric_cells(vws,row)]
    assert any(abs(value+1000)<1e-6 for value in consideration_values),consideration_values
    assert any(abs(value-400)<1e-6 for value in debt_values),debt_values

    # Formula and cache integrity must still pass after the funded transaction.
    cache_report=outdir/"cache-parity.json"
    completed=run(["node",str(HERE/"validate_cache_parity.mjs"),str(workbook),"--json",str(cache_report)])
    assert completed.returncode==0,completed.stderr
    if cache_report.exists():
        report=json.loads(cache_report.read_text("utf8"))
        assert len(report.get("mismatches",[]))==0
        assert len(report.get("circular_mismatches",[]))==0

    report={
      "schema_version":"acquisition-workbook-proof/1.0","status":"PASS",
      "case":str(case),"workbook":str(workbook),"workbook_bytes":workbook.stat().st_size,
      "build":build_receipt,
      "consideration_rows":consideration_rows,"debt_proceeds_rows":debt_rows,
      "consideration_formulas":consideration_formulas,"debt_proceeds_formulas":debt_formulas,
      "consideration_values":consideration_values,"debt_proceeds_values":debt_values,
    }
    (outdir/"acquisition-workbook-proof.json").write_text(json.dumps(report,indent=2)+"\n")
    print(json.dumps({"status":"PASS","workbook":str(workbook),"bytes":workbook.stat().st_size},sort_keys=True))
    return 0
if __name__=="__main__":raise SystemExit(main())
